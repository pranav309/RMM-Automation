import openpyxl
import smtplib

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

from utilities.customLogger import LogGen

logger = LogGen.loggen()


def merge_html_with_css(html_file_path, css_file_path):
    # Read the HTML file
    with open(html_file_path, 'r') as html_file:
        html_content = html_file.read()

    # Read the CSS file
    with open(css_file_path, 'r') as css_file:
        css_content = css_file.read()

    # Inject the CSS styles into the HTML
    merged_content = f'<style>{css_content}</style>{html_content}'

    # Write the merged content to the output file
    with open(html_file_path, 'w') as output_file:
        output_file.write(merged_content)


def sendEmail():
    path = "./TestData/email.xlsx"
    workBook = openpyxl.load_workbook(path)
    sheet = workBook.active
    rows = sheet.max_row
    for r in range(2, rows+1):
        subject = sheet.cell(row=r, column=1).value
        body = sheet.cell(row=r, column=2).value
        fromEmail = sheet.cell(row=r, column=3).value
        toEmails = sheet.cell(row=r, column=4).value
        reportPaths = sheet.cell(row=r, column=5).value
        cssFilePath = sheet.cell(row=r, column=6).value

        res = tuple(map(str, toEmails.split(', ')))
        for toEmail in res:
            # Create a multipart message object
            message = MIMEMultipart()
            message["Subject"] = subject
            message["From"] = fromEmail
            message["To"] = toEmail

            eachPath = tuple(map(str, reportPaths.split(', ')))
            for reportPath in eachPath:
                merge_html_with_css(reportPath, cssFilePath)
                tmp = tuple(map(str, reportPath.split('/')))
                cnt = len(tmp) - 1
                # Attach the HTML report to the email
                with open(reportPath, "rb") as attachment_file:
                    attachment = MIMEApplication(attachment_file.read(), _subtype="html")
                    attachment.add_header("Content-Disposition", "attachment", filename=str(tmp[cnt]))
                    message.attach(attachment)

            # Add the email body as HTML
            message.attach(MIMEText(body, "html"))

            # Send the email
            smtp_server = "smtp.office365.com"  # Replace with your SMTP server address
            smtp_port = 587  # Replace with your SMTP server port
            smtp_username = "pranav.pawar@rackwareinc.com"  # Replace with your SMTP username
            smtp_password = "P8H2ak%#wd"  # Replace with your SMTP password

            try:
                with smtplib.SMTP(smtp_server, smtp_port) as server:
                    server.starttls()
                    server.login(smtp_username, smtp_password)
                    server.send_message(message)
                logger.info("********** Email Send Successfully To " + toEmail + " **********")
            except smtplib.SMTPException as e:
                logger.info("********** Failed To Send Email To " + toEmail + ", " + str(e) + " **********")


sendEmail()
