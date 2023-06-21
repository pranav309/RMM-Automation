import openpyxl
import datetime
import argparse
import unittest
import HtmlTestRunner


from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from pageObjects.testAllInOne import Test_AllInOne

from utilities.readDriver import get_driver
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen

baseURL = ReadConfig.getApplicationURL()
username = ReadConfig.getUserName()
password = ReadConfig.getPassword()
logger = LogGen.loggen()


def fun1():
    test_suite = unittest.TestSuite()

    # path = r"C:\Users\Pranav Pawar\PycharmProjects\RMM_DataDriven\TestData\tpp.xlsx"
    path = r"C:\Users\Pranav Pawar\PycharmProjects\RMM_DataDriven\TestData\firstFlow\firstFlow.xlsx"
    # path = r"C:\Users\Pranav Pawar\PycharmProjects\RMM_DataDriven\TestData\secondFlow\secondFlow.xlsx"

    workBook = openpyxl.load_workbook(path)
    sheet = workBook.active
    rows = sheet.max_row

    test_suite.addTest(Test_AllInOne('test_clickOnLogin', driver=driver, uName=username, password=password))

    count = 1
    for r in range(2, rows + 1):
        operation = sheet.cell(row=r, column=1).value.lower()

        # Wave Page
        if (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("wave") != -1 and operation.find("without") != -1:
            waveName = sheet.cell(row=r, column=5).value
            passthrough = sheet.cell(row=r, column=6).value
            test_suite.addTest(Test_AllInOne('test_createWaveWithoutHost', driver=driver, count=count, waveName=waveName, trueFalse=passthrough))
            count += 1

        elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("wave") != -1 and operation.find("file") != -1:
            filePath = sheet.cell(row=r, column=2).value
            test_suite.addTest(Test_AllInOne('test_createWaveWithFile', driver=driver, count=count, path=filePath))
            count += 1

        elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("wave") != -1 and operation.find("with") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_createWaveWithHost', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif operation.find("add") != -1 and operation.find("host") != -1 and operation.find("wave") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_addHostToWave', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif operation.find("search") != -1 and operation.find("wave") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_searchWave', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("search") != -1 and operation.find("host") != -1:
            waveName = sheet.cell(row=r, column=5).value
            hostName = sheet.cell(row=r, column=7).value
            test_suite.addTest(Test_AllInOne('test_searchHost', driver=driver, count=count, waveName=waveName, hostName=hostName))
            count += 1

        # Wave Operations
        elif operation.find("start") != -1 and operation.find("wave") != -1 and operation.find("one") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_startWaveOneByOne', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("start") != -1 and operation.find("wave") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_startWave', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("delete") != -1 and operation.find("sync") != -1 and operation.find("relation") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_deleteSRDetails', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("verify") != -1 and operation.find("sync") != -1 and operation.find("success") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_verifySyncSuccess', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("parallel") != -1 and operation.find("count") != -1:
            waveName = sheet.cell(row=r, column=5).value
            parallelCount = sheet.cell(row=r, column=8).value
            test_suite.addTest(Test_AllInOne('test_setParallelCount', driver=driver, count=count, waveName=waveName, val=parallelCount))
            count += 1

        elif (operation.find("change") != -1 or operation.find("remove") != -1) and operation.find("policy") != -1:
            waveName = sheet.cell(row=r, column=5).value
            startNow = sheet.cell(row=r, column=10).value
            test_suite.addTest(Test_AllInOne('test_changePolicy', driver=driver, count=count, waveName=waveName, trueFalse=startNow))
            count += 1

        elif operation.find("stop") != -1 and operation.find("wave") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_stopWave', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("pause") != -1 and operation.find("wave") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_pauseWave', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("restart") != -1 and operation.find("wave") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_restartWave', driver=driver, count=count, waveName=waveName))
            count += 1

        # Wave Edits
        elif (operation.find("add") != -1 or operation.find("set") != -1) and operation.find("autoprovision") != -1:
            filePath = sheet.cell(row=r, column=2).value
            test_suite.addTest(Test_AllInOne('test_setAutoprovision', driver=driver, count=count, path=filePath))
            count += 1

        elif operation.find("bulk") != -1 and operation.find("edit") != -1 and operation.find("sync") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_bulkEditSyncOption', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif operation.find("edit") != -1 and operation.find("sync") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_setSyncOptions', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif operation.find("change") != -1 and operation.find("target") != -1 and operation.find("type") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_changeTargetType', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif (operation.find("move") != -1 or operation.find("shift") != -1) and operation.find("host") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_moveHosts', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif (operation.find("change") != -1 or operation.find("edit") != -1) and (operation.find("vcenter") != -1 or operation.find("vcentre") != -1):
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_changeVcenterData', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        # Wave Details
        elif operation.find("verify") != -1 and operation.find("wave") != -1 and operation.find("sync") != -1 and (operation.find("detail") != -1 or operation.find("details") != -1 or operation.find("data") != -1):
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_verifySyncDetails', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("wave") != -1 and operation.find("status") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_checkWaveStatus', driver=driver, count=count, waveName=waveName))
            count += 1

        elif operation.find("sync") != -1 and (operation.find("successful") != -1 or operation.find("count") != -1):
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_totalSuccessfulSyncs', driver=driver, count=count, waveName=waveName))
            count += 1

        elif (operation.find("check") != -1 or operation.find("find") != -1) and operation.find("host") != -1:
            waveName = sheet.cell(row=r, column=5).value
            hostName = sheet.cell(row=r, column=7).value
            test_suite.addTest(Test_AllInOne('test_checkHosts', driver=driver, count=count, waveName=waveName, hostName=hostName))
            count += 1

        # DR Policy
        elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("dr") != -1 and operation.find("policy") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_createDRPolicy', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif (operation.find("find") != -1 or operation.find("search") != -1) and operation.find("policy") != -1:
            policyName = sheet.cell(row=r, column=9).value
            test_suite.addTest(Test_AllInOne('test_findPolicy', driver=driver, count=count, policyName=policyName))
            count += 1

        elif (operation.find("add") != -1 or operation.find("assign") != -1) and operation.find("policy") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_addDRPolicyToWave', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif operation.find("policy") != -1 and operation.find("status") != -1:
            policyName = sheet.cell(row=r, column=9).value
            test_suite.addTest(Test_AllInOne('test_checkDrPolicyState', driver=driver, count=count, policyName=policyName))
            count += 1

        elif (operation.find("resume") != -1 or operation.find("restart") != -1) and operation.find("policy") != -1:
            policyName = sheet.cell(row=r, column=9).value
            test_suite.addTest(Test_AllInOne('test_resumePolicy', driver=driver, count=count, policyName=policyName))
            count += 1

        elif operation.find("verify") != -1 and operation.find("policy") != -1 and operation.find("sync") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_verifyDRHostSyncStatus', driver=driver, count=count, waveName=waveName))
            count += 1

        elif (operation.find("pause") != -1 or operation.find("stop") != -1) and operation.find("policy") != -1:
            policyName = sheet.cell(row=r, column=9).value
            test_suite.addTest(Test_AllInOne('test_pauseDRPolicy', driver=driver, count=count, policyName=policyName))
            count += 1

        elif operation.find("failover") != -1:
            waveName = sheet.cell(row=r, column=5).value
            testMode = sheet.cell(row=r, column=10).value
            test_suite.addTest(Test_AllInOne('test_failoverHost', driver=driver, count=count, waveName=waveName, trueFalse=testMode))
            count += 1

        elif operation.find("fallback") != -1:
            waveName = sheet.cell(row=r, column=5).value
            test_suite.addTest(Test_AllInOne('test_fallbackHost', driver=driver, count=count, waveName=waveName))
            count += 1

        # Configuration
        elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("cloud") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_addNewCloudUser', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif (operation.find("find") != -1 or operation.find("search") != -1) and operation.find("cloud") != -1:
            userName = sheet.cell(row=r, column=11).value
            test_suite.addTest(Test_AllInOne('test_findClouduser', driver=driver, count=count, uName=userName))
            count += 1

        elif (operation.find("add") != -1 or operation.find("create") != -1) and (operation.find("vcenter") != -1 or operation.find("vcentre") != -1):
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_addVCenter', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif (operation.find("find") != -1 or operation.find("search") != -1) and (operation.find("vcenter") != -1 or operation.find("vcentre") != -1):
            userName = sheet.cell(row=r, column=11).value
            test_suite.addTest(Test_AllInOne('test_findVCenter', driver=driver, count=count, uName=userName))
            count += 1

        # Retention Policy
        elif operation.find("create") != -1 and operation.find("retention") != -1 and operation.find("policy") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_createRetentionPolicy', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        elif operation.find("edit") != -1 and operation.find("retention") != -1 and operation.find("policy") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_editRetentionPolicy', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        # Tear Down
        elif operation.find("tear") != -1 and operation.find("down") != -1:
            filePath = sheet.cell(row=r, column=2).value
            start = sheet.cell(row=r, column=3).value
            end = sheet.cell(row=r, column=4).value
            test_suite.addTest(Test_AllInOne('test_tearDown', driver=driver, count=count, path=filePath, start=start, end=end))
            count += 1

        else:
            logger.info("********** TestCase " + str(count) + ": There Are Some Mistakes In The Operation Keywords '" + str(operation) + "' ... Please Check Once **********\n \n")
            count += 1

    test_suite.addTest(Test_AllInOne('test_clickOnLogout', driver=driver))

    return test_suite


# Generate the report
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Automation Test Flow')
    parser.add_argument('--browser', default='firefox', choices=['chrome', 'firefox', 'edge', 'safari', 'chromium'], help='Specify the browser to use')
    parser.add_argument('--reportLog', help='Specify the report log filename')
    parser.add_argument('--report', help='Specify the report log filename')
    args = parser.parse_args()
    browser_name = args.browser
    logFileName = args.reportLog
    fileName = args.report

    report_file = open('./Reports/Report Log Files/'+logFileName+'.log', 'w')
    runner = HtmlTestRunner.HTMLTestRunner(
        stream=report_file,
        report_name=fileName,
        add_timestamp=False,
        report_title='Automation Flow Report'
    )

    driver = get_driver(browser_name)
    driver.get(baseURL)

    logger.info("********** Test_000_OneForAll ********** ")

    logger.info("********** Opening Browser **********")
    driver.get(baseURL)
    driver.maximize_window()
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "username"))
    )
    logger.info("********** Browser Opened Successfully **********\n \n")

    timestamp = datetime.datetime.now()
    timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
    logger.info(f'Timestamp: {timestamp_str}')
    logger.info('\n')

    runner.run(fun1())
    report_file.close()

    driver.close()
    driver.quit()
    logger.info("********** Browser Closed Successfully **********")

    timestamp = datetime.datetime.now()
    timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
    logger.info('\n')
    logger.info(f'Timestamp: {timestamp_str}')
    logger.info('\n \n \n')
