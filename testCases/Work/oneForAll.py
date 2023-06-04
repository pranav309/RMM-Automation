import openpyxl
import datetime


from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen

from pageObjects.loginPage import LoginPage
from pageObjects.wavePage import WavePage
from pageObjects.waveOperations import WaveOperations
from pageObjects.waveEdit import WaveEdit
from pageObjects.waveDetails import WaveDetails
from pageObjects.retentionPolicyPage import RetentionPolicy
from pageObjects.drPolicyPage import DRPolicy
from pageObjects.configurationPage import Configuration
from pageObjects.tearDown import TearDown


class Test_000_OneForAll:
    baseURL = ReadConfig.getApplicationURL()
    username = ReadConfig.getUserName()
    password = ReadConfig.getPassword()
    logger = LogGen.loggen()

    def test_allInOne(self, setup):
        self.driver = setup
        self.lp = LoginPage(self.driver)
        self.wp = WavePage(self.driver)
        self.wo = WaveOperations(self.driver)
        self.we = WaveEdit(self.driver)
        self.wd = WaveDetails(self.driver)
        self.rp = RetentionPolicy(self.driver)
        self.dr = DRPolicy(self.driver)
        self.conf = Configuration(self.driver)
        self.td = TearDown(self.driver)

        self.logger.info("********** Test_019_SecondFlow ********** ")
        self.logger.info("********** Opening Browser ********** ")
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.logger.info("********** Browser Opened Successfully **********\n \n")

        self.logger.info("********** Logging In **********")
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickOnLogin()
        self.logger.info("********** Login Successful **********\n \n")

        # path = r"C:\Users\Pranav Pawar\PycharmProjects\RMM_DataDriven\TestData\firstFlow\firstFlow.xlsx"
        path = r"C:\Users\Pranav Pawar\PycharmProjects\RMM_DataDriven\TestData\test\drPolicy.xlsx"
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row

        count = 1
        for r in range(2, rows+1):
            operation = sheet.cell(row=r, column=1).value.lower()

            # Wave Page
            if (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("wave") != -1 and operation.find("without") != -1:
                waveName = sheet.cell(row=r, column=5).value
                passthrough = sheet.cell(row=r, column=6).value
                self.logger.info("********** Starting TestCase "+str(count)+": Create New Wave Without Hosts **********")
                self.wp.createWaveWithoutHost(waveName, passthrough)
                self.logger.info("********** Successfully Executed TestCase: Create New Wave Without Hosts **********\n \n")
                count += 1

            elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("wave") != -1 and operation.find("file") != -1:
                filePath = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Create New Wave With Upload .csv File Method **********")
                self.wp.createWaveWithFile(filePath)
                self.logger.info("********** Successfully Executed TestCase: Create New Wave With Upload .csv File Method **********\n \n")
                count += 1

            elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("wave") != -1 and operation.find("with") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Create New Wave With Host **********")
                self.wp.createWaveWithHost(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Create New Wave With Host **********\n \n")
                count += 1

            elif operation.find("add") != -1 and operation.find("host") != -1 and operation.find("wave") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Add New Hosts To Wave **********")
                self.wp.addHostToWave(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Add New Hosts To Wave **********\n \n")
                count += 1

            elif operation.find("search") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Search A Wave **********")
                self.wp.searchWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Search A Wave **********\n \n")
                count += 1

            elif operation.find("search") != -1 and operation.find("host") != -1:
                waveName = sheet.cell(row=r, column=5).value
                hostName = sheet.cell(row=r, column=7).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Search A Host **********")
                self.wp.searchHost(waveName, hostName)
                self.logger.info("********** Successfully Executed TestCase: Search A Host **********\n \n")
                count += 1

            # Wave Operations
            elif operation.find("start") != -1 and operation.find("wave") != -1 and operation.find("one") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Start Wave And Verify **********")
                self.wo.startWaveOneByOne(waveName)
                self.logger.info("********** Successfully Executed TestCase: Start Wave And Verify **********\n \n")
                count += 1

            elif operation.find("start") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Start Wave And Verify **********")
                self.wo.startWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Start Wave And Verify **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and operation.find("shh") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Delete Sync Relation Data From RMM SSH **********")
                self.wo.startWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Sync Relation Data From RMM SSH **********\n \n")
                count += 1

            elif operation.find("verify") != -1 and operation.find("sync") != -1 and operation.find("success") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Verify Sync Success Or Fail Details **********")
                self.wo.verifySyncSuccess(waveName)
                self.logger.info("********** Successfully Executed TestCase: Verify Sync Success Or Fail Details **********\n \n")
                count += 1

            elif operation.find("parallel") != -1 and operation.find("count") != -1:
                waveName = sheet.cell(row=r, column=5).value
                parallelCount = sheet.cell(row=r, column=8).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Set Parallel Count **********")
                self.wo.setParallelCount(waveName, parallelCount)
                self.logger.info("********** Successfully Executed TestCase: Set Parallel Count **********\n \n")
                count += 1

            elif (operation.find("change") != -1 or operation.find("remove") != -1) and operation.find("policy") != -1:
                waveName = sheet.cell(row=r, column=5).value
                if operation.find("remove") != -1:
                    policyName = "No Policy"
                    self.logger.info("********** Starting TestCase " + str(count) + ": Remove DR Policy Of Wave **********")
                else:
                    policyName = sheet.cell(row=r, column=9).value
                    self.logger.info("********** Starting TestCase " + str(count) + ": Change DR Policy Of Wave **********")
                startNow = sheet.cell(row=r, column=10).value
                self.wo.changePolicy(waveName, policyName, startNow)
                if operation.find("remove") != -1:
                    self.logger.info("********** Successfully Executed TestCase: Remove DR Policy Of Wave **********\n \n")
                else:
                    self.logger.info("********** Successfully Executed TestCase: Change DR Policy Of Wave **********\n \n")
                count += 1

            elif operation.find("stop") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Stop Wave **********")
                self.wo.stopWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Stop Wave **********\n \n")
                count += 1

            elif operation.find("pause") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Pause Wave **********")
                self.wo.pauseWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Pause Wave **********\n \n")
                count += 1

            elif operation.find("restart") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Restart Wave **********")
                self.wo.restartWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Restart Wave **********\n \n")
                count += 1

            # Wave Edits
            elif (operation.find("add") != -1 or operation.find("set") != -1) and operation.find("autoprovision") != -1:
                filePath = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Set Autoprovision **********")
                self.we.setAutoprovision(filePath)
                self.logger.info("********** Successfully Executed TestCase: Set Autoprovision **********\n \n")
                count += 1

            elif operation.find("bulk") != -1 and operation.find("edit") != -1 and operation.find("sync") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Bulk Edit Sync Options **********")
                self.we.bulkEditSyncOption(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Bulk Edit Sync Options **********\n \n")
                count += 1

            elif operation.find("edit") != -1 and operation.find("sync") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                if start == "NA":
                    start = 3
                if end == "NA":
                    end = rows+1
                self.logger.info("********** Starting TestCase " + str(count) + ": Set Sync Options **********")
                self.we.setSyncOptions(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Set Sync Options **********\n \n")
                count += 1

            elif operation.find("change") != -1 and operation.find("target") != -1 and operation.find("type") != -1:
                filePath = sheet.cell(row=r, column=2).value
                startRow = sheet.cell(row=r, column=3).value
                endRow = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Change Target Type **********")
                self.we.changeTargetType(filePath, startRow, endRow)
                self.logger.info("********** Successfully Executed TestCase: Change Target Type **********\n \n")
                count += 1

            elif (operation.find("move") != -1 or operation.find("shift") != -1) and operation.find("host") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Move Host Between Waves **********")
                self.we.moveHosts(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Move Host Between Waves **********\n \n")
                count += 1

            elif operation.find("edit") != -1 and (operation.find("vcenter") != -1 or operation.find("vcentre") != -1):
                path = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Change vCenter Details **********")
                self.we.changeVcenterData(path, start, end)
                self.logger.info("********** Successfully Executed TestCase: Change vCenter Details **********\n \n")
                count += 1

            # Wave Details
            elif operation.find("verify") != -1 and operation.find("sync") != -1 and (operation.find("detail") != -1 or operation.find("details") != -1 or operation.find("data") != -1):
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Verify Sync System And Summary Details **********")
                self.wd.verifySyncDetails(waveName)
                self.logger.info("********** Successfully Executed TestCase: Verify Sync System And Summary Details **********\n \n")
                count += 1

            elif operation.find("wave") != -1 and operation.find("status") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Check Wave Status **********")
                self.wd.checkWaveStatus(waveName)
                self.logger.info("********** Successfully Executed TestCase: Check Wave Status **********\n \n")
                count += 1

            elif operation.find("sync") != -1 and (operation.find("successful") != -1 or operation.find("count") != -1):
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Total Successful Syncs **********")
                self.wd.totalSuccessfulSyncs(waveName)
                self.logger.info("********** Successfully Executed TestCase: Total Successful Syncs **********\n \n")
                count += 1

            elif (operation.find("check") != -1 or operation.find("find") != -1) and operation.find("host") != -1:
                waveName = sheet.cell(row=r, column=5).value
                hostName = sheet.cell(row=r, column=7).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Check Host **********")
                self.wd.checkHosts(waveName,hostName)
                self.logger.info("********** Successfully Executed TestCase: Check Host **********\n \n")
                count += 1

            # DR Policy
            elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("dr") != -1 and operation.find("policy") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Create New DR Policy **********")
                self.dr.createDRPolicy(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Create New DR Policy **********\n \n")
                count += 1

            elif (operation.find("find") != -1 or operation.find("search") != -1) and operation.find("policy") != -1:
                policyName = sheet.cell(row=r, column=9).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Search DR Policy **********")
                self.dr.findPolicy(policyName)
                self.logger.info("********** Successfully Executed TestCase: Search DR Policy **********\n \n")
                count += 1

            elif (operation.find("add") != -1 or operation.find("assign") != -1) and operation.find("policy") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Assign DR Policy To Wave **********")
                self.dr.addDRPolicyToWave(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Assign DR Policy To Wave **********\n \n")
                count += 1

            elif operation.find("policy") != -1 and operation.find("status") != -1:
                policyName = sheet.cell(row=r, column=9).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Check DR Policy Status **********")
                self.dr.checkDrPolicyState(policyName)
                self.logger.info("********** Successfully Executed TestCase: Check DR Policy Status **********\n \n")
                count += 1

            elif (operation.find("resume") != -1 or operation.find("restart") != -1) and operation.find("policy") != -1:
                policyName = sheet.cell(row=r, column=9).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Resume Policy **********")
                self.dr.resumePolicy(policyName)
                self.logger.info("********** Successfully Executed TestCase: Resume Policy **********\n \n")
                count += 1

            elif operation.find("verify") != -1 and operation.find("policy") != -1 and operation.find("sync") != -1:
                waveName = sheet.cell(row=r, column=5).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Check DR Policy Host Sync Status **********")
                self.dr.verifyDRHostSyncStatus(waveName)
                self.logger.info("********** Successfully Executed TestCase: Check DR Policy Host Sync Status **********\n \n")
                count += 1

            elif (operation.find("pause") != -1 or operation.find("stop") != -1) and operation.find("policy") != -1:
                policyName = sheet.cell(row=r, column=9).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Pause DR Policy **********")
                self.dr.pauseDRPolicy(policyName)
                self.logger.info("********** Successfully Executed TestCase: Pause DR Policy **********\n \n")
                count += 1

            elif operation.find("failover") != -1:
                waveName = sheet.cell(row=r, column=5).value
                testMode = sheet.cell(row=r, column=10).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Failover Host **********")
                self.dr.failoverHost(waveName, testMode)
                self.logger.info("********** Successfully Executed TestCase: Failover Host **********\n \n")
                count += 1

            # Configuration
            elif (operation.find("add") != -1 or operation.find("create") != -1) and operation.find("cloud") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Create New Cloud User **********")
                self.conf.addNewCloudUser(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Create New Cloud User **********\n \n")
                count += 1

            elif (operation.find("find") != -1 or operation.find("search") != -1) and operation.find("cloud") != -1:
                userName = sheet.cell(row=r, column=11).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Find Cloud User **********")
                self.conf.findClouduser(userName)
                self.logger.info("********** Successfully Executed TestCase: Find Cloud User **********\n \n")
                count += 1

            elif (operation.find("add") != -1 or operation.find("create") != -1) and (operation.find("vcenter") != -1 or operation.find("vcentre") != -1):
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Create New vCenter **********")
                self.conf.addVCenter(filePath, start, end)
                self.logger.info("********** Successfully Executed TestCase: Create New vCenter **********\n \n")
                count += 1

            elif (operation.find("find") != -1 or operation.find("search") != -1) and (operation.find("vcenter") != -1 or operation.find("vcentre") != -1):
                userName = sheet.cell(row=r, column=11).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Create New Cloud User **********")
                self.conf.findVCenter(userName)
                self.logger.info("********** Successfully Executed TestCase: Create New Cloud User **********\n \n")
                count += 1

            # Tear Down
            elif operation.find("tear") != -1 and operation.find("down") != -1:
                filePath = sheet.cell(row=r, column=2).value
                start = sheet.cell(row=r, column=3).value
                end = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase " + str(count) + ": Perform Tear Down **********")
                self.td.tearDown(filePath, start, end, count)
                self.logger.info("********** Successfully Executed TestCase: Perform Tear Down **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and operation.find("SSH") != -1 and operation.find("entry") != -1:
                source = sheet.cell(row=r, column=5).value
                target = sheet.cell(row=r, column=6).value
                self.logger.info("********** Starting TestCase "+str(count)+": Delete SSH Host Sync Entry **********")
                self.td.deleteSR(source, target)
                self.logger.info("********** Successfully Executed TestCase: Delete SSH Host Sync Entry **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and operation.find("host") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=2).value
                hostNames = sheet.cell(row=r, column=3).value
                self.logger.info("********** Starting TestCase "+str(count)+": Delete Host In Wave **********")
                self.td.deleteHostInWave(waveName, hostNames)
                self.logger.info("********** Successfully Executed TestCase: Delete Host In Wave **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and operation.find("single") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+": Delete Wave **********")
                self.td.deleteWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Wave **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and operation.find("replication") != -1 and (operation.find("wave") != -1 or operation.find("waves") != -1):
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+": Delete Wave **********")
                self.td.deleteRepWaves(waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Wave **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and operation.find("DR") != -1 and (operation.find("wave") != -1 or operation.find("waves") != -1):
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+": Delete Wave **********")
                self.td.deleteDrWaves(waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Wave **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and operation.find("dr") != -1 and (operation.find("policy") != -1 or operation.find("policies") != -1):
                policyName = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase "+str(count)+": Delete SSH Host Sync Entry **********")
                self.td.deleteDRPolicy(policyName)
                self.logger.info("********** Successfully Executed TestCase: Delete SSH Host Sync Entry **********\n \n")
                count += 1

            elif operation.find("delete") != -1 and (operation.find("host") != -1 or operation.find("hosts") != -1):
                hostNames = sheet.cell(row=r, column=3).value
                VM = sheet.cell(row=r, column=7).value
                environment = sheet.cell(row=r, column=8).value
                name = sheet.cell(row=r, column=9).value
                self.logger.info("********** Starting TestCase "+str(count)+": Delete Host In Wave **********")
                self.td.deleteHost(hostNames, VM, environment, name)
                self.logger.info("********** Successfully Executed TestCase: Delete Host In Wave **********\n \n")
                count += 1

            else:
                self.logger.info("********** TestCase " + str(count) + ": There Are Some Mistakes In The Operation Keywords '" + str(operation) + "' ... Please Check Once **********\n \n")
                count += 1

        self.lp.clickOnLogout()
        self.logger.info("********** Logout Successful **********")
        self.driver.close()
        self.logger.info("********** Browser Closed Successfully **********")
        timestamp = datetime.datetime.now()
        timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
        self.logger.info('\n')
        self.logger.info(f'Timestamp: {timestamp_str}')
        self.logger.info('\n \n \n')
