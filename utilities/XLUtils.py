import openpyxl


class XLUtil:
    @staticmethod
    def getRowCount(file, sheetName):
        wb = openpyxl.load_workbook(file)
        sheet = wb.get_sheet_by_name(sheetName)
        return sheet.max_row

    @staticmethod
    def getColumnCount(file, sheetName):
        wb = openpyxl.load_workbook(file)
        sheet = wb.get_sheet_by_name(sheetName)
        return sheet.max_column

    @staticmethod
    def readData(file, sheetName, rowNum, colNum):
        wb = openpyxl.load_workbook(file)
        sheet = wb.get_sheet_by_name(sheetName)
        return sheet.cell(row=rowNum, column=colNum).value

    @staticmethod
    def writeData(file, sheetName, rowNum, colNum, data):
        wb = openpyxl.load_workbook(file)
        sheet = wb.get_sheet_by_name(sheetName)
        sheet.cell(row=rowNum, column=colNum).value = data
        wb.save(file)
