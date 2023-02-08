import time
import openpyxl

from selenium.webdriver.common.by import By


class RetentionPolicy:
    btn_add_xpath = "//*[@id='retention_add_retention']/span/i"
    btn_create_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[3]/div/button[2]"
    txt_policyName_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[1]/div/input"

    # By Interval
    rd_byInterval_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[2]/div/div[1]/div/input"

    txt_shortMinInt_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[2]/div[1]/input"
    btn_shortMinIntHours_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[2]/div[2]/p-selectbutton/div/div[1]/span"
    btn_shortMinIntDays_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[2]/div[2]/p-selectbutton/div/div[2]/span"
    btn_shortMinIntWeeks_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[2]/div[2]/p-selectbutton/div/div[3]/span"

    txt_shortMaxWin_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[3]/div[1]/input"
    btn_shortMaxWinHours_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[3]/div[2]/p-selectbutton/div/div[1]/span"
    btn_shortMaxWinDays_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[3]/div[2]/p-selectbutton/div/div[2]/span"
    btn_shortMaxWinWeeks_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[3]/div[2]/p-selectbutton/div/div[3]/span"

    txt_longMinInt_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[2]/div[1]/input"
    btn_longMinIntDays_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[2]/div[2]/p-selectbutton/div/div[1]/span"
    btn_longMinIntWeeks_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[2]/div[2]/p-selectbutton/div/div[2]/span"

    txt_longMaxWin_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[3]/div[1]/input"
    btn_longMaxWinDays_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[3]/div[2]/p-selectbutton/div/div[1]/span"
    btn_longMaxWinWeeks_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[3]/div[2]/p-selectbutton/div/div[2]/span"

    # By Count
    rd_byCount_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[2]/div/div[2]/div/input"
    txt_retainCount_xpath = "//*[@id='main']/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div/input"

    # Start
    txt_imageName_xpath = "//*[@id='main']/app-retention-policies/apply-retention-policy/div/div/div/form/div[2]/div[2]/div/p-multiselect/div/div[2]/span"
    txt_searchImage_xpath = "//*[@id='main']/app-retention-policies/apply-retention-policy/div/div/div/form/div[2]/div[2]/div/p-multiselect/div/div[4]/div[1]/div[2]/input"

    # Delete
    btn_deleteRetentionPolicy_id = "conf_cu_del_cloud_modal_del_btn"

    def __init__(self, driver):
        self.driver = driver

    def createRetentionPolicy(self, path):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        for r in range(2, rows+1):
            policyName = sheet.cell(row=r, column=2).value
            retentionType = sheet.cell(row=r, column=3).value
            retainCount = sheet.cell(row=r, column=4).value
            shortMinInt = sheet.cell(row=r, column=5).value
            shortMinIntType = sheet.cell(row=r, column=6).value
            shortMaxWin = sheet.cell(row=r, column=7).value
            shortMaxWinType = sheet.cell(row=r, column=8).value
            longMinInt = sheet.cell(row=r, column=9).value
            longMinIntType = sheet.cell(row=r, column=10).value
            longMaxWin = sheet.cell(row=r, column=11).value
            longMaxWinType = sheet.cell(row=r, column=12).value

            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_add_xpath).click()
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.txt_policyName_xpath).send_keys(policyName)
            if retentionType == "By Interval":
                self.driver.find_element(By.XPATH, self.rd_byInterval_xpath).click()
                self.driver.find_element(By.XPATH, self.txt_shortMinInt_xpath).send_keys(shortMinInt)
                self.selectShortType(shortMinIntType, 2)
                self.driver.find_element(By.XPATH, self.txt_shortMaxWin_xpath).send_keys(shortMaxWin)
                self.selectShortType(shortMaxWinType, 3)
                self.driver.find_element(By.XPATH, self.txt_longMinInt_xpath).send_keys(longMinInt)
                self.selectLongType(longMinIntType, 2)
                self.driver.find_element(By.XPATH, self.txt_longMaxWin_xpath).send_keys(longMaxWin)
                self.selectLongType(longMaxWinType, 3)

            elif retentionType == "By Count":
                self.driver.find_element(By.XPATH, self.rd_byCount_xpath).click()
                self.driver.find_element(By.XPATH, self.txt_retainCount_xpath).send_keys(retainCount)

            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_create_xpath).click()

        time.sleep(10)

    def selectShortType(self, val, minMax):
        if val == "Hours":
            self.driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[1]/span').click()
        elif val == "Days":
            self.driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[2]/span').click()
        elif val == "Weeks":
            self.driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[3]/span').click()

    def selectLongType(self, val, minMax):
        if val == "Days":
            self.driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[1]/span').click()
        elif val == "Weeks":
            self.driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[2]/span').click()

    def startRetentionPolicy(self, vals, images):
        for i in vals:
            self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[7]/span/div/i[1]').click()
            time.sleep(2)
            self.driver.find_element(By.XPATH, self.txt_imageName_xpath).click()
            time.sleep(2)
            self.driver.find_element(By.XPATH, self.txt_searchImage_xpath).send_keys()

    def editRetentionPolicy(self, path):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        for r in range(2, rows+1):
            srn = sheet.cell(row=r, column=1).value
            self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(srn)+']/td[7]/span/div/i[2]').click()
            time.sleep(3)
            self.createRetentionPolicy(path)

    def deleteRetentionPolicy(self, vals):
        for i in vals:
            self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[7]/span/div/i[3]').click()
            self.driver.find_element(By.ID, self.btn_deleteRetentionPolicy_id).click()
            time.sleep(3)
