import time
import unittest
import openpyxl
import keyboard

import Locators.locWavePage as LOC

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from utilities.customLogger import LogGen
from utilities.commonObjects import CommonObjects


class WavePage(unittest.TestCase):

    logger = LogGen.loggen()

    def createWaveWithoutHost(self, driver, waveName, passthrough):
        replication_class = driver.find_element(By.XPATH, LOC.txt_replication_xpath).get_attribute("class")
        if replication_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        wave_class = driver.find_element(By.XPATH, LOC.txt_rWave_xpath).get_attribute("class")
        if wave_class != "active":
            driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
        if len(driver.find_elements(By.LINK_TEXT, waveName)) == 0:
            if len(driver.find_elements(By.XPATH, LOC.img_createWave_xpath)) != 0:
                driver.find_element(By.XPATH, LOC.img_createWave_xpath).click()
            else:
                driver.find_element(By.XPATH, LOC.btn_createWave_xpath).click()
                time.sleep(3)
                driver.find_element(By.XPATH, LOC.img_createNewWave_xpath).click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_createWave_xpath)) != 0:
                self.logger.info("********** Create New Wave Pop-up Banner Was Opened For Wave, "+str(waveName)+" **********")
                driver.find_element(By.ID, LOC.txt_waveName_id).send_keys(waveName)
                if not passthrough:
                    driver.find_element(By.ID, LOC.chBox_passthrough_id).click()
                ele1 = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_create_id))
                )
                ele1.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Create New Wave Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Create New Wave Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
        else:
            self.logger.info("********** Wave : " + waveName + ", Is Already Present **********")

    def createWaveWithFile(self, driver, path):
        replication_class = driver.find_element(By.XPATH, LOC.txt_replication_xpath).get_attribute("class")
        if replication_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        wave_class = driver.find_element(By.XPATH, LOC.txt_rWave_xpath).get_attribute("class")
        if wave_class != "active":
            driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
        btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_createWave_xpath))
        )
        btn.click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_createWave_xpath)) != 0:
            self.logger.info("********** Create New Wave Pop-up Banner Was Opened **********")
            driver.find_element(By.XPATH, LOC.img_uploadCSV_xpath).click()
            keyboard.write(path)
            keyboard.send('enter')
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
            )
            note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
            self.logger.info("********** Create Status of Wave Name, ")
            self.logger.info(note + "\n")
            time.sleep(2)
            driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
        else:
            self.logger.info("********** Create New Wave Pop-up Banner Was Not Opened **********")

    def createWaveWithHost(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 3
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            passthrough = sheet.cell(row=r, column=2).value
            replication_class = driver.find_element(By.XPATH, LOC.txt_replication_xpath).get_attribute("class")
            if replication_class == "ng-star-inserted":
                driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(5)
            wave_class = driver.find_element(By.XPATH, LOC.txt_rWave_xpath).get_attribute("class")
            if wave_class != "active":
                driver.find_element(By.LINK_TEXT, "Waves").click()
                time.sleep(5)
            if len(driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                if len(driver.find_elements(By.XPATH, LOC.img_createWave_xpath)) != 0:
                    driver.find_element(By.XPATH, LOC.img_createWave_xpath).click()
                else:
                    driver.find_element(By.XPATH, LOC.btn_createWave_xpath).click()
                    time.sleep(3)
                    driver.find_element(By.XPATH, LOC.img_createNewWave_xpath).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.pop_createWave_xpath)) != 0:
                    self.logger.info("********** Create Wave Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
                    self.logger.info("********** Creating No. " + str(r - 2) + " Wave With Target Type " + sheet.cell(row=r,column=3).value + " **********")
                    driver.find_element(By.ID, LOC.rd_waveWithHost_id).click()
                    driver.find_element(By.ID, LOC.txt_waveName_id).send_keys(waveName)
                    if not passthrough:
                        driver.find_element(By.ID, LOC.chBox_passthrough_id).click()
                    time.sleep(3)
                    self.enterData(driver, sheet, r)
                    ele1 = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, LOC.btn_create_id))
                    )
                    ele1.click()
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                    )
                    note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                    self.logger.info("********** Create New Wave Status of Wave : " + waveName + ",")
                    self.logger.info(note + "\n")
                    time.sleep(2)
                    driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
                else:
                    self.logger.info("********** Create Wave Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
            else:
                self.logger.info("********** Wave : " + waveName + ", Is Already Present **********")

    def addHostToWave(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            hostName = sheet.cell(row=r, column=5).value
            tmp = "none"
            if tmp != waveName:
                co = CommonObjects(driver)
                val = co.findWave(waveName)
                if val == 2:
                    continue
                if val == 1:
                    driver.find_element(By.LINK_TEXT, waveName).click()
                    time.sleep(5)
                    if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                        self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
                    else:
                        self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                        continue
                if val == 0:
                    if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                        self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
            host = self.searchHostInsideWave(driver, hostName)
            if host:
                self.logger.info("********** The Host " + hostName + " Was Already Present In The Wave : " + waveName + " **********")
                continue
            ele1 = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, LOC.btn_addHost_id))
            )
            ele1.click()
            time.sleep(3)
            if len(driver.find_elements(By.XPATH, LOC.pop_addHost_xpath)) != 0:
                self.logger.info("********** Add New Host Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
                self.enterData(driver, sheet, r)
                ele1 = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_createHost_id))
                )
                ele1.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Add New Host Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Add New Host Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")

    def enterData(self, driver, sheet, r):
        # Source Data
        sourceDNS_IP = sheet.cell(row=r, column=4).value
        sourceFriendlyName = sheet.cell(row=r, column=5).value
        sourceOS = sheet.cell(row=r, column=6).value
        sourceSSHport = sheet.cell(row=r, column=7).value
        sourceUserName = sheet.cell(row=r, column=8).value

        driver.find_element(By.ID, LOC.txt_sourceDNSIP_id).send_keys(sourceDNS_IP)
        driver.find_element(By.ID, LOC.txt_sourceFriendlyName_id).clear()
        driver.find_element(By.ID, LOC.txt_sourceFriendlyName_id).send_keys(sourceFriendlyName)
        if sourceOS == "Linux":
            driver.find_element(By.ID, LOC.rd_OS0_id).click()
        elif sourceOS == "Windows":
            driver.find_element(By.ID, LOC.rd_OS1_id).click()
        driver.find_element(By.CSS_SELECTOR, LOC.txt_sourceSshPort_css).send_keys(sourceSSHport)
        driver.find_element(By.ID, LOC.txt_sourceUserName_id).clear()
        driver.find_element(By.ID, LOC.txt_sourceUserName_id).send_keys(sourceUserName)

        # Target Data
        targetType = sheet.cell(row=r, column=3).value
        syncType = sheet.cell(row=r, column=9).value
        targetFriendlyName = sheet.cell(row=r, column=10).value
        targetImageName = sheet.cell(row=r, column=11).value
        targetDNS_IP = sheet.cell(row=r, column=12).value
        sourceUserPassword = sheet.cell(row=r, column=13).value
        targetUsername = sheet.cell(row=r, column=14).value
        targetSSHport = sheet.cell(row=r, column=15).value

        # Autoprovision
        if targetType == "Autoprovision":
            driver.find_element(By.ID, LOC.rd_targetType0_id).click()
            if syncType == "Direct Sync":
                driver.find_element(By.ID, LOC.rd_syncType0_id).click()
            else:
                if syncType == "Stage 1 & 2":
                    driver.find_element(By.ID, LOC.rd_syncType1_id).click()
                elif syncType == "Stage 1":
                    driver.find_element(By.ID, LOC.rd_syncType2_id).click()
                elif syncType == "Stage 2":
                    driver.find_element(By.ID, LOC.rd_syncType3_id).click()
                driver.find_element(By.ID, LOC.txt_imageName_id).clear()
                driver.find_element(By.ID, LOC.txt_imageName_id).send_keys(targetImageName)
            driver.find_element(By.ID, LOC.txt_targetFriendlyName_id).clear()
            driver.find_element(By.ID, LOC.txt_targetFriendlyName_id).send_keys(targetFriendlyName)

        # Exiting System
        elif targetType == "Existing System":
            driver.find_element(By.ID, LOC.rd_targetType1_id).click()
            if syncType == "Direct Sync":
                driver.find_element(By.ID, LOC.rd_syncType0_id).click()
                self.selectSyncType(driver, targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
                driver.find_element(By.CSS_SELECTOR, LOC.txt_targetSshPort_css).send_keys(targetSSHport)
            elif targetType == "Stage 1 & 2":
                driver.find_element(By.ID, LOC.rd_syncType1_id).click()
                driver.find_element(By.ID, LOC.txt_imageName_id).clear()
                driver.find_element(By.ID, LOC.txt_imageName_id).send_keys(targetImageName)
                self.selectSyncType(driver, targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
                driver.find_element(By.CSS_SELECTOR, LOC.txt_targetSshPort_css).send_keys(targetSSHport)
            elif syncType == "Stage 1":
                driver.find_element(By.ID, LOC.rd_syncType2_id).click()
                driver.find_element(By.ID, LOC.txt_imageName_id).clear()
                driver.find_element(By.ID, LOC.txt_imageName_id).send_keys(targetImageName)
                self.selectSyncType(driver, targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
            elif syncType == "Stage 2":
                driver.find_element(By.ID, LOC.rd_syncType3_id).click()
                driver.find_element(By.ID, LOC.txt_imageName_id).clear()
                driver.find_element(By.ID, LOC.txt_imageName_id).send_keys(targetImageName)
                self.selectSyncType(driver, targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
                driver.find_element(By.CSS_SELECTOR, LOC.txt_targetSshPort_css).send_keys(targetSSHport)

        # Capture
        elif targetType == "Capture":
            driver.find_element(By.ID, LOC.rd_targetType2_id).click()
            driver.find_element(By.ID, LOC.txt_imageName_id).clear()
            driver.find_element(By.ID, LOC.txt_imageName_id).send_keys(targetImageName)

    def selectSyncType(self, driver, DNS_IP, friendlyName, sourceUserPassword, targetUsername):
        time.sleep(5)
        driver.find_element(By.ID, LOC.txt_targetDNSIP_id).send_keys(DNS_IP)
        driver.find_element(By.ID, LOC.txt_targetFriendlyName_id).clear()
        driver.find_element(By.ID, LOC.txt_targetFriendlyName_id).send_keys(friendlyName)
        if sourceUserPassword:
            driver.find_element(By.ID, LOC.chBox_sourceUserPassword_id).click()
        else:
            driver.find_element(By.ID, LOC.txt_targetUserName_id).send_keys(targetUsername)

    def searchWave(self, driver, waveNames):
        res = tuple(map(str, waveNames.split(', ')))
        co = CommonObjects(driver)
        replication_class = driver.find_element(By.XPATH, LOC.txt_replication_xpath).get_attribute("class")
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if replication_class == "ng-star-inserted" and dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        for waveName in res:
            flag = 0
            if replication_class == "ng-star-inserted" and dr_class == "ng-star-inserted":
                driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(3)
            replication_class = driver.find_element(By.XPATH, LOC.txt_replication_xpath).get_attribute("class")
            if replication_class == "ng-star-inserted open":
                val = co.findWaveCommonOne(waveName, 0)
                if val == 0:
                    self.logger.info("********** Wave: " + waveName + ", Was Found In Replication Waves **********")
                    flag += 1
            elif dr_class == "ng-star-inserted open":
                val = co.findWaveCommonTwo(waveName, 0)
                if val == 0:
                    self.logger.info("********** Wave: " + waveName + ", Was Found In DR Waves **********")
                    flag += 1
            if flag > 0:
                self.logger.info("********** Wave: " + waveName + ", Was Not Found **********")

    def searchHost(self, driver, waveName, hostName):
        co = CommonObjects(driver)
        val = co.findWave(waveName)
        if val == 2:
            return
        if val == 1:
            driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                return
        if val == 0:
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, LOC.txt_waveName_xpath))
        )
        totalHosts = len(driver.find_elements(By.ID, LOC.txt_totalHosts_id))
        for hostNo in range(1, totalHosts + 1):
            if totalHosts == 1:
                tmp = driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
            else:
                tmp = driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
            if hostName == tmp:
                self.logger.info("********** The Host " + hostName + " Was Found In The Wave : " + waveName + " **********")
                break
            elif hostNo == totalHosts:
                self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")

    def searchHostInsideWave(self, driver, hostName):
        time.sleep(5)
        totalHosts = len(driver.find_elements(By.ID, LOC.txt_totalHosts_id))
        for hostNo in range(1, totalHosts + 1):
            if totalHosts == 1:
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
            else:
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
            if hostName == tmp:
                return True
        return False
