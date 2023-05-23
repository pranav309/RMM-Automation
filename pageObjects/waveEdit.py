import time
import openpyxl
import keyboard

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from utilities.customLogger import LogGen


class WaveEdit:
    # Autoprovision
    txt_autoprovision_id = "wave_policy_wave_policy_wave_detail_autoprovision"
    txt_environment_xpath = "//*[@id='clouduser']"
    txt_clusterName_xpath = "//*[@id='wave_detail_cu_edit_vc_clustername']/div/input"
    txt_ESXHost_xpath = "//*[@id='wave_detail_cu_edit_vc_esx_host']/div/input"
    txt_Datastore_xpath = "//*[@id='wave_detail_cu_edit_vc_dc']/div/input"
    btn_applyChanges_id = "wave_detail_cu_edit_apply_changes_btn"
    btn_cancelEdit_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[1]/button/span'

    # Single Edit Set NIC
    btn_NICAdd_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[3]/div/div/div[4]/div[2]/a"
    txt_deviceName_id = "wave_detail_edit_item_nic_device_name"
    txt_type_id = "wave_detail_edit_item_nic_type"
    txt_networkName_id = "wave_detail_edit_item_nic_networkName"
    rd_DHCP_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/edit-item-nic/div/div/div/form/div[2]/div[1]/div[4]/div/div[1]/div/div/input"
    rd_staticIP_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/edit-item-nic/div/div/div/form/div[2]/div[1]/div[4]/div/div[2]/div/div/input"
    txt_CIDR_id = "wave_detail_edit_item_nic_cidr"
    txt_gateway_id = "wave_detail_edit_item_nic_gateway"
    txt_DNS1_id = "wave_detail_edit_item_nic_dns1"
    txt_DNS2_id = "wave_detail_edit_item_nic_dns2"
    btn_save_id = "wave_detail_edit_item_nic_save_btn"

    # Bulk Edit Set NIC
    btn_NICAddBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[4]/div/p-tabview/div/div/p-tabpanel[2]/div/div/div[5]/div[3]/a/i"
    txt_deviceNameBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[1]/div[1]/span/input"
    txt_typeBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[1]/div[2]/span/input"
    txt_networkNameBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[1]/div[3]/span/input"
    rd_DHCPBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[1]/div[4]/div/div[1]/div/div/input"
    rd_staticIPBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[1]/div[4]/div/div[2]/div/div/input"
    txt_CIDRBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[2]/div[1]/span/input"
    txt_gatewayBulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[2]/div[2]/span/input"
    txt_DNS1BulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[2]/div[3]/span/input"
    txt_DNS2BulkEdit_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[2]/div[2]/div[4]/span/input"
    btn_save_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[3]/div/button[2]"

    # Sync Options
    btn_syncOptions_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/ul/li[2]/a/span"
    btn_modify_id = "wave_detail_edit_item_modify_btn"
    rd_tng_id = "wave_detail_edit_item_options_tng"
    rd_verbose_id = "wave_detail_edit_item_options_verbose"
    rd_directFScopy_id = "wave_detail_edit_item_options_allowdirectfscopy"
    rd_FSDeletion_id = "wave_detail_edit_item_options_allowFsDeletion"
    rd_NoTransfer_id = "wave_detail_edit_item_options_no_xfer"
    rd_transferCompress_id = "wave_detail_edit_item_options_xfer_compress"
    rd_noTransferCompress_id = "wave_detail_edit_item_options_no_xfer_compress"
    rd_ignoreMissing_id = "wave_detail_edit_item_options_ignore_missing"
    rd_noInPlace_id = "wave_detail_edit_item_options_noInPlace"
    rd_noReboot_id = "wave_detail_edit_item_options_noReboot"
    rd_includeSAN_id = "wave_detail_edit_item_options_include_san"
    rd_excludeSAN_id = "wave_detail_edit_item_options_exclude_san"
    rd_overrideRMMStorageCheck_id = "wave_detail_edit_item_options_storage_override"
    rd_deleteAllTargetFS_id = "wave_detail_edit_item_options_delete_all_target_fs"
    rd_keepTargetLayout_id = "wave_detail_edit_item_options_keep_target_layout"
    rd_cloudInit_id = "wave_detail_edit_item_options_cloud_init"
    txt_eventScript_id = "wave_detail_edit_item_options_eventScript"
    txt_eventScriptArgs_id = "wave_detail_edit_item_options_eventArgs"
    txt_excludeFile_id = "wave_detail_edit_item_options_exclude"
    txt_includeFile_id = "wave_detail_edit_item_options_include"
    rd_uploadLocalFile_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[2]/div/div[2]/div[3]/div[1]/div[1]/div/div/input"
    rd_filePathOnRmm_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[2]/div/div[2]/div[3]/div[1]/div[2]/div/div/input"
    btn_uploadLocalFile_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[2]/div/div[2]/div[3]/div[2]/div/div/label"
    txt_filePathOnRmm_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[2]/div/div[2]/div[3]/div[2]/input"

    # Bulk Edit Sync Options
    btn_selectAll_xpath = "//*[@id='content']/article/div/div[2]/p-table/div/div[2]/table/thead/tr/th[1]/p-tableheadercheckbox/div/div[2]"
    btn_bulkEdit_xpath = "//*[@id='content']/article/div/div[2]/p-table/div/div[1]/div[1]/button[8]/span/i"
    drp_goal_id = "wave_detail_bulk_edit_item_existing_capture_type"
    btn_next_xpath = "//*[@id='main']/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[5]/div/button[2]"
    btn_modifyAll_xpath = "//*[@id='main']/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[5]/div/button[2]"
    btn_yes_id = "wave_policy_wave_policy_wave_detail_autoprov_not_conf_yes_btn"
    txt_blkExcludeFile_id = "wave_detail_bulk_edit_item_options_exclude"
    txt_blkIncludeFile_id = "wave_detail_bulk_edit_item_options_include"

    # AWS CU
    txt_AWSVPCID_id = "wave_detail_cu_edit_aws_options_vpc_id"
    txt_AWSSubnetID_id = "wave_detail_cu_edit_aws_options_subnet_id"

    # OCI Sync Options
    txt_OCIVCNName_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[3]/div/div[1]/div[3]/div/input"
    txt_OCISubnetName_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[3]/div/div[1]/div[4]/div/input"
    txt_OCIAVDomain_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[3]/div/div[1]/div[9]/div/input"

    # Change Target Type
    rd_autoprovision_id = "wave_detail_edit_item_provisioning_model_radio_0"
    rd_existingSystem_id = "wave_detail_edit_item_provisioning_model_radio_1"
    rd_capture_id = "wave_detail_edit_item_provisioning_model_radio_2"
    txt_captureImage_xpath = "//*[@id='wave_detail_edit_item_clone_name']"
    rd_directSync_id = "wave_detail_edit_item_existing_capture_type_0"
    rd_stage12_id = "wave_detail_edit_item_existing_capture_type_1"
    rd_stage1_id = "wave_detail_edit_item_existing_capture_type_2"
    rd_stage2_id = "wave_detail_edit_item_existing_capture_type_3"
    txt_imageName_id = "wave_detail_edit_item_clone_name"
    ch_passthrough_id = "wave_detail_edit_item_options_sshonly_target"
    txt_targetIP_id = "wave_detail_edit_item_target_dns_ip"
    txt_friendlyName_id = "wave_detail_edit_item_target_friendlyname"
    txt_userName_id = "wave_detail_edit_item_target_username"

    # Change Datastore
    txt_editDatastore_xpath = "//*[@id='target_vcenter_datastore']/div/input"

    # Pop-up Banners
    var_waveDetails_xpath = '//*[@id="rmm_lite_header"]/div/div[1]/div[2]'
    pop_setAutoprovision_xpath = '//*[@id="wave_policy_wave_policy_wave_detail_cloudUserEdit"]/div/div/div/form/div[1]/h4'
    pop_edit_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[2]/div/div[1]/div[1]/div/label'
    pop_system_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[1]/div/div[1]/div/div/label'
    pop_syncOption_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[2]/div/div[1]/div[1]/div/label'
    pop_vCenter_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[3]/div/div/div[1]/label'
    pop_addNIC_xpath = '//*[@id="wave_detail_edit_item_edit_item_nic"]/div/div/div/form/div[1]/h4'
    var_addNIC_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[3]/div/div/div[4]/div[1]/div/div/div[1]/div[1]/b'
    pop_bulkEdit_xpath = '//*[@id="main"]/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[1]/h4'
    pop_bulkSO_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[4]/div/p-tabview/div/ul/li[1]/a/span'
    pop_bulkVC_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[4]/div/p-tabview/div/ul/li[2]/a/span'
    pop_bulkNIC_xpath = '//*[@id="main"]/rw-wave-detail/bulk-edit/edit-item-nic/div/div/div/form/div[1]/h4'
    var_bulkNIC_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[4]/div/p-tabview/div/div/p-tabpanel[2]/div/div/div[5]/div[1]/div/div/div[1]/div[1]/b'
    pop_bulkSyncOpt_xpath = '/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[4]/div/p-tabview/div/div/p-tabpanel[1]/div/div[1]/div[1]/label'
    pop_moveHost_xpath = '//*[@id="wave_policy_wave_policy_wave_detail_move_item"]/div/div/div/div[1]/h4'
    pop_successful_xpath = '/html/body/app-root/simple-notifications/div/simple-notification/div'
    pop_deleteSuccessful_xpath = '/html/body/app-root/simple-notifications/div/simple-notification[2]/div'

    # Move hosts
    drp_selectWave_xpath = "//*[@id='wave_policy_wave_policy_wave_detail_move_item']/div/div/div/div[2]/div/div/select"
    btn_moveMachine_id = "wave_detail_move_item_move_machine_btn"

    logger = LogGen.loggen()

    def __init__(self, driver):
        self.driver = driver

    def findWave(self, waveName):
        time.sleep(5)
        flag = 0
        if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
            if (len(self.driver.find_elements(By.LINK_TEXT, "Summary"))) == 0:
                self.driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
            if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                self.driver.find_element(By.LINK_TEXT, "DR").click()
                flag += 1
                time.sleep(5)
                self.driver.find_element(By.LINK_TEXT, "Waves").click()
                time.sleep(5)
                if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                    flag += 1
                    self.logger.info("********** Wave : " + waveName + " Is Not Present **********")
        return flag

    def setAutoprovision(self, path):
        time.sleep(5)
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        time.sleep(5)
        for r in range(3, rows+1):
            waveName = sheet.cell(row=r, column=1).value
            CUType = sheet.cell(row=r, column=2).value
            environment = sheet.cell(row=r, column=3).value

            val = self.findWave(waveName)
            if val == 2:
                return
            time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.var_waveDetails_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Opened Successfully **********")
                self.driver.find_element(By.ID, self.txt_autoprovision_id).click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.pop_setAutoprovision_xpath)) != 0:
                    self.logger.info("********** Select An Environment Pop-up Banner Is Opened For Wave, " + str(waveName) + " **********")
                    env = Select(self.driver.find_element(By.XPATH, self.txt_environment_xpath))
                    env.select_by_visible_text(environment)
                    time.sleep(5)
                    if CUType == "VCenter":
                        self.setVCenter(sheet, r)
                else:
                    self.logger.info("********** Select An Environment Pop-up Banner Is Not Opened For Wave, " + str(waveName) + " **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def setVCenter(self, sheet, r):
        time.sleep(5)
        waveName = sheet.cell(row=r, column=1).value
        clusterName = sheet.cell(row=r, column=4).value
        esxHost = sheet.cell(row=r, column=5).value
        datastore = sheet.cell(row=r, column=6).value

        time.sleep(5)
        self.driver.find_element(By.XPATH, self.txt_clusterName_xpath).send_keys(clusterName)
        self.driver.find_element(By.XPATH, self.txt_ESXHost_xpath).send_keys(esxHost)
        self.driver.find_element(By.XPATH, self.txt_Datastore_xpath).send_keys(datastore)
        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_applyChanges_id).click()
        time.sleep(5)
        note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
        self.logger.info("********** Set Autoprovision Status of Wave : " + waveName + ",")
        self.logger.info(note + "\n")
        totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
        if totalHosts == 1:
            self.setNICEdit(sheet, r)
        else:
            self.setNICBulkEdit(sheet, r)

    def setNICEdit(self, sheet, r):
        time.sleep(5)
        waveName = sheet.cell(row=r, column=1).value
        deviceName = sheet.cell(row=r, column=7).value
        Type = sheet.cell(row=r, column=8).value
        networkName = sheet.cell(row=r, column=9).value
        ipType = sheet.cell(row=r, column=10).value
        CIDR = sheet.cell(row=r, column=11).value
        gateway = sheet.cell(row=r, column=12).value
        DNS1 = sheet.cell(row=r, column=13).value
        DNS2 = sheet.cell(row=r, column=14).value

        self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.pop_edit_xpath)) != 0:
            self.logger.info("********** Edit Host Pop-Up Banner Was Opened For Wave : "+waveName+" **********")
            self.driver.find_element(By.LINK_TEXT, "vCenter Options").click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_vCenter_xpath)) != 0:
                self.logger.info("********** vCenter Option Pop-Up Banner Was Opened For Wave : "+waveName+" **********")
                self.driver.find_element(By.XPATH, self.btn_NICAdd_xpath).click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.pop_addNIC_xpath)) != 0:
                    self.logger.info("********** Add NIC Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                    self.driver.find_element(By.XPATH, self.btn_NICAdd_xpath).click()
                    if ipType == "DHCP":
                        self.driver.find_element(By.XPATH, self.rd_DHCP_xpath).click()
                    elif ipType == "Static IP":
                        self.driver.find_element(By.XPATH, self.rd_staticIP_xpath).click()
                        self.driver.find_element(By.ID, self.txt_CIDR_id).send_keys(CIDR)
                        self.driver.find_element(By.ID, self.txt_gateway_id).send_keys(gateway)
                        self.driver.find_element(By.ID, self.txt_DNS1_id).send_keys(DNS1)
                        self.driver.find_element(By.ID, self.txt_DNS2_id).send_keys(DNS2)
                    self.driver.find_element(By.ID, self.txt_deviceName_id).send_keys(deviceName)
                    self.driver.find_element(By.ID, self.txt_type_id).send_keys(Type)
                    self.driver.find_element(By.ID, self.txt_networkName_id).send_keys(networkName)
                    time.sleep(5)
                    self.driver.find_element(By.ID, self.btn_save_id).click()
                    time.sleep(5)
                    if len(self.driver.find_elements(By.XPATH, self.var_addNIC_xpath)) != 0:
                        self.logger.info("********** NIC Added Successfully For Wave : " + waveName + " **********")
                    else:
                        self.logger.info("********** Failed To Add NIC For Wave : " + waveName + " **********")
                else:
                    self.logger.info("********** Add NIC Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                time.sleep(3)
                self.driver.find_element(By.ID, self.btn_modify_id).click()
                time.sleep(5)
                note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                self.logger.info("********** Set NIC Bulk Edit Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
            else:
                self.logger.info("********** vCenter Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
        else:
            self.logger.info("********** Sync Option Wave Pop-Up Banner Was Not Opened For Wave : "+waveName+" **********")

    def setNICBulkEdit(self, sheet, r):
        time.sleep(5)
        waveName = sheet.cell(row=r, column=1).value
        deviceName = sheet.cell(row=r, column=7).value
        Type = sheet.cell(row=r, column=8).value
        networkName = sheet.cell(row=r, column=9).value
        ipType = sheet.cell(row=r, column=10).value
        CIDR = sheet.cell(row=r, column=11).value
        gateway = sheet.cell(row=r, column=12).value
        DNS1 = sheet.cell(row=r, column=13).value
        DNS2 = sheet.cell(row=r, column=14).value

        self.driver.find_element(By.XPATH, self.btn_selectAll_xpath).click()
        time.sleep(5)
        self.driver.find_element(By.XPATH, self.btn_bulkEdit_xpath).click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.pop_bulkEdit_xpath)) != 0:
            self.logger.info("********** Bulk Edit Wave Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
            self.driver.find_element(By.LINK_TEXT, "vCenter Options").click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_bulkVC_xpath)) != 0:
                self.logger.info("********** Bulk Edit vCenter Option Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                self.driver.find_element(By.XPATH, self.btn_NICAddBulkEdit_xpath).click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.pop_bulkNIC_xpath)) != 0:
                    self.logger.info("********** Bulk Edit Add NIC Option Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                    self.driver.find_element(By.XPATH, self.txt_deviceNameBulkEdit_xpath).send_keys(deviceName)
                    self.driver.find_element(By.XPATH, self.txt_typeBulkEdit_xpath).send_keys(Type)
                    self.driver.find_element(By.XPATH, self.txt_networkNameBulkEdit_xpath).send_keys(networkName)
                    if ipType == "DHCP":
                        self.driver.find_element(By.XPATH, self.rd_DHCPBulkEdit_xpath).click()
                    elif ipType == "Static IP":
                        self.driver.find_element(By.XPATH, self.rd_staticIPBulkEdit_xpath).click()
                        self.driver.find_element(By.XPATH, self.txt_CIDRBulkEdit_xpath).send_keys(CIDR)
                        self.driver.find_element(By.XPATH, self.txt_gatewayBulkEdit_xpath).send_keys(gateway)
                        self.driver.find_element(By.XPATH, self.txt_DNS1BulkEdit_xpath).send_keys(DNS1)
                        self.driver.find_element(By.XPATH, self.txt_DNS2BulkEdit_xpath).send_keys(DNS2)
                    time.sleep(5)
                    self.driver.find_element(By.XPATH, self.btn_save_xpath).click()
                    time.sleep(5)
                    if len(self.driver.find_elements(By.XPATH, self.var_bulkNIC_xpath)) != 0:
                        self.logger.info("********** Successfully Added NIC For Wave : " + waveName + " **********")
                    else:
                        self.logger.info("********** Failed To Add NIC For Wave : " + waveName + " **********")
                else:
                    self.logger.info("********** Bulk Edit Add NIC Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                time.sleep(3)
                self.driver.find_element(By.XPATH, self.btn_next_xpath).click()
                time.sleep(3)
                self.driver.find_element(By.XPATH, self.btn_modifyAll_xpath).click()
                time.sleep(3)
                self.driver.find_element(By.ID, self.btn_yes_id).click()
                time.sleep(5)
                note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                self.logger.info("********** Set NIC Bulk Edit Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
            else:
                self.logger.info("********** Bulk Edit vCenter Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
        else:
            self.logger.info("********** Bulk Edit Wave Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")

    def setAWS(self, waveName, environment, vpcID, subnetID):
        time.sleep(5)
        val = self.findWave(waveName)
        if val == 2:
            return
        time.sleep(5)
        self.driver.find_element(By.ID, self.txt_autoprovision_id).click()
        time.sleep(5)
        env = Select(self.driver.find_element(By.XPATH, self.txt_environment_xpath))
        time.sleep(5)
        env.select_by_visible_text(environment)
        time.sleep(5)
        self.driver.find_element(By.ID, self.txt_AWSVPCID_id).send_keys(vpcID)
        self.driver.find_element(By.ID, self.txt_AWSSubnetID_id).send_keys(subnetID)
        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_applyChanges_id).click()
        time.sleep(5)
        if val == 1:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def setOCI(self, waveName, environment, VCNName, SubnetName, AVDomain):
        time.sleep(5)
        val = self.findWave(waveName)
        if val == 2:
            return
        time.sleep(5)
        self.driver.find_element(By.ID, self.txt_autoprovision_id).click()
        time.sleep(5)
        env = Select(self.driver.find_element(By.XPATH, self.txt_environment_xpath))
        time.sleep(5)
        env.select_by_visible_text(environment)
        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_applyChanges_id).click()
        time.sleep(5)
        totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
        if totalHosts == 1:
            self.setOCISync(VCNName, SubnetName, AVDomain)
        else:
            self.setOCIBulkSync(VCNName, SubnetName, AVDomain)

        time.sleep(5)
        if val == 1:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def setOCISync(self, VCNName, SubnetName, AVDomain):
        time.sleep(5)
        self.driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "OCI Options").click()
        time.sleep(5)
        self.driver.find_element(By.XPATH, self.txt_OCIVCNName_xpath).send_keys(VCNName)
        self.driver.find_element(By.XPATH, self.txt_OCISubnetName_xpath).send_keys(SubnetName)
        self.driver.find_element(By.XPATH, self.txt_OCIAVDomain_xpath).send_keys(AVDomain)
        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_modify_id).click()

    def setOCIBulkSync(self, VCNName, SubnetName, AVDomain):
        time.sleep(5)
        self.driver.find_element(By.XPATH, self.btn_selectAll_xpath).click()
        time.sleep(5)
        self.driver.find_element(By.XPATH, self.btn_bulkEdit_xpath).click()
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "OCI Options").click()
        time.sleep(5)
        self.driver.find_element(By.ID, "oci_vcn_name").send_keys(VCNName)
        self.driver.find_element(By.ID, "oci_subnet_name").send_keys(SubnetName)
        self.driver.find_element(By.ID, "oci_av_domain").send_keys(AVDomain)
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.btn_next_xpath).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.btn_modifyAll_xpath).click()
        time.sleep(3)
        self.driver.find_element(By.ID, self.btn_yes_id).click()

    def setSyncOptions(self, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 3
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = rows
        time.sleep(5)
        tmp = "None"
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            hostName = sheet.cell(row=r, column=2).value

            tng = sheet.cell(row=r, column=3).value
            verbose = sheet.cell(row=r, column=4).value
            directFScopy = sheet.cell(row=r, column=5).value
            FSDeletion = sheet.cell(row=r, column=6).value
            NoTransfer = sheet.cell(row=r, column=7).value
            transferCompress = sheet.cell(row=r, column=8).value
            noTransferCompress = sheet.cell(row=r, column=9).value
            ignoreMissing = sheet.cell(row=r, column=10).value
            noInPlace = sheet.cell(row=r, column=11).value
            noReboot = sheet.cell(row=r, column=12).value
            includeSAN = sheet.cell(row=r, column=13).value
            excludeSAN = sheet.cell(row=r, column=14).value
            overrideRMMStorageCheck = sheet.cell(row=r, column=15).value
            deleteAllTargetFS = sheet.cell(row=r, column=16).value
            keepTargetLayout = sheet.cell(row=r, column=17).value
            cloudInit = sheet.cell(row=r, column=18).value
            eventScript = sheet.cell(row=r, column=19).value
            eventScriptArgs = sheet.cell(row=r, column=20).value
            excludeFileSystem = sheet.cell(row=r, column=21).value
            includeFileSystem = sheet.cell(row=r, column=22).value
            filterFileOption = sheet.cell(row=r, column=23).value
            filterFilePath = sheet.cell(row=r, column=24).value

            time.sleep(5)
            if tmp != waveName:
                tmp = waveName
                val = self.findWave(waveName)
                if val == 2:
                    continue
                time.sleep(5)
                self.driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.var_waveDetails_xpath)) != 0:
                    self.logger.info("********** Wave " + waveName + " Was Opened **********")
                else:
                    self.logger.info("********** Wave " + waveName + " Was Not Opened **********")
                    continue
            time.sleep(5)
            totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
            if totalHosts == 1:
                self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
            else:
                count = 1
                for hostNo in range(1, totalHosts+1):
                    if totalHosts == 1:
                        host = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                    else:
                        host = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                    if hostName == host:
                        break
                    count += 1
                time.sleep(5)
                if count == totalHosts+1:
                    self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
                    continue
                else:
                    self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(count)+']/td[8]/span/div/i[2]').click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_edit_xpath)) != 0:
                self.logger.info("********** Edit Host Pop-Up Banner Was Opened For Host : " + hostName + " **********")
                time.sleep(5)
                syncOption = WebDriverWait(self.driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, self.btn_syncOptions_xpath))
                )
                syncOption.click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.pop_syncOption_xpath)) != 0:
                    self.logger.info("********** Sync Options Pop-up Banner Is Opened For Host : " + hostName + " **********")
                    if tng == "Yes":
                        self.driver.find_element(By.ID, self.rd_tng_id).click()
                        self.logger.info("********** Sync Option 'TNG' is set **********")
                    if verbose == "Yes":
                        self.driver.find_element(By.ID, self.rd_verbose_id).click()
                        self.logger.info("********** Sync Option 'Verbose' is set **********")
                    if directFScopy == "Yes":
                        self.driver.find_element(By.ID, self.rd_directFScopy_id).click()
                        self.logger.info("********** Sync Option 'Direct FScopy' is set **********")
                    if FSDeletion == "Yes":
                        self.driver.find_element(By.ID, self.rd_FSDeletion_id).click()
                        self.logger.info("********** Sync Option 'FS Deletion' is set **********")
                    if NoTransfer == "Yes":
                        self.driver.find_element(By.ID, self.rd_NoTransfer_id).click()
                        self.logger.info("********** Sync Option 'No Transfer' is set **********")
                    if transferCompress == "Yes":
                        self.driver.find_element(By.ID, self.rd_transferCompress_id).click()
                        self.logger.info("********** Sync Option 'Transfer Compress' is set **********")
                    if noTransferCompress == "Yes":
                        self.driver.find_element(By.ID, self.rd_noTransferCompress_id).click()
                        self.logger.info("********** Sync Option 'No Transfer Compress' is set **********")
                    if ignoreMissing == "Yes":
                        self.driver.find_element(By.ID, self.rd_ignoreMissing_id).click()
                        self.logger.info("********** Sync Option 'Ignore Missing' is set **********")
                    if noInPlace == "Yes":
                        self.driver.find_element(By.ID, self.rd_noInPlace_id).click()
                        self.logger.info("********** Sync Option 'No In Place' is set **********")
                    if noReboot == "Yes":
                        self.driver.find_element(By.ID, self.rd_noReboot_id).click()
                        self.logger.info("********** Sync Option 'No Reboot' is set **********")
                    if includeSAN == "Yes":
                        self.driver.find_element(By.ID, self.rd_includeSAN_id).click()
                        self.logger.info("********** Sync Option 'Include SAN' is set **********")
                    if excludeSAN == "Yes":
                        self.driver.find_element(By.ID, self.rd_excludeSAN_id).click()
                        self.logger.info("********** Sync Option 'Exclude SAN' is set **********")
                    if overrideRMMStorageCheck == "Yes":
                        self.driver.find_element(By.ID, self.rd_overrideRMMStorageCheck_id).click()
                        self.logger.info("********** Sync Option 'Override RMM Storage' Check is set **********")
                    if deleteAllTargetFS == "Yes":
                        self.driver.find_element(By.ID, self.rd_deleteAllTargetFS_id).click()
                        self.logger.info("********** Sync Option 'Delete All Target' FS is set **********")
                    if keepTargetLayout == "Yes":
                        self.driver.find_element(By.ID, self.rd_keepTargetLayout_id).click()
                        self.logger.info("********** Sync Option 'Keep Target Layout' is set **********")
                    if cloudInit == "Yes":
                        self.driver.find_element(By.ID, self.rd_cloudInit_id).click()
                        self.logger.info("********** Sync Option 'Cloud Init' is set **********")

                    if eventScript != "NA":
                        self.driver.find_element(By.ID, self.txt_eventScript_id).clear()
                        self.driver.find_element(By.ID, self.txt_eventScript_id).send_keys(eventScript)
                        self.logger.info("********** Sync Option 'Event Script' is set **********")

                    if eventScriptArgs != "NA":
                        self.driver.find_element(By.ID, self.txt_eventScriptArgs_id).clear()
                        self.driver.find_element(By.ID, self.txt_eventScriptArgs_id).send_keys(eventScriptArgs)
                        self.logger.info("********** Sync Option 'Event Script Arguments' is set **********")

                    if excludeFileSystem != "NA":
                        self.driver.find_element(By.ID, self.txt_excludeFile_id).clear()
                        self.driver.find_element(By.ID, self.txt_excludeFile_id).send_keys(excludeFileSystem)
                        self.logger.info("********** Sync Option 'Exclude File System(s)' is set **********")

                    if includeFileSystem != "NA":
                        self.driver.find_element(By.ID, self.txt_includeFile_id).clear()
                        self.driver.find_element(By.ID, self.txt_includeFile_id).send_keys(includeFileSystem)
                        self.logger.info("********** Sync Option 'Include File System(s)' is set **********")

                    if filterFileOption != "NA":
                        if filterFileOption == "Upload Local File":
                            self.driver.find_element(By.XPATH, self.rd_uploadLocalFile_xpath).click()
                            self.driver.find_element(By.XPATH, self.btn_uploadLocalFile_xpath).click()
                            time.sleep(5)
                            keyboard.write(filterFilePath)
                            time.sleep(5)
                            keyboard.send('enter')
                        elif filterFileOption == "File Path On RMM":
                            self.driver.find_element(By.XPATH, self.rd_filePathOnRmm_xpath).click()
                            self.driver.find_element(By.XPATH, self.txt_filePathOnRmm_xpath).send_keys(filterFilePath)
                        self.logger.info("********** Sync Option 'Filter File' is set **********")

                    self.driver.find_element(By.ID, self.btn_modify_id).click()
                    time.sleep(5)
                    note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                    self.logger.info("********** Set Edit Sync Options Status of Wave : " + waveName + ",")
                    self.logger.info(note + "\n")
                else:
                    self.logger.info("********** Sync Options Pop-up Banner Is Not Opened For Host : " + hostName + " **********")
            else:
                self.logger.info("********** Edit Host Pop-Up Banner Was Not Opened For Host : " + hostName + " **********")
                self.driver.find_element(By.XPATH, self.btn_cancelEdit_xpath).click()
            time.sleep(5)
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def bulkEditSyncOption(self, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        time.sleep(5)
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            hostNames = sheet.cell(row=r, column=2).value
            tng = sheet.cell(row=r, column=3).value
            verbose = sheet.cell(row=r, column=4).value
            passwordLess = sheet.cell(row=r, column=5).value
            directFScopy = sheet.cell(row=r, column=6).value
            FSDeletion = sheet.cell(row=r, column=7).value
            NoTransfer = sheet.cell(row=r, column=8).value
            transferCompress = sheet.cell(row=r, column=9).value
            noTransferCompress = sheet.cell(row=r, column=10).value
            ignoreMissing = sheet.cell(row=r, column=11).value
            noInPlace = sheet.cell(row=r, column=12).value
            noReboot = sheet.cell(row=r, column=13).value
            includeSAN = sheet.cell(row=r, column=14).value
            excludeSAN = sheet.cell(row=r, column=15).value
            overrideRMMStorageCheck = sheet.cell(row=r, column=16).value
            deleteAllTargetFS = sheet.cell(row=r, column=17).value
            keepTargetLayout = sheet.cell(row=r, column=18).value
            cloudInit = sheet.cell(row=r, column=19).value
            excludeFile = sheet.cell(row=r, column=20).value
            includeFile = sheet.cell(row=r, column=21).value

            if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                val = self.findWave(waveName)
                if val == 2:
                    return
                time.sleep(5)
            ele = WebDriverWait(self.driver, 30).until(
                EC.element_to_be_clickable((By.LINK_TEXT, waveName))
            )
            ele.click()
            if len(self.driver.find_elements(By.XPATH, self.var_waveDetails_xpath)) != 0:
                self.logger.info("********** Wave " + waveName + " Was Opened **********")
            else:
                self.logger.info("********** Wave " + waveName + " Was Not Opened **********")
                continue
            time.sleep(5)
            if type(hostNames) != str:
                self.driver.find_element(By.XPATH, self.btn_selectAll_xpath).click()
            else:
                res = tuple(map(str, hostNames.split(', ')))
                totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
                for hostName in res:
                    count = 1
                    for hostNo in range(1, totalHosts+1):
                        if totalHosts == 1:
                            tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                        else:
                            tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                        if hostName == tmp:
                            break
                        elif hostNo == totalHosts:
                            self.logger.info("********** The Host " + str(hostName) + " Was Not Found In The Wave : " + waveName + " **********")
                        count += 1
                    self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(count)+']/td[1]/p-tablecheckbox/div/div[2]').click()
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_bulkEdit_xpath).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_bulkEdit_xpath)) != 0:
                self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Is Opened For Wave, " + waveName + " **********")
                time.sleep(5)
                if tng == "Yes":
                    self.driver.find_element(By.ID, "tng_yes").click()
                    self.logger.info("********** Sync Option 'TNG' is set as yes **********")
                if tng == "No":
                    self.driver.find_element(By.ID, "tng_no").click()
                    self.logger.info("********** Sync Option 'TNG' is set as no **********")

                if verbose == "Yes":
                    self.driver.find_element(By.ID, "verbose_yes").click()
                    self.logger.info("********** Sync Option 'Verbose' is set as yes **********")
                if verbose == "No":
                    self.driver.find_element(By.ID, "verbose_no").click()
                    self.logger.info("********** Sync Option 'Verbose' is set as no **********")

                if passwordLess == "Yes":
                    self.driver.find_element(By.ID, "ssh_yes").click()
                    self.logger.info("********** Sync Option 'Passwordless' is set as yes **********")
                if passwordLess == "No":
                    self.driver.find_element(By.ID, "ssh_no").click()
                    self.logger.info("********** Sync Option 'Passwordless' is set as no **********")

                if directFScopy == "Yes":
                    self.driver.find_element(By.ID, "allowDirectFscopy_yes").click()
                    self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as yes **********")
                if directFScopy == "No":
                    self.driver.find_element(By.ID, "allowDirectFscopy_no").click()
                    self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as no **********")

                if FSDeletion == "Yes":
                    self.driver.find_element(By.ID, "allowFsDeletion_yes").click()
                    self.logger.info("********** Sync Option 'Allow FS Deletion' is set as yes **********")
                if FSDeletion == "No":
                    self.driver.find_element(By.ID, "allowFsDeletion_no").click()
                    self.logger.info("********** Sync Option 'Allow FS Deletion' is set as no **********")

                if NoTransfer == "Yes":
                    self.driver.find_element(By.ID, "no_xfer_yes").click()
                    self.logger.info("********** Sync Option 'No Transfer' is set as yes **********")
                if NoTransfer == "No":
                    self.driver.find_element(By.ID, "no_xfer_no").click()
                    self.logger.info("********** Sync Option 'No Transfer' is set as no **********")

                if transferCompress == "Yes":
                    self.driver.find_element(By.ID, "xferCompress_yes").click()
                    self.logger.info("********** Sync Option 'Transfer Compress' is set as yes **********")
                if transferCompress == "No":
                    self.driver.find_element(By.ID, "xferCompress_no").click()
                    self.logger.info("********** Sync Option 'Transfer Compress' is set as no **********")

                if noTransferCompress == "Yes":
                    self.driver.find_element(By.ID, "no_xfer_compress_yes").click()
                    self.logger.info("********** Sync Option 'No Transfer Compress' is set as yes **********")
                if noTransferCompress == "No":
                    self.driver.find_element(By.ID, "no_xfer_compress_no").click()
                    self.logger.info("********** Sync Option 'No Transfer Compress' is set as no **********")

                if ignoreMissing == "Yes":
                    self.driver.find_element(By.ID, "ignoreMissing_yes").click()
                    self.logger.info("********** Sync Option 'Ignore Missing' is set as yes **********")
                if ignoreMissing == "No":
                    self.driver.find_element(By.ID, "ignoreMissing_no").click()
                    self.logger.info("********** Sync Option 'Ignore Missing' is set as no **********")

                if noInPlace == "Yes":
                    self.driver.find_element(By.ID, "noInPlace_yes").click()
                    self.logger.info("********** Sync Option 'No In Place' is set as yes **********")
                if noInPlace == "No":
                    self.driver.find_element(By.ID, "noInPlace_no").click()
                    self.logger.info("********** Sync Option 'No In Place' is set as no **********")

                if noReboot == "Yes":
                    self.driver.find_element(By.ID, "noReboot_yes").click()
                    self.logger.info("********** Sync Option 'No Reboot' is set as yes **********")
                if noReboot == "No":
                    self.driver.find_element(By.ID, "noReboot_no").click()
                    self.logger.info("********** Sync Option 'No Reboot' is set as no **********")

                if includeSAN == "Yes":
                    self.driver.find_element(By.ID, "include_san_yes").click()
                    self.logger.info("********** Sync Option 'Include SAN' is set as yes **********")
                if includeSAN == "No":
                    self.driver.find_element(By.ID, "include_san_no").click()
                    self.logger.info("********** Sync Option 'Include SAN' is set as no **********")

                if excludeSAN == "Yes":
                    self.driver.find_element(By.ID, "exclude_san_yes").click()
                    self.logger.info("********** Sync Option 'Exclude SAN' is set as yes **********")
                if excludeSAN == "No":
                    self.driver.find_element(By.ID, "exclude_san_no").click()
                    self.logger.info("********** Sync Option 'Exclude SAN' is set as no **********")

                if overrideRMMStorageCheck == "Yes":
                    self.driver.find_element(By.ID, "storage_override_yes").click()
                    self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as yes **********")
                if overrideRMMStorageCheck == "No":
                    self.driver.find_element(By.ID, "storage_override_no").click()
                    self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as no **********")

                if deleteAllTargetFS == "Yes":
                    self.driver.find_element(By.ID, "delete_all_target_fs_yes").click()
                    self.logger.info("********** Sync Option 'Delete All Target FS' is set as yes **********")
                if deleteAllTargetFS == "No":
                    self.driver.find_element(By.ID, "delete_all_target_fs_no").click()
                    self.logger.info("********** Sync Option 'Delete All Target FS' is set as no **********")

                if keepTargetLayout == "Yes":
                    self.driver.find_element(By.ID, "keep_target_layout_yes").click()
                    self.logger.info("********** Sync Option 'Keep Target Layout' is set as yes **********")
                if keepTargetLayout == "No":
                    self.driver.find_element(By.ID, "keep_target_layout_no").click()
                    self.logger.info("********** Sync Option 'Keep Target Layout' is set as no **********")

                if cloudInit == "Yes":
                    self.driver.find_element(By.ID, "cloud_init_yes").click()
                    self.logger.info("********** Sync Option 'Cloud Init' is set as yes **********")
                if cloudInit == "No":
                    self.driver.find_element(By.ID, "cloud_init_no").click()
                    self.logger.info("********** Sync Option 'Cloud Init' is set as no **********")

                if excludeFile != "NA":
                    self.driver.find_element(By.ID, self.txt_blkExcludeFile_id).clear()
                    self.driver.find_element(By.ID, self.txt_blkExcludeFile_id).send_keys(excludeFile)
                    self.logger.info("********** Sync Option 'Exclude File System(s)' is set as no **********")

                if includeFile != "NA":
                    self.driver.find_element(By.ID, self.txt_blkIncludeFile_id).clear()
                    self.driver.find_element(By.ID, self.txt_blkIncludeFile_id).send_keys(includeFile)
                    self.logger.info("********** Sync Option 'Exclude File System(s)' is set as no **********")

                time.sleep(5)
                self.driver.find_element(By.XPATH, self.btn_next_xpath).click()
                time.sleep(5)
                self.driver.find_element(By.XPATH, self.btn_modifyAll_xpath).click()
                time.sleep(5)
                self.driver.find_element(By.ID, self.btn_yes_id).click()
                time.sleep(5)
                note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                self.logger.info("********** Set Bulk Edit Sync Options Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
            else:
                self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Is Not Opened For Wave, " + waveName + " **********")
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def changeBulkEditOption(self, waveName, option, yesOrNo):
        if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
            val = self.findWave(waveName)
            if val == 2:
                return
            time.sleep(5)
        ele = WebDriverWait(self.driver, 30).until(
            EC.element_to_be_clickable((By.LINK_TEXT, waveName))
        )
        ele.click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.var_waveDetails_xpath)) != 0:
            self.logger.info("********** Wave " + waveName + " Was Opened **********")
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_selectAll_xpath).click()
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_bulkEdit_xpath).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_bulkEdit_xpath)) != 0:
                self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Is Opened For Wave, " + waveName + " **********")
                time.sleep(5)
                if option == "TNG":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "tng_yes").click()
                        self.logger.info("********** Sync Option 'TNG' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "tng_no").click()
                        self.logger.info("********** Sync Option 'TNG' is set as no **********")

                elif option == "Verbose":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "verbose_yes").click()
                        self.logger.info("********** Sync Option 'Verbose' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "verbose_no").click()
                        self.logger.info("********** Sync Option 'Verbose' is set as no **********")

                elif option == "Passwordless":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "ssh_yes").click()
                        self.logger.info("********** Sync Option 'Passwordless' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "ssh_no").click()
                        self.logger.info("********** Sync Option 'Passwordless' is set as no **********")

                elif option == "Allow Direct Fscopy":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "allowDirectFscopy_yes").click()
                        self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "allowDirectFscopy_no").click()
                        self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as no **********")

                elif option == "Allow FS Deletion":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "allowFsDeletion_yes").click()
                        self.logger.info("********** Sync Option 'Allow FS Deletion' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "allowFsDeletion_no").click()
                        self.logger.info("********** Sync Option 'Allow FS Deletion' is set as no **********")

                elif option == "No Transfer":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "no_xfer_yes").click()
                        self.logger.info("********** Sync Option 'No Transfer' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "no_xfer_no").click()
                        self.logger.info("********** Sync Option 'No Transfer' is set as no **********")

                elif option == "Transfer Compress":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "xferCompress_yes").click()
                        self.logger.info("********** Sync Option 'Transfer Compress' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "xferCompress_no").click()
                        self.logger.info("********** Sync Option 'Transfer Compress' is set as no **********")

                elif option == "No Transfer Compress":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "no_xfer_compress_yes").click()
                        self.logger.info("********** Sync Option 'No Transfer Compress' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "no_xfer_compress_no").click()
                        self.logger.info("********** Sync Option 'No Transfer Compress' is set as no **********")

                elif option == "Ignore Missing":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "ignoreMissing_yes").click()
                        self.logger.info("********** Sync Option 'Ignore Missing' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "ignoreMissing_no").click()
                        self.logger.info("********** Sync Option 'Ignore Missing' is set as no **********")

                elif option == "No In Place":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "noInPlace_yes").click()
                        self.logger.info("********** Sync Option 'No In Place' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "noInPlace_no").click()
                        self.logger.info("********** Sync Option 'No In Place' is set as no **********")

                elif option == "No Reboot":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "noReboot_yes").click()
                        self.logger.info("********** Sync Option 'No Reboot' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "noReboot_no").click()
                        self.logger.info("********** Sync Option 'No Reboot' is set as no **********")

                elif option == "Include SAN":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "include_san_yes").click()
                        self.logger.info("********** Sync Option 'Include SAN' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "include_san_no").click()
                        self.logger.info("********** Sync Option 'Include SAN' is set as no **********")

                elif option == "Exclude SAN":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "exclude_san_yes").click()
                        self.logger.info("********** Sync Option 'Exclude SAN' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "exclude_san_no").click()
                        self.logger.info("********** Sync Option 'Exclude SAN' is set as no **********")

                elif option == "Override RMM Storage Check":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "storage_override_yes").click()
                        self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "storage_override_no").click()
                        self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as no **********")

                elif option == "Delete All Target FS":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "delete_all_target_fs_yes").click()
                        self.logger.info("********** Sync Option 'Delete All Target FS' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "delete_all_target_fs_no").click()
                        self.logger.info("********** Sync Option 'Delete All Target FS' is set as no **********")

                elif option == "Keep Target Layout":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "keep_target_layout_yes").click()
                        self.logger.info("********** Sync Option 'Keep Target Layout' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "keep_target_layout_no").click()
                        self.logger.info("********** Sync Option 'Keep Target Layout' is set as no **********")

                elif option == "Cloud Init":
                    if yesOrNo == "Yes":
                        self.driver.find_element(By.ID, "cloud_init_yes").click()
                        self.logger.info("********** Sync Option 'Cloud Init' is set as yes **********")
                    if yesOrNo == "No":
                        self.driver.find_element(By.ID, "cloud_init_no").click()
                        self.logger.info("********** Sync Option 'Cloud Init' is set as no **********")

                time.sleep(5)
                self.driver.find_element(By.XPATH, self.btn_next_xpath).click()
                time.sleep(5)
                self.driver.find_element(By.XPATH, self.btn_modifyAll_xpath).click()
                time.sleep(5)
                self.driver.find_element(By.ID, self.btn_yes_id).click()
                time.sleep(5)
                note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                self.logger.info("********** Set Bulk Edit Sync Options Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
            else:
                self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Was Not Opened For Wave, " + waveName + " **********")
        else:
            self.logger.info("********** Wave " + waveName + " Was Not Opened **********")
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def changeTargetType(self, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        tmp = "none"

        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = rows

        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            hostName = sheet.cell(row=r, column=2).value
            targetType = sheet.cell(row=r, column=3).value

            if tmp != waveName:
                time.sleep(5)
                if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                    val = self.findWave(waveName)
                    if val == 2:
                        return
                    time.sleep(5)
                self.driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.var_waveDetails_xpath)) != 0:
                    self.logger.info("********** Wave " + waveName + " Was Opened **********")
                else:
                    self.logger.info("********** Wave " + waveName + " Was Not Opened **********")
                    continue
            time.sleep(5)
            totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
            count = 1
            for hostNo in range(1, totalHosts + 1):
                if totalHosts == 1:
                    tmp = self.driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                else:
                    tmp = self.driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                if hostName == tmp:
                    break
                count += 1
            if count == totalHosts + 1:
                self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
            else:
                self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(count)+']/td[8]/span/div/i[2]').click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.pop_system_xpath)) != 0:
                    self.logger.info("********** Wave System Options Pop-up Banner Is Opened For Wave, " + waveName + " **********")

                    if targetType == "Autoprovision":
                        self.driver.find_element(By.ID, self.rd_autoprovision_id).click()

                    elif targetType == "Existing System":
                        imageName = sheet.cell(row=r, column=4).value
                        syncType = sheet.cell(row=r, column=5).value
                        passthrough = sheet.cell(row=r, column=6).value
                        targetIP = sheet.cell(row=r, column=7).value
                        friendlyName = sheet.cell(row=r, column=8).value
                        userName = sheet.cell(row=r, column=9).value

                        self.driver.find_element(By.ID, self.rd_existingSystem_id).click()
                        time.sleep(5)
                        if syncType == "Direct Sync":
                            self.driver.find_element(By.ID, self.rd_directSync_id).click()
                            time.sleep(5)
                            tmp = self.driver.find_element(By.ID, self.ch_passthrough_id).is_selected()
                            if passthrough and not tmp:
                                self.driver.find_element(By.ID, self.ch_passthrough_id).click()
                            elif not passthrough and tmp:
                                self.driver.find_element(By.ID, self.ch_passthrough_id).click()
                            self.driver.find_element(By.ID, self.txt_targetIP_id).clear()
                            self.driver.find_element(By.ID, self.txt_targetIP_id).send_keys(targetIP)
                            self.driver.find_element(By.ID, self.txt_friendlyName_id).clear()
                            self.driver.find_element(By.ID, self.txt_friendlyName_id).send_keys(friendlyName)

                        elif syncType == "Stage 1 & 2" or syncType == "Stage 2":
                            if syncType == "Stage 1 & 2":
                                self.driver.find_element(By.ID, self.rd_stage12_id).click()
                            elif syncType == "Stage 2":
                                self.driver.find_element(By.ID, self.rd_stage2_id).click()
                            time.sleep(5)
                            self.driver.find_element(By.ID, self.txt_imageName_id).clear()
                            self.driver.find_element(By.ID, self.txt_imageName_id).send_keys(imageName)
                            time.sleep(5)
                            tmp = self.driver.find_element(By.ID, self.ch_passthrough_id).is_selected()
                            if passthrough and not tmp:
                                self.driver.find_element(By.ID, self.ch_passthrough_id).click()
                            elif not passthrough and tmp:
                                self.driver.find_element(By.ID, self.ch_passthrough_id).click()
                            self.driver.find_element(By.ID, self.txt_targetIP_id).clear()
                            self.driver.find_element(By.ID, self.txt_targetIP_id).send_keys(targetIP)
                            self.driver.find_element(By.ID, self.txt_friendlyName_id).clear()
                            self.driver.find_element(By.ID, self.txt_friendlyName_id).send_keys(friendlyName)

                        elif syncType == "Stage 1":
                            self.driver.find_element(By.ID, self.rd_stage1_id).click()
                            time.sleep(5)
                            self.driver.find_element(By.ID, self.txt_imageName_id).clear()
                            self.driver.find_element(By.ID, self.txt_imageName_id).send_keys(imageName)

                        self.driver.find_element(By.ID, self.txt_userName_id).clear()
                        self.driver.find_element(By.ID, self.txt_userName_id).send_keys(userName)
                        time.sleep(5)

                    elif targetType == "Capture":
                        imageName = sheet.cell(row=r, column=4).value
                        self.driver.find_element(By.ID, self.rd_capture_id).click()
                        self.driver.find_element(By.XPATH, self.txt_captureImage_xpath).clear()
                        self.driver.find_element(By.XPATH, self.txt_captureImage_xpath).send_keys(imageName)

                    time.sleep(5)
                    self.driver.find_element(By.ID, self.btn_modify_id).click()
                    time.sleep(5)
                    note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                    self.logger.info("********** Changed Target Type For Host : " + hostName + ",")
                    self.logger.info(note + "\n")
                    time.sleep(5)
                else:
                    self.logger.info("********** Wave System Options Pop-up Banner Is Not Opened For Wave, " + waveName + " **********")
            tmp = waveName
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def changeDatastore(self, waveName, datastore):
        val = self.findWave(waveName)
        if val == 2:
            return
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, waveName).click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.var_waveDetails_xpath)) != 0:
            self.logger.info("********** Wave " + waveName + " Was Opened **********")
            totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
            for hostNo in range(1, totalHosts + 1):
                if totalHosts == 1:
                    hostName = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                    self.driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
                else:
                    time.sleep(5)
                    hostName = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(hostNo)+']/td[3]/span/span').text
                    self.driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(hostNo)+']/td[8]/span/div/i[2]').click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.pop_edit_xpath)) != 0:
                    self.logger.info("********** Edit Host Pop-Up Banner Was Opened For Host : " + hostName + " **********")
                    self.driver.find_element(By.LINK_TEXT, "vCenter Options").click()
                    time.sleep(5)
                    if len(self.driver.find_elements(By.XPATH, self.pop_vCenter_xpath)) != 0:
                        self.logger.info("********** Wave Edit vCenter Options Pop-up Banner Is Opened For Wave, " + waveName + " **********")
                        self.driver.find_element(By.XPATH, self.txt_editDatastore_xpath).clear()
                        self.driver.find_element(By.XPATH, self.txt_editDatastore_xpath).send_keys(datastore)
                        time.sleep(5)
                        self.driver.find_element(By.ID, self.btn_modify_id).click()
                        self.logger.info("********** Successfully Changes Datastore for Host Number : " + str(hostNo) + " **********")
                    else:
                        self.logger.info("********** Wave Edit vCenter Options Pop-up Banner Is Not Open For Wave, " + waveName + " **********")
                else:
                    self.logger.info("********** Edit Host Pop-Up Banner Was Not Opened For Host : " + hostName + " **********")
            time.sleep(5)
        else:
            self.logger.info("********** Wave " + waveName + " Was Not Opened **********")
        if val == 1:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def moveHosts(self, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = rows
        for r in range(st, ed+1):
            sourceWave = sheet.cell(row=r, column=1).value
            targetWave = sheet.cell(row=r, column=2).value
            hostNames = sheet.cell(row=r, column=3).value
            time.sleep(5)
            if len(self.driver.find_elements(By.LINK_TEXT, sourceWave)) == 0:
                val = self.findWave(sourceWave)
                if val == 2:
                    return
                time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, sourceWave).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.var_waveDetails_xpath)) != 0:
                self.logger.info("********** Wave " + sourceWave + " Was Opened **********")
                res = tuple(map(str, hostNames.split(', ')))
                totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
                count = 1
                for hostName in res:
                    time.sleep(5)
                    for hostNo in range(1, totalHosts + 1):
                        if totalHosts == 1:
                            tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                        else:
                            tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                        if hostName == tmp:
                            break
                        elif hostNo == totalHosts:
                            self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + sourceWave + " **********")
                        count += 1
                    self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(count)+']/td[8]/span/div/i[3]').click()
                    time.sleep(5)
                    if len(self.driver.find_elements(By.XPATH, self.pop_moveHost_xpath)) != 0:
                        self.logger.info("********** Move Machine To New Wave Pop-up Banner Was Opened For Host : " + hostName + " **********")
                        env = Select(self.driver.find_element(By.XPATH, self.drp_selectWave_xpath))
                        env.select_by_visible_text(targetWave)
                        time.sleep(5)
                        self.driver.find_element(By.ID, self.btn_moveMachine_id).click()
                        time.sleep(5)
                        note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                        self.logger.info("********** Changed Target Type For Host : " + hostName + ",")
                        self.logger.info(note + "\n")
                    else:
                        self.logger.info("********** Move Machine To New Wave Pop-up Banner Was Not Opened For Host : " + hostName + " **********")
            else:
                self.logger.info("********** Wave " + sourceWave + " Was Not Opened **********")
            time.sleep(5)
            if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
                self.driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
        time.sleep(5)
