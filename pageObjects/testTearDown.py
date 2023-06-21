import time
import unittest
import paramiko
import openpyxl

import Locators.locTearDown as LOC

from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from utilities.customLogger import LogGen
from utilities.commonObjects import CommonObjects


class TearDown(unittest.TestCase):

    logger = LogGen.loggen()

    def tearDown1(self, driver, count, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 3
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        time.sleep(5)
        cnt = 1
        for r in range(st, ed+1):
            operation = sheet.cell(row=r, column=1).value.lower()

            if operation.find("delete") != -1 and operation.find("sync") != -1 and operation.find("relation") != -1:
                source = sheet.cell(row=r, column=5).value
                target = sheet.cell(row=r, column=6).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Sync Relation **********")
                self.deleteSR(source, target)
                self.logger.info("********** Successfully Executed TestCase: Delete Sync Relation **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("host") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=2).value
                hostNames = sheet.cell(row=r, column=3).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Host In Wave **********")
                self.deleteHostInWave(driver, waveName, hostNames)
                self.logger.info("********** Successfully Executed TestCase: Delete Host In Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("replication") != -1 and (operation.find("wave") != -1 or operation.find("waves") != -1):
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Replication Wave **********")
                self.deleteRepWaves(driver, waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Replication Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("DR") != -1 and (operation.find("wave") != -1 or operation.find("waves") != -1):
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete DR Wave **********")
                self.deleteDrWaves(driver, waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete DR Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Wave **********")
                self.deleteWave(driver, waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("dr") != -1 and (operation.find("policy") != -1 or operation.find("policies") != -1):
                policyName = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete DR Policy **********")
                self.deleteDRPolicy(driver, policyName)
                self.logger.info("********** Successfully Executed TestCase: Delete DR Policy **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("source") != -1 and (operation.find("host") != -1 or operation.find("hosts") != -1):
                hostNames = sheet.cell(row=r, column=3).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Source Host **********")
                self.deleteSourceHost(driver, hostNames)
                self.logger.info("********** Successfully Executed TestCase: Delete Source Host **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("target") != -1 and (operation.find("host") != -1 or operation.find("hosts") != -1):
                hostNames = sheet.cell(row=r, column=3).value
                VM = sheet.cell(row=r, column=7).value
                environment = sheet.cell(row=r, column=8).value
                name = sheet.cell(row=r, column=9).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Target Host **********")
                self.deleteTargetHost(driver, hostNames, VM, environment, name)
                self.logger.info("********** Successfully Executed TestCase: Delete Target Host **********\n \n")
                cnt += 1

            else:
                self.logger.info("********** TestCase "+str(count)+"."+str(cnt)+": There Are Some Mistakes In The Operation Keywords '" + str(operation) + "' ... Please Check Once **********\n \n")
                cnt += 1

    def deleteSR(self, source, target):
        vm_ip = "172.29.30.127"
        vm_username = "root"
        vm_password = "rackware"

        ssh = paramiko.SSHClient()
        ssh.load_system_host_keys()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(vm_ip,
                    username=vm_username,
                    password=vm_password,
                    look_for_keys=False)
        stdin, stdout, stderr = ssh.exec_command("rw ic srd " + str(source) + " --target " + str(target))
        output = stdout.read().decode()
        self.logger.info("********** Delete Sync Relation Status: \n" + output + "\n")
        ssh.close()

    def deleteHostInWave(self, driver, waveName, hostNames):
        co = CommonObjects(driver)
        val = co.findWave(waveName)
        if val == 2:
            return
        if val == 1:
            driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                return
        if val == 0:
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
        totalHosts = len(driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
        res = tuple(map(str, hostNames.split(', ')))
        for hostName in res:
            if totalHosts == 1:
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                if tmp == hostName:
                    driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[4]').click()
                    self.hostDeleteState(driver, hostName)
                else:
                    self.logger.info("********** There Was No Host With Name : " + hostName + ", In Wave : " + waveName + " **********")
            else:
                count = 1
                for i in range(1, totalHosts + 1):
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[3]/span/span').text
                    if tmp == hostName:
                        driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(count) + ']/td[8]/span/div/i[4]').click()
                        self.hostDeleteState(driver, hostName)
                        break
                    count += 1
                if count == totalHosts + 1:
                    self.logger.info("********** There Was No Host With Name : " + hostName + ", In Wave : " + waveName + " **********")

    def hostDeleteState(self, driver, hostName):
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, LOC.btn_deleteWaveHost_id))
        )
        driver.find_element(By.ID, LOC.btn_deleteWaveHost_id).click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
        )
        note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
        self.logger.info("********** Delete Status For Host : " + hostName + ",")
        self.logger.info(note + "\n")
        time.sleep(2)
        driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()

    def deleteWave(self, driver, waveName):
        co = CommonObjects(driver)
        val = co.findWave(waveName)
        if val == 2:
            return
        if val == 0:
            replication_class = driver.find_element(By.XPATH, LOC.txt_replication_xpath).get_attribute("class")
            if replication_class == "ng-star-inserted":
                driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(3)
            driver.find_element(By.LINK_TEXT, "Waves").click()
            val = co.findWave(waveName)
            if val == 2:
                return
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "waves_"+waveName+"_wave_actions"))
        )
        driver.find_element(By.ID, "waves_"+waveName+"_wave_actions").click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_delete_xpath)) != 0:
            self.logger.info("********** Delete Wave Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
            driver.find_element(By.ID, LOC.btn_deleteWave_id).click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
            )
            note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
            self.logger.info("********** Delete Wave Status of Wave : " + waveName + ",")
            self.logger.info(note + "\n")
            time.sleep(2)
            driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
        else:
            self.logger.info("********** Delete Wave Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
        time.sleep(5)

    def deleteRepWaves(self, driver, waveNames):
        time.sleep(5)
        res = tuple(map(str, waveNames.split(', ')))
        replication_class = driver.find_element(By.XPATH, LOC.txt_replication_xpath).get_attribute("class")
        if replication_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        wave_class = driver.find_element(By.XPATH, LOC.txt_rWave_xpath).get_attribute("class")
        if wave_class != "active":
            driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
        flag = 0
        for waveName in res:
            driver.find_element(By.XPATH, LOC.txt_waveSearch_xpath).click()
            driver.find_element(By.XPATH, LOC.txt_waveSearch_xpath).clear()
            driver.find_element(By.XPATH, LOC.txt_waveSearch_xpath).send_keys(waveName)
            time.sleep(2)
            totalWaves = len(driver.find_elements(By.XPATH, LOC.txt_totalPolicies_xpath))
            if totalWaves == 0:
                self.logger.info("********** " + waveName + " Was Not Found In Replication Waves **********")
            elif totalWaves == 1:
                tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span/div/a').text
                if tmp == waveName:
                    driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** " + waveName + " Was Not Found In Replication Waves **********")
            else:
                for i in range(1, totalWaves + 1):
                    tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span/div/a').text
                    if tmp == waveName:
                        driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalWaves:
                        self.logger.info("********** " + waveName + " Was Not Found In Replication Waves **********")
        if flag > 0:
            ele1 = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_deleteWaves_xpath))
            )
            ele1.click()
            time.sleep(5)
            if len(driver.find_elements(By.ID, LOC.pop_delMultWaves_id)) != 0:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Opened **********")
                driver.find_element(By.ID, LOC.btn_bulkDelete_id).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_deleteSuccessful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).text
                self.logger.info("********** Delete Wave Status,")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Not Opened **********")

    def deleteDrWaves(self, driver, waveNames):
        time.sleep(5)
        res = tuple(map(str, waveNames.split(', ')))
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(5)
        wave_class = driver.find_element(By.XPATH, LOC.txt_drWave_xpath).get_attribute("class")
        if wave_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
        flag = 0
        for waveName in res:
            driver.find_element(By.XPATH, LOC.txt_waveSearch_xpath).click()
            driver.find_element(By.XPATH, LOC.txt_waveSearch_xpath).clear()
            driver.find_element(By.XPATH, LOC.txt_waveSearch_xpath).send_keys(waveName)
            time.sleep(2)
            totalWaves = len(driver.find_elements(By.XPATH, LOC.txt_totalPolicies_xpath))
            if totalWaves == 0:
                self.logger.info("********** "+waveName+" Was Not Found In DR Waves **********")
            elif totalWaves == 1:
                tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span/div/a').text
                if tmp == waveName:
                    driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** " + waveName + " Was Not Found In DR Waves **********")
            else:
                for i in range(1, totalWaves+1):
                    tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span/div/a').text
                    if tmp == waveName:
                        driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalWaves:
                        self.logger.info("********** "+waveName+" Was Not Found In DR Waves **********")
        if flag > 0:
            ele1 = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_deleteWaves_xpath))
            )
            ele1.click()
            time.sleep(5)
            if len(driver.find_elements(By.ID, LOC.pop_delMultWaves_id)) != 0:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Opened **********")
                driver.find_element(By.ID, LOC.btn_bulkDelete_id).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_deleteSuccessful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).text
                self.logger.info("********** Delete Wave Status,")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Not Opened **********")

    def deleteDRPolicy(self, driver, policyNames):
        time.sleep(5)
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(2)
        policies_class = driver.find_element(By.XPATH, LOC.txt_policies_xpath).get_attribute("class")
        if policies_class != "active":
            driver.find_element(By.LINK_TEXT, "Policies").click()
            time.sleep(5)
        res = tuple(map(str, policyNames.split(', ')))
        flag = 0
        for policyName in res:
            driver.find_element(By.XPATH, LOC.txt_policySearch_xpath).click()
            driver.find_element(By.XPATH, LOC.txt_policySearch_xpath).clear()
            driver.find_element(By.XPATH, LOC.txt_policySearch_xpath).send_keys(policyName)
            time.sleep(2)
            totalPolicies = len(driver.find_elements(By.XPATH, LOC.txt_totalPolicies_xpath))
            if totalPolicies == 0:
                self.logger.info("********** Policy : " + policyName + ", Was Not Present **********")
            if totalPolicies == 1:
                tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
                if tmp == policyName:
                    driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[3]').click()
                    flag += 1
                else:
                    self.logger.info("********** Policy : " + policyName + ", Was Not Present **********")
            else:
                for i in range(1, totalPolicies+1):
                    tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/span').text
                    if tmp == policyName:
                        driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[9]/span/i[3]').click()
                        flag += 1
                        break
                    if i == totalPolicies:
                        self.logger.info("********** Policy : " + policyName + ", Was Not Present **********")
        if flag > 0:
            time.sleep(2)
            if len(driver.find_elements(By.XPATH, LOC.pop_deletePolicy_xpath)) != 0:
                self.logger.info("********** Delete Policy Pop Up Banner Was Opened **********")
                driver.find_element(By.ID, LOC.btn_deleteDRP_id).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Delete Policy Status, ")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Delete Policy Pop Up Banner Was Not Opened **********")

    def deleteSourceHost(self, driver, hostNames):
        time.sleep(5)
        hi_class = driver.find_element(By.XPATH, LOC.txt_hostNImages_xpath).get_attribute("class")
        if hi_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Hosts & Images").click()
            time.sleep(2)
        hosts_class = driver.find_element(By.XPATH, LOC.txt_hosts_xpath).get_attribute("class")
        if hosts_class != "active":
            driver.find_element(By.LINK_TEXT, "Hosts").click()
            time.sleep(5)
        res = tuple(map(str, hostNames.split(', ')))
        flag = 0
        for hostName in res:
            time.sleep(2)
            driver.find_element(By.XPATH, LOC.txt_hostSearch_xpath).click()
            driver.find_element(By.XPATH, LOC.txt_hostSearch_xpath).clear()
            driver.find_element(By.XPATH, LOC.txt_hostSearch_xpath).send_keys(hostName)
            time.sleep(3)
            totalHosts = len(driver.find_elements(By.XPATH, LOC.txt_totalHosts_xpath))
            if totalHosts == 0:
                self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            elif totalHosts == 1:
                if len(driver.find_elements(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span')) != 0:
                    tp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span').text
                    if tp == "* No hosts to show.":
                        continue
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span[1]').text
                if tmp == hostName:
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]'))
                    )
                    driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            else:
                for i in range(1, totalHosts + 1):
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span[1]').text
                    if tmp == hostName:
                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]'))
                        )
                        driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalHosts:
                        self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
        if flag > 0:
            ele = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_delete_xpath))
            )
            ele.click()
            time.sleep(3)
            if len(driver.find_elements(By.XPATH, LOC.pop_deleteHost_xpath)) != 0:
                self.logger.info("********** Delete Host Pop Up Banner Was Opened **********")
                driver.find_element(By.ID, LOC.btn_deleteHost_id).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_deleteSuccessful_xpath))
                )
                self.logger.info("********** Delete Status For Hosts,")
                for i in range(1, len(res)+1):
                    if len(driver.find_elements(By.XPATH, LOC.pop_deleteSuccessful_xpath)) != 0:
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, LOC.pop_deleteSuccessful_xpath))
                        )
                        note = driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).text
                        self.logger.info(note + "\n")
                        driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Delete Host Pop Up Banner Was Not Opened **********")

    def deleteTargetHost(self, driver, hostNames, VM, environment, name):
        time.sleep(5)
        hi_class = driver.find_element(By.XPATH, LOC.txt_hostNImages_xpath).get_attribute("class")
        if hi_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Hosts & Images").click()
            time.sleep(2)
        hosts_class = driver.find_element(By.XPATH, LOC.txt_hosts_xpath).get_attribute("class")
        if hosts_class != "active":
            driver.find_element(By.LINK_TEXT, "Hosts").click()
            time.sleep(5)
        res = tuple(map(str, hostNames.split(', ')))
        flag = 0
        for hostName in res:
            driver.find_element(By.XPATH, LOC.txt_hostSearch_xpath).click()
            driver.find_element(By.XPATH, LOC.txt_hostSearch_xpath).clear()
            driver.find_element(By.XPATH, LOC.txt_hostSearch_xpath).send_keys(hostName)
            time.sleep(2)
            totalHosts = len(driver.find_elements(By.XPATH, LOC.txt_totalHosts_xpath))
            if totalHosts == 0:
                self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            if totalHosts == 1:
                if len(driver.find_elements(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span')) != 0:
                    tp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span').text
                    if tp == "* No hosts to show.":
                        continue
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span[1]').text
                if tmp == hostName:
                    driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            else:
                for i in range(1, totalHosts + 1):
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span[1]').text
                    if tmp == hostName:
                        driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalHosts:
                        self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
        if flag > 0:
            ele = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_delete_xpath))
            )
            ele.click()
            time.sleep(3)
            if len(driver.find_elements(By.XPATH, LOC.pop_deleteHost_xpath)) != 0:
                self.logger.info("********** Delete Host Pop Up Banner Was Opened **********")
                if VM:
                    driver.find_element(By.XPATH, LOC.ch_deleteVM_xpath).click()
                    time.sleep(3)
                    if environment == "clouduser":
                        driver.find_element(By.XPATH, LOC.rd_cloudUSer_xpath).click()
                    elif environment == "vCenter":
                        driver.find_element(By.XPATH, LOC.rd_vCenter_xpath).click()
                    elif environment == "Hypervisor":
                        driver.find_element(By.XPATH, LOC.rd_hypervisor_xpath).click()
                    time.sleep(3)
                    select = Select(driver.find_element(By.XPATH, LOC.txt_select_xpath))
                    select.select_by_visible_text(str(name))
                ele2 = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_deleteHost_id))
                )
                ele2.click()
                time.sleep(10)
            else:
                self.logger.info("********** Delete Host Pop Up Banner Was Not Opened **********")
