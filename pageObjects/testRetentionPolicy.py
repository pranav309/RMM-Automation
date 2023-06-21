import time
import unittest
import openpyxl

import Locators.locRetentionPolicy as LOC

from selenium.webdriver.common.by import By
from utilities.customLogger import LogGen
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class RetentionPolicy(unittest.TestCase):

    logger = LogGen.loggen()

    def createRetentionPolicy(self, driver, path, start, end):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        backup_class = driver.find_element(By.XPATH, LOC.txt_backup_xpath).get_attribute("class")
        if backup_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Backup & Restore").click()
        rp_class = driver.find_element(By.XPATH, LOC.txt_retentionPolicy_xpath).get_attribute("class")
        if rp_class == "active":
            driver.find_element(By.LINK_TEXT, "Retention Policies").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_add_xpath))
        )
        for r in range(st, ed+1):
            policyName = sheet.cell(row=r, column=1).value
            retentionType = sheet.cell(row=r, column=2).value
            retainCount = sheet.cell(row=r, column=3).value
            shortMinInt = sheet.cell(row=r, column=4).value
            shortMinIntType = sheet.cell(row=r, column=5).value
            shortMaxWin = sheet.cell(row=r, column=6).value
            shortMaxWinType = sheet.cell(row=r, column=7).value
            longMinInt = sheet.cell(row=r, column=8).value
            longMinIntType = sheet.cell(row=r, column=9).value
            longMaxWin = sheet.cell(row=r, column=10).value
            longMaxWinType = sheet.cell(row=r, column=11).value
            driver.find_element(By.XPATH, LOC.btn_add_xpath).click()
            time.sleep(3)
            if len(driver.find_elements(By.ID, LOC.val_createRP_id)) != 0:
                self.logger.info("********** Create New Retention Policy Pop-up Banner Was Opened For Retention Policy, " + str(policyName) + " **********")
                driver.find_element(By.XPATH, LOC.txt_policyName_xpath).send_keys(policyName)
                if retentionType == "By Interval":
                    driver.find_element(By.XPATH, LOC.rd_byInterval_xpath).click()
                    driver.find_element(By.XPATH, LOC.txt_shortMinInt_xpath).send_keys(shortMinInt)
                    self.selectShortType(driver, shortMinIntType, 2)
                    driver.find_element(By.XPATH, LOC.txt_shortMaxWin_xpath).send_keys(shortMaxWin)
                    self.selectShortType(driver, shortMaxWinType, 3)
                    driver.find_element(By.XPATH, LOC.txt_longMinInt_xpath).send_keys(longMinInt)
                    self.selectLongType(driver, longMinIntType, 2)
                    driver.find_element(By.XPATH, LOC.txt_longMaxWin_xpath).send_keys(longMaxWin)
                    self.selectLongType(driver, longMaxWinType, 3)

                elif retentionType == "By Count":
                    driver.find_element(By.XPATH, LOC.rd_byCount_xpath).click()
                    driver.find_element(By.XPATH, LOC.txt_retainCount_xpath).send_keys(retainCount)

                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_create_xpath))
                )
                btn.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Create New Retention Policy Status, ")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Create New Retention Policy Pop-up Banner Was Not Opened For Retention Policy, " + str(policyName) + " **********")

    @staticmethod
    def selectShortType(driver, val, minMax):
        if val == "Hours":
            driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[1]/span').click()
        elif val == "Days":
            driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[2]/span').click()
        elif val == "Weeks":
            driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[3]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[3]/span').click()

    @staticmethod
    def selectLongType(driver, val, minMax):
        if val == "Days":
            driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[1]/span').click()
        elif val == "Weeks":
            driver.find_element(By.XPATH, '//*[@id="main"]/app-retention-policies/create-retention-policy/div/div/div/form/div[2]/div[4]/div[' + str(minMax) + ']/div[2]/p-selectbutton/div/div[2]/span').click()

    def startRetentionPolicy(self, driver, vals, images):
        for i in vals:
            driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[7]/span/div/i[1]').click()
            time.sleep(2)
            driver.find_element(By.XPATH, LOC.txt_imageName_xpath).click()
            time.sleep(2)
            for image in images:
                driver.find_element(By.XPATH, LOC.txt_searchImage_xpath).send_keys(image)

    def editRetentionPolicy(self, driver, path, start, end):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        backup_class = driver.find_element(By.XPATH, LOC.txt_backup_xpath).get_attribute("class")
        if backup_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Backup & Restore").click()
        rp_class = driver.find_element(By.XPATH, LOC.txt_retentionPolicy_xpath).get_attribute("class")
        if rp_class == "active":
            driver.find_element(By.LINK_TEXT, "Retention Policies").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_add_xpath))
        )
        for r in range(st, ed+1):
            policyName = sheet.cell(row=r, column=1).value
            retentionType = sheet.cell(row=r, column=2).value
            retainCount = sheet.cell(row=r, column=3).value
            shortMinInt = sheet.cell(row=r, column=4).value
            shortMinIntType = sheet.cell(row=r, column=5).value
            shortMaxWin = sheet.cell(row=r, column=6).value
            shortMaxWinType = sheet.cell(row=r, column=7).value
            longMinInt = sheet.cell(row=r, column=8).value
            longMinIntType = sheet.cell(row=r, column=9).value
            longMaxWin = sheet.cell(row=r, column=10).value
            longMaxWinType = sheet.cell(row=r, column=11).value

            totalPolicies = len(driver.find_elements(By.XPATH, LOC.cnt_totalPolicies_xpath))
            if totalPolicies == 1:
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
                if tmp == policyName:
                    driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[7]/span/div/i[2]').click()
                else:
                    self.logger.info("********** There Was No Retention Policy With Name : " + policyName + "**********")
                    continue
            else:
                for i in range(1, totalPolicies+1):
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/span').text
                    if tmp == policyName:
                        driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[7]/span/div/i[2]').click()
                        break
                    if i == totalPolicies:
                        self.logger.info("********** There Was No Retention Policy With Name : " + policyName + "**********")
                        continue
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.val_modRP_xpath)) != 0:
                self.logger.info("********** Modify Retention Policy Pop-up Banner Was Opened For Policy, " + str(policyName) + " **********")
                driver.find_element(By.XPATH, LOC.txt_policyName_xpath).send_keys(policyName)
                if retentionType == "By Interval":
                    driver.find_element(By.XPATH, LOC.rd_byInterval_xpath).click()
                    driver.find_element(By.XPATH, LOC.txt_shortMinInt_xpath).send_keys(shortMinInt)
                    self.selectShortType(driver, shortMinIntType, 2)
                    driver.find_element(By.XPATH, LOC.txt_shortMaxWin_xpath).send_keys(shortMaxWin)
                    self.selectShortType(driver, shortMaxWinType, 3)
                    driver.find_element(By.XPATH, LOC.txt_longMinInt_xpath).send_keys(longMinInt)
                    self.selectLongType(driver, longMinIntType, 2)
                    driver.find_element(By.XPATH, LOC.txt_longMaxWin_xpath).send_keys(longMaxWin)
                    self.selectLongType(driver, longMaxWinType, 3)

                elif retentionType == "By Count":
                    driver.find_element(By.XPATH, LOC.rd_byCount_xpath).click()
                    driver.find_element(By.XPATH, LOC.txt_retainCount_xpath).send_keys(retainCount)

                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_modifyRP_xpath))
                )
                driver.find_element(By.XPATH, LOC.btn_modifyRP_xpath).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Create New Retention Policy Status of Policy : " + policyName + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Modify Retention Policy Pop-up Banner Was Not Opened For Policy, " + str(policyName) + " **********")

    def deleteRetentionPolicy(self, driver, policies, force):
        backup_class = driver.find_element(By.XPATH, LOC.txt_backup_xpath).get_attribute("class")
        if backup_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Backup & Restore").click()
        rp_class = driver.find_element(By.XPATH, LOC.txt_retentionPolicy_xpath).get_attribute("class")
        if rp_class == "active":
            driver.find_element(By.LINK_TEXT, "Retention Policies").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_add_xpath))
        )
        for policy in policies:
            totalPolicies = len(driver.find_elements(By.XPATH, LOC.cnt_totalPolicies_xpath))
            if totalPolicies == 1:
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
                if tmp == policy:
                    driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[7]/span/div/i[3]').click()
                else:
                    self.logger.info("********** There Was No Retention Policy With Name : " + policy + "**********")
                    continue
            else:
                for i in range(1, totalPolicies+1):
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/span').text
                    if tmp == policy:
                        driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[7]/span/div/i[3]').click()
                        break
                    if i == totalPolicies:
                        self.logger.info("********** There Was No Retention Policy With Name : " + policy + "**********")
                        continue
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_delete_xpath)) != 0:
                self.logger.info("********** Delete Retention Policy Pop-Up Banner Was Opened For Retention Policy With Name : " + policy + " **********")
                if len(driver.find_elements(By.XPATH, LOC.ch_forceDelete_xpath)) != 0 and force:
                    driver.find_element(By.XPATH, LOC.pop_delete_xpath).click()
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_deleteRetentionPolicy_id))
                )
                driver.find_element(By.ID, LOC.btn_deleteRetentionPolicy_id).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Delete Retention Policy Status For Retention Policy: " + policy + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Delete Retention Policy Pop-Up Banner Was Not Opened For Retention Policy With Name : " + policy + " **********")
