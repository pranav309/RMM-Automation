import time
import unittest
import openpyxl
import keyboard

import Locators.locWaveEdit as LOC

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select

from utilities.customLogger import LogGen
from utilities.commonObjects import CommonObjects


class WaveEdit(unittest.TestCase):

    logger = LogGen.loggen()

    def setAutoprovision(self, driver, path):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        for r in range(3, rows+1):
            waveName = sheet.cell(row=r, column=1).value
            CUType = sheet.cell(row=r, column=2).value
            environment = sheet.cell(row=r, column=3).value
            co = CommonObjects(driver)
            val = co.findWave(waveName)
            if val == 2:
                continue
            if val == 1:
                driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
                else:
                    self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                    continue
            if val == 0:
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
            cu = driver.find_element(By.ID, LOC.txt_autoprovision_id).text
            if cu == environment:
                self.logger.info("********** Wave : " + waveName + ", Already Have Cloud User " + environment + " Added To It **********")
                continue
            ele = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, LOC.txt_autoprovision_id))
            )
            ele.click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_setAutoprovision_xpath)) != 0:
                self.logger.info("********** Select An Environment Pop-up Banner Is Opened For Wave, " + str(waveName) + " **********")
                env = Select(driver.find_element(By.XPATH, LOC.txt_environment_xpath))
                env.select_by_visible_text(environment)
                time.sleep(5)
                if CUType == "VCenter":
                    self.setVCenter(driver, sheet, r)
                elif CUType == "AWS":
                    self.setAWS(driver, sheet, r)
                elif CUType == "OCI":
                    self.setOCI(driver, sheet, r)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Set Autoprovision Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Select An Environment Pop-up Banner Is Not Opened For Wave, " + str(waveName) + " **********")

    def setVCenter(self, driver, sheet, r):
        waveName = sheet.cell(row=r, column=1).value
        clusterName = sheet.cell(row=r, column=4).value
        esxHost = sheet.cell(row=r, column=5).value
        datastore = sheet.cell(row=r, column=6).value
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="wave_policy_wave_policy_wave_detail_cloudUserEdit"]/div/div/div/form/div[2]/div/div[2]/div[1]'))
        )
        driver.find_element(By.XPATH, LOC.txt_clusterName_xpath).send_keys(clusterName)
        driver.find_element(By.XPATH, LOC.txt_ESXHost_xpath).send_keys(esxHost)
        driver.find_element(By.XPATH, LOC.txt_Datastore_xpath).send_keys(datastore)
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, LOC.btn_applyChanges_id))
        )
        btn.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
        )
        note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
        self.logger.info("********** Set Autoprovision Status of Wave : " + waveName + ",")
        self.logger.info(note + "\n")
        time.sleep(2)
        driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
        totalHosts = len(driver.find_elements(By.ID, LOC.txt_totalHosts_id))
        if totalHosts == 1:
            self.setNICEdit(driver, sheet, r)
        else:
            self.setNICBulkEdit(driver, sheet, r)

    def setAWS(self, driver, sheet, r):
        waveName = sheet.cell(row=r, column=1).value
        vpcID = sheet.cell(row=r, column=4).value
        subnetID = sheet.cell(row=r, column=5).value
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="wave_policy_wave_policy_wave_detail_cloudUserEdit"]/div/div/div/form/div[2]/div/div[2]/div[1]'))
        )
        driver.find_element(By.ID, LOC.txt_AwsVpcId_id).send_keys(vpcID)
        driver.find_element(By.ID, LOC.txt_AWSSubnetID_id).send_keys(subnetID)
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, LOC.btn_applyChanges_id))
        )
        btn.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
        )
        note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
        self.logger.info("********** Set Autoprovision Status of Wave : " + waveName + ",")
        self.logger.info(note + "\n")
        time.sleep(2)
        driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()

    def setOCI(self, driver, sheet, r):
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="wave_policy_wave_policy_wave_detail_cloudUserEdit"]/div/div/div/form/div[2]/div/div[2]/div[1]'))
        )
        waveName = sheet.cell(row=r, column=1).value
        region = sheet.cell(row=r, column=4).value
        driver.find_element(By.XPATH, LOC.txt_OCIRegion_xpath).clear()
        driver.find_element(By.XPATH, LOC.txt_OCIRegion_xpath).send_keys(region)
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, LOC.btn_applyChanges_id))
        )
        btn.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
        )
        note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
        self.logger.info("********** Set Autoprovision Status of Wave : " + waveName + ",")
        self.logger.info(note + "\n")
        time.sleep(2)
        driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
        totalHosts = len(driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
        if totalHosts == 1:
            self.setOCISync(driver, sheet, r)
        else:
            self.setOCIBulkSync(driver, sheet, r)

    def setNICEdit(self, driver, sheet, r):
        time.sleep(5)
        waveName = sheet.cell(row=r, column=1).value
        deviceName = sheet.cell(row=r, column=7).value
        Type = sheet.cell(row=r, column=8).value
        networkName = sheet.cell(row=r, column=9).value
        ipType = sheet.cell(row=r, column=10).value
        CIDR = sheet.cell(row=r, column=11).value
        gateway = sheet.cell(row=r, column=12).value
        DNS1 = sheet.cell(row=r, column=13).value
        DNS2 = sheet.cell(row=r, column=14).value

        driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_edit_xpath)) != 0:
            self.logger.info("********** Edit Host Pop-Up Banner Was Opened For Wave : "+waveName+" **********")
            driver.find_element(By.LINK_TEXT, "vCenter Options").click()
            time.sleep(5)
            vCenter_class = driver.find_element(By.XPATH, LOC.txt_editCU_xpath).get_attribute("class")
            if vCenter_class == "ui-state-default ui-corner-top ng-star-inserted ui-tabview-selected ui-state-active":
                self.logger.info("********** vCenter Option Pop-Up Banner Was Opened For Wave : "+waveName+" **********")
                driver.find_element(By.XPATH, LOC.btn_NICAdd_xpath).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.pop_addNIC_xpath)) != 0:
                    self.logger.info("********** Add NIC Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                    totalNICs = len(driver.find_elements(By.XPATH, LOC.txt_totalNICs_xpath))
                    self.logger.info("********** Wave : " + waveName + " Have " + str(totalNICs) + " NICs **********")
                    driver.find_element(By.XPATH, LOC.btn_NICAdd_xpath).click()
                    if ipType == "DHCP":
                        driver.find_element(By.XPATH, LOC.rd_DHCP_xpath).click()
                    elif ipType == "Static IP":
                        driver.find_element(By.XPATH, LOC.rd_staticIP_xpath).click()
                        driver.find_element(By.ID, LOC.txt_CIDR_id).send_keys(CIDR)
                        driver.find_element(By.ID, LOC.txt_gateway_id).send_keys(gateway)
                        driver.find_element(By.ID, LOC.txt_DNS1_id).send_keys(DNS1)
                        driver.find_element(By.ID, LOC.txt_DNS2_id).send_keys(DNS2)
                    driver.find_element(By.ID, LOC.txt_deviceName_id).send_keys(deviceName)
                    driver.find_element(By.ID, LOC.txt_type_id).send_keys(Type)
                    driver.find_element(By.ID, LOC.txt_networkName_id).send_keys(networkName)
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, LOC.btn_save_id))
                    )
                    btn.click()
                    time.sleep(5)
                    if len(driver.find_elements(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/div/p-tabpanel[3]/div/div/div[4]/div[1]/div/div['+str(totalNICs + 1)+']')) != 0:
                        self.logger.info("********** NIC Added Successfully For Wave : " + waveName + " **********")
                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.ID, LOC.btn_modify_id))
                        )
                        driver.find_element(By.ID, LOC.btn_modify_id).click()
                    else:
                        self.logger.info("********** Failed To Add NIC For Wave : " + waveName + " **********")
                        driver.find_element(By.XPATH, LOC.btn_cancelEdit_xpath).click()
                else:
                    self.logger.info("********** Add NIC Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                    driver.find_element(By.XPATH, LOC.btn_cancelEdit_xpath).click()
            else:
                self.logger.info("********** vCenter Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.btn_cancelEdit_xpath).click()
        else:
            self.logger.info("********** Sync Option Wave Pop-Up Banner Was Not Opened For Wave : "+waveName+" **********")

    def setNICBulkEdit(self, driver, sheet, r):
        time.sleep(5)
        waveName = sheet.cell(row=r, column=1).value
        deviceName = sheet.cell(row=r, column=7).value
        Type = sheet.cell(row=r, column=8).value
        networkName = sheet.cell(row=r, column=9).value
        ipType = sheet.cell(row=r, column=10).value
        CIDR = sheet.cell(row=r, column=11).value
        gateway = sheet.cell(row=r, column=12).value
        DNS1 = sheet.cell(row=r, column=13).value
        DNS2 = sheet.cell(row=r, column=14).value

        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
        )
        btn.click()
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_bulkEdit_xpath))
        )
        btn.click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_bulkEdit_xpath)) != 0:
            self.logger.info("********** Bulk Edit Wave Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
            driver.find_element(By.LINK_TEXT, "vCenter Options").click()
            vCenter_class = driver.find_element(By.XPATH, LOC.txt_bulkEditCU_xpath).get_attribute("class")
            if vCenter_class == "ui-state-default ui-corner-top ng-star-inserted ui-tabview-selected ui-state-active":
                self.logger.info("********** Bulk Edit vCenter Option Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                beforeTotalNIcs = len(driver.find_elements(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[4]/div/p-tabview/div/div/p-tabpanel[2]/div/div/div[5]/div[1]/div/div'))
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_NICAddBulkEdit_xpath))
                )
                btn.click()
                time.sleep(3)
                if len(driver.find_elements(By.XPATH, LOC.pop_bulkNIC_xpath)) != 0:
                    self.logger.info("********** Bulk Edit Add NIC Option Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                    driver.find_element(By.XPATH, LOC.txt_deviceNameBulkEdit_xpath).send_keys(deviceName)
                    driver.find_element(By.XPATH, LOC.txt_typeBulkEdit_xpath).send_keys(Type)
                    driver.find_element(By.XPATH, LOC.txt_networkNameBulkEdit_xpath).send_keys(networkName)
                    if ipType == "DHCP":
                        driver.find_element(By.XPATH, LOC.rd_DHCPBulkEdit_xpath).click()
                    elif ipType == "Static IP":
                        driver.find_element(By.XPATH, LOC.rd_staticIPBulkEdit_xpath).click()
                        driver.find_element(By.XPATH, LOC.txt_CIDRBulkEdit_xpath).send_keys(CIDR)
                        driver.find_element(By.XPATH, LOC.txt_gatewayBulkEdit_xpath).send_keys(gateway)
                        driver.find_element(By.XPATH, LOC.txt_DNS1BulkEdit_xpath).send_keys(DNS1)
                        driver.find_element(By.XPATH, LOC.txt_DNS2BulkEdit_xpath).send_keys(DNS2)
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, LOC.btn_save_xpath))
                    )
                    btn.click()
                    time.sleep(5)
                    afterTotalNIcs = len(driver.find_elements(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[4]/div/p-tabview/div/div/p-tabpanel[2]/div/div/div[5]/div[1]/div/div'))
                    if afterTotalNIcs > beforeTotalNIcs:
                        self.logger.info("********** Successfully Added NIC For Wave : " + waveName + " **********")
                    else:
                        self.logger.info("********** Failed To Add NIC For Wave : " + waveName + " **********")
                else:
                    self.logger.info("********** Bulk Edit Add NIC Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                    return
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_next_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_modifyAll_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_yes_id))
                )
                btn.click()
            else:
                self.logger.info("********** Bulk Edit vCenter Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.btn_cancelBulkEdit_xpath).click()
        else:
            self.logger.info("********** Bulk Edit Wave Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
        time.sleep(3)
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
        )
        btn.click()

    def setOCISync(self, driver, sheet, r):
        waveName = sheet.cell(row=r, column=1).value
        VCNName = sheet.cell(row=r, column=5).value
        SubnetName = sheet.cell(row=r, column=6).value
        AVDomain = sheet.cell(row=r, column=7).value
        driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_edit_xpath)) != 0:
            self.logger.info("********** Edit Host Pop-Up Banner Was Opened For Wave : "+waveName+" **********")
            driver.find_element(By.LINK_TEXT, "OCI Options").click()
            time.sleep(5)
            oci_class = driver.find_element(By.XPATH, LOC.txt_editCU_xpath).get_attribute("class")
            if oci_class == "ui-state-default ui-corner-top ng-star-inserted ui-tabview-selected ui-state-active":
                self.logger.info("********** OCI Option Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.txt_OCIVCNName_xpath).send_keys(VCNName)
                driver.find_element(By.XPATH, LOC.txt_OCISubnetName_xpath).send_keys(SubnetName)
                driver.find_element(By.XPATH, LOC.txt_OCIAVDomain_xpath).send_keys(AVDomain)
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_modify_id))
                )
                btn.click()
            else:
                self.logger.info("********** OCI Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.btn_cancelEdit_xpath).click()
        else:
            self.logger.info("********** Edit Host Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")

    def setOCIBulkSync(self, driver, sheet, r):
        waveName = sheet.cell(row=r, column=1).value
        VCNName = sheet.cell(row=r, column=5).value
        SubnetName = sheet.cell(row=r, column=6).value
        AVDomain = sheet.cell(row=r, column=7).value
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
        )
        btn.click()
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_selected((By.XPATH, LOC.btn_bulkEdit_xpath))
        )
        btn.click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_bulkEdit_xpath)) != 0:
            self.logger.info("********** Bulk Edit Wave Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
            driver.find_element(By.LINK_TEXT, "OCI Options").click()
            time.sleep(5)
            oci_class = driver.find_element(By.XPATH, LOC.txt_bulkEditCU_xpath).get_attribute("class")
            if oci_class == "ui-state-default ui-corner-top ng-star-inserted ui-tabview-selected ui-state-active":
                self.logger.info("********** Bulk Edit OCI Option Pop-Up Banner Was Opened For Wave : " + waveName + " **********")
                driver.find_element(By.ID, "oci_vcn_name").send_keys(VCNName)
                driver.find_element(By.ID, "oci_subnet_name").send_keys(SubnetName)
                driver.find_element(By.ID, "oci_av_domain").send_keys(AVDomain)
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_next_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_modifyAll_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_yes_id))
                )
                btn.click()
            else:
                self.logger.info("********** Bulk Edit OCI Option Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.btn_cancelBulkEdit_xpath).click()
        else:
            self.logger.info("********** Bulk Edit Wave Pop-Up Banner Was Not Opened For Wave : " + waveName + " **********")
        time.sleep(3)
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
        )
        btn.click()

    def setSyncOptions(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 3
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        time.sleep(5)
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            hostName = sheet.cell(row=r, column=2).value

            tng = sheet.cell(row=r, column=3).value
            verbose = sheet.cell(row=r, column=4).value
            directFScopy = sheet.cell(row=r, column=5).value
            FSDeletion = sheet.cell(row=r, column=6).value
            NoTransfer = sheet.cell(row=r, column=7).value
            transferCompress = sheet.cell(row=r, column=8).value
            noTransferCompress = sheet.cell(row=r, column=9).value
            ignoreMissing = sheet.cell(row=r, column=10).value
            noInPlace = sheet.cell(row=r, column=11).value
            noReboot = sheet.cell(row=r, column=12).value
            includeSAN = sheet.cell(row=r, column=13).value
            excludeSAN = sheet.cell(row=r, column=14).value
            overrideRMMStorageCheck = sheet.cell(row=r, column=15).value
            deleteAllTargetFS = sheet.cell(row=r, column=16).value
            keepTargetLayout = sheet.cell(row=r, column=17).value
            cloudInit = sheet.cell(row=r, column=18).value
            eventScript = sheet.cell(row=r, column=19).value
            eventScriptArgs = sheet.cell(row=r, column=20).value
            excludeFileSystem = sheet.cell(row=r, column=21).value
            includeFileSystem = sheet.cell(row=r, column=22).value
            filterFileOption = sheet.cell(row=r, column=23).value
            filterFilePath = sheet.cell(row=r, column=24).value

            co = CommonObjects(driver)
            val = co.findWave(waveName)
            if val == 2:
                continue
            if val == 1:
                driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
                else:
                    self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                    continue
            if val == 0:
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
            totalHosts = len(driver.find_elements(By.ID, LOC.txt_totalHosts_id))
            if totalHosts == 1:
                host = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                if hostName == host:
                    driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
                else:
                    self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
                    continue
            else:
                count = 0
                for hostNo in range(1, totalHosts+1):
                    host = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                    if hostName == host:
                        driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[8]/span/div/i[2]').click()
                        break
                    count += 1
                if count > totalHosts:
                    self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
                    continue
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_edit_xpath)) != 0:
                self.logger.info("********** Edit Host Pop-Up Banner Was Opened For Host : " + hostName + " **********")
                driver.find_element(By.XPATH, LOC.btn_syncOptions_xpath).click()
                time.sleep(5)
                syncOpt_class = driver.find_element(By.XPATH, LOC.txt_editSyncOption_xpath).get_attribute("class")
                if syncOpt_class == "ui-state-default ui-corner-top ng-star-inserted ui-tabview-selected ui-state-active":
                    self.logger.info("********** Sync Options Pop-up Banner Is Opened For Host : " + hostName + " **********")
                    if tng == "Yes":
                        driver.find_element(By.ID, LOC.rd_tng_id).click()
                        self.logger.info("********** Sync Option 'TNG' is set **********")
                    if verbose == "Yes":
                        driver.find_element(By.ID, LOC.rd_verbose_id).click()
                        self.logger.info("********** Sync Option 'Verbose' is set **********")
                    if directFScopy == "Yes":
                        driver.find_element(By.ID, LOC.rd_directFScopy_id).click()
                        self.logger.info("********** Sync Option 'Direct FScopy' is set **********")
                    if FSDeletion == "Yes":
                        driver.find_element(By.ID, LOC.rd_FSDeletion_id).click()
                        self.logger.info("********** Sync Option 'FS Deletion' is set **********")
                    if NoTransfer == "Yes":
                        driver.find_element(By.ID, LOC.rd_NoTransfer_id).click()
                        self.logger.info("********** Sync Option 'No Transfer' is set **********")
                    if transferCompress == "Yes":
                        driver.find_element(By.ID, LOC.rd_transferCompress_id).click()
                        self.logger.info("********** Sync Option 'Transfer Compress' is set **********")
                    if noTransferCompress == "Yes":
                        driver.find_element(By.ID, LOC.rd_noTransferCompress_id).click()
                        self.logger.info("********** Sync Option 'No Transfer Compress' is set **********")
                    if ignoreMissing == "Yes":
                        driver.find_element(By.ID, LOC.rd_ignoreMissing_id).click()
                        self.logger.info("********** Sync Option 'Ignore Missing' is set **********")
                    if noInPlace == "Yes":
                        driver.find_element(By.ID, LOC.rd_noInPlace_id).click()
                        self.logger.info("********** Sync Option 'No In Place' is set **********")
                    if noReboot == "Yes":
                        driver.find_element(By.ID, LOC.rd_noReboot_id).click()
                        self.logger.info("********** Sync Option 'No Reboot' is set **********")
                    if includeSAN == "Yes":
                        driver.find_element(By.ID, LOC.rd_includeSAN_id).click()
                        self.logger.info("********** Sync Option 'Include SAN' is set **********")
                    if excludeSAN == "Yes":
                        driver.find_element(By.ID, LOC.rd_excludeSAN_id).click()
                        self.logger.info("********** Sync Option 'Exclude SAN' is set **********")
                    if overrideRMMStorageCheck == "Yes":
                        driver.find_element(By.ID, LOC.rd_overrideRMMStorageCheck_id).click()
                        self.logger.info("********** Sync Option 'Override RMM Storage' Check is set **********")
                    if deleteAllTargetFS == "Yes":
                        driver.find_element(By.ID, LOC.rd_deleteAllTargetFS_id).click()
                        self.logger.info("********** Sync Option 'Delete All Target' FS is set **********")
                    if keepTargetLayout == "Yes":
                        driver.find_element(By.ID, LOC.rd_keepTargetLayout_id).click()
                        self.logger.info("********** Sync Option 'Keep Target Layout' is set **********")
                    if cloudInit == "Yes":
                        driver.find_element(By.ID, LOC.rd_cloudInit_id).click()
                        self.logger.info("********** Sync Option 'Cloud Init' is set **********")
                    if eventScript != "NA":
                        driver.find_element(By.ID, LOC.txt_eventScript_id).clear()
                        driver.find_element(By.ID, LOC.txt_eventScript_id).send_keys(eventScript)
                        self.logger.info("********** Sync Option 'Event Script' is set **********")
                    if eventScriptArgs != "NA":
                        driver.find_element(By.ID, LOC.txt_eventScriptArgs_id).clear()
                        driver.find_element(By.ID, LOC.txt_eventScriptArgs_id).send_keys(eventScriptArgs)
                        self.logger.info("********** Sync Option 'Event Script Arguments' is set **********")
                    if excludeFileSystem != "NA":
                        driver.find_element(By.ID, LOC.txt_excludeFile_id).clear()
                        driver.find_element(By.ID, LOC.txt_excludeFile_id).send_keys(excludeFileSystem)
                        self.logger.info("********** Sync Option 'Exclude File System(s)' is set **********")
                    if includeFileSystem != "NA":
                        driver.find_element(By.ID, LOC.txt_includeFile_id).clear()
                        driver.find_element(By.ID, LOC.txt_includeFile_id).send_keys(includeFileSystem)
                        self.logger.info("********** Sync Option 'Include File System(s)' is set **********")
                    if filterFileOption != "NA":
                        if filterFileOption == "Upload Local File":
                            driver.find_element(By.XPATH, LOC.rd_uploadLocalFile_xpath).click()
                            driver.find_element(By.XPATH, LOC.btn_uploadLocalFile_xpath).click()
                            time.sleep(5)
                            keyboard.write(filterFilePath)
                            time.sleep(5)
                            keyboard.send('enter')
                        elif filterFileOption == "File Path On RMM":
                            driver.find_element(By.XPATH, LOC.rd_filePathOnRmm_xpath).click()
                            driver.find_element(By.XPATH, LOC.txt_filePathOnRmm_xpath).send_keys(filterFilePath)
                        self.logger.info("********** Sync Option 'Filter File' is set **********")
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, LOC.btn_modify_id))
                    )
                    btn.click()
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                    )
                    note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                    self.logger.info("********** Set Edit Sync Options Status of Wave : " + waveName + ",")
                    self.logger.info(note + "\n")
                    time.sleep(2)
                    driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
                else:
                    self.logger.info("********** Sync Options Pop-up Banner Is Not Opened For Host : " + hostName + " **********")
                    driver.find_element(By.XPATH, LOC.btn_cancelEdit_xpath).click()
            else:
                self.logger.info("********** Edit Host Pop-Up Banner Was Not Opened For Host : " + hostName + " **********")

    def bulkEditSyncOption(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        time.sleep(5)
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            hostNames = sheet.cell(row=r, column=2).value
            tng = sheet.cell(row=r, column=3).value
            verbose = sheet.cell(row=r, column=4).value
            passwordLess = sheet.cell(row=r, column=5).value
            directFScopy = sheet.cell(row=r, column=6).value
            FSDeletion = sheet.cell(row=r, column=7).value
            NoTransfer = sheet.cell(row=r, column=8).value
            transferCompress = sheet.cell(row=r, column=9).value
            noTransferCompress = sheet.cell(row=r, column=10).value
            ignoreMissing = sheet.cell(row=r, column=11).value
            noInPlace = sheet.cell(row=r, column=12).value
            noReboot = sheet.cell(row=r, column=13).value
            includeSAN = sheet.cell(row=r, column=14).value
            excludeSAN = sheet.cell(row=r, column=15).value
            overrideRMMStorageCheck = sheet.cell(row=r, column=16).value
            deleteAllTargetFS = sheet.cell(row=r, column=17).value
            keepTargetLayout = sheet.cell(row=r, column=18).value
            cloudInit = sheet.cell(row=r, column=19).value
            excludeFile = sheet.cell(row=r, column=20).value
            includeFile = sheet.cell(row=r, column=21).value

            co = CommonObjects(driver)
            val = co.findWave(waveName)
            if val == 2:
                continue
            if val == 1:
                driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
                else:
                    self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                    continue
            if val == 0:
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
            totalHosts = len(driver.find_elements(By.ID, LOC.txt_totalHosts_id))
            if type(hostNames) != str:
                driver.find_element(By.XPATH, LOC.btn_selectAll_xpath).click()
            elif totalHosts == 1:
                self.logger.info("********** Can't Perform Bulk Edit On Wave : " + waveName + ", As There Is Only One Host Available **********")
            else:
                res = tuple(map(str, hostNames.split(', ')))
                for hostName in res:
                    for hostNo in range(1, totalHosts+1):
                        tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                        if hostName == tmp:
                            driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[1]/p-tablecheckbox/div/div[2]').click()
                            break
                        if hostNo == totalHosts:
                            self.logger.info("********** The Host " + str(hostName) + " Was Not Found In The Wave : " + waveName + " **********")
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_bulkEdit_xpath))
            )
            btn.click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_bulkEdit_xpath)) != 0:
                self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Is Opened For Wave, " + waveName + " **********")
                if tng == "Yes":
                    driver.find_element(By.ID, "tng_yes").click()
                    self.logger.info("********** Sync Option 'TNG' is set as yes **********")
                if tng == "No":
                    driver.find_element(By.ID, "tng_no").click()
                    self.logger.info("********** Sync Option 'TNG' is set as no **********")

                if verbose == "Yes":
                    driver.find_element(By.ID, "verbose_yes").click()
                    self.logger.info("********** Sync Option 'Verbose' is set as yes **********")
                if verbose == "No":
                    driver.find_element(By.ID, "verbose_no").click()
                    self.logger.info("********** Sync Option 'Verbose' is set as no **********")

                if passwordLess == "Yes":
                    driver.find_element(By.ID, "ssh_yes").click()
                    self.logger.info("********** Sync Option 'Passwordless' is set as yes **********")
                if passwordLess == "No":
                    driver.find_element(By.ID, "ssh_no").click()
                    self.logger.info("********** Sync Option 'Passwordless' is set as no **********")

                if directFScopy == "Yes":
                    driver.find_element(By.ID, "allowDirectFscopy_yes").click()
                    self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as yes **********")
                if directFScopy == "No":
                    driver.find_element(By.ID, "allowDirectFscopy_no").click()
                    self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as no **********")

                if FSDeletion == "Yes":
                    driver.find_element(By.ID, "allowFsDeletion_yes").click()
                    self.logger.info("********** Sync Option 'Allow FS Deletion' is set as yes **********")
                if FSDeletion == "No":
                    driver.find_element(By.ID, "allowFsDeletion_no").click()
                    self.logger.info("********** Sync Option 'Allow FS Deletion' is set as no **********")

                if NoTransfer == "Yes":
                    driver.find_element(By.ID, "no_xfer_yes").click()
                    self.logger.info("********** Sync Option 'No Transfer' is set as yes **********")
                if NoTransfer == "No":
                    driver.find_element(By.ID, "no_xfer_no").click()
                    self.logger.info("********** Sync Option 'No Transfer' is set as no **********")

                if transferCompress == "Yes":
                    driver.find_element(By.ID, "xferCompress_yes").click()
                    self.logger.info("********** Sync Option 'Transfer Compress' is set as yes **********")
                if transferCompress == "No":
                    driver.find_element(By.ID, "xferCompress_no").click()
                    self.logger.info("********** Sync Option 'Transfer Compress' is set as no **********")

                if noTransferCompress == "Yes":
                    driver.find_element(By.ID, "no_xfer_compress_yes").click()
                    self.logger.info("********** Sync Option 'No Transfer Compress' is set as yes **********")
                if noTransferCompress == "No":
                    driver.find_element(By.ID, "no_xfer_compress_no").click()
                    self.logger.info("********** Sync Option 'No Transfer Compress' is set as no **********")

                if ignoreMissing == "Yes":
                    driver.find_element(By.ID, "ignoreMissing_yes").click()
                    self.logger.info("********** Sync Option 'Ignore Missing' is set as yes **********")
                if ignoreMissing == "No":
                    driver.find_element(By.ID, "ignoreMissing_no").click()
                    self.logger.info("********** Sync Option 'Ignore Missing' is set as no **********")

                if noInPlace == "Yes":
                    driver.find_element(By.ID, "noInPlace_yes").click()
                    self.logger.info("********** Sync Option 'No In Place' is set as yes **********")
                if noInPlace == "No":
                    driver.find_element(By.ID, "noInPlace_no").click()
                    self.logger.info("********** Sync Option 'No In Place' is set as no **********")

                if noReboot == "Yes":
                    driver.find_element(By.ID, "noReboot_yes").click()
                    self.logger.info("********** Sync Option 'No Reboot' is set as yes **********")
                if noReboot == "No":
                    driver.find_element(By.ID, "noReboot_no").click()
                    self.logger.info("********** Sync Option 'No Reboot' is set as no **********")

                if includeSAN == "Yes":
                    driver.find_element(By.ID, "include_san_yes").click()
                    self.logger.info("********** Sync Option 'Include SAN' is set as yes **********")
                if includeSAN == "No":
                    driver.find_element(By.ID, "include_san_no").click()
                    self.logger.info("********** Sync Option 'Include SAN' is set as no **********")

                if excludeSAN == "Yes":
                    driver.find_element(By.ID, "exclude_san_yes").click()
                    self.logger.info("********** Sync Option 'Exclude SAN' is set as yes **********")
                if excludeSAN == "No":
                    driver.find_element(By.ID, "exclude_san_no").click()
                    self.logger.info("********** Sync Option 'Exclude SAN' is set as no **********")

                if overrideRMMStorageCheck == "Yes":
                    driver.find_element(By.ID, "storage_override_yes").click()
                    self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as yes **********")
                if overrideRMMStorageCheck == "No":
                    driver.find_element(By.ID, "storage_override_no").click()
                    self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as no **********")

                if deleteAllTargetFS == "Yes":
                    driver.find_element(By.ID, "delete_all_target_fs_yes").click()
                    self.logger.info("********** Sync Option 'Delete All Target FS' is set as yes **********")
                if deleteAllTargetFS == "No":
                    driver.find_element(By.ID, "delete_all_target_fs_no").click()
                    self.logger.info("********** Sync Option 'Delete All Target FS' is set as no **********")

                if keepTargetLayout == "Yes":
                    driver.find_element(By.ID, "keep_target_layout_yes").click()
                    self.logger.info("********** Sync Option 'Keep Target Layout' is set as yes **********")
                if keepTargetLayout == "No":
                    driver.find_element(By.ID, "keep_target_layout_no").click()
                    self.logger.info("********** Sync Option 'Keep Target Layout' is set as no **********")

                if cloudInit == "Yes":
                    driver.find_element(By.ID, "cloud_init_yes").click()
                    self.logger.info("********** Sync Option 'Cloud Init' is set as yes **********")
                if cloudInit == "No":
                    driver.find_element(By.ID, "cloud_init_no").click()
                    self.logger.info("********** Sync Option 'Cloud Init' is set as no **********")

                if excludeFile != "NA":
                    driver.find_element(By.ID, LOC.txt_blkExcludeFile_id).clear()
                    driver.find_element(By.ID, LOC.txt_blkExcludeFile_id).send_keys(excludeFile)
                    self.logger.info("********** Sync Option 'Exclude File System(s)' is set as no **********")

                if includeFile != "NA":
                    driver.find_element(By.ID, LOC.txt_blkIncludeFile_id).clear()
                    driver.find_element(By.ID, LOC.txt_blkIncludeFile_id).send_keys(includeFile)
                    self.logger.info("********** Sync Option 'Exclude File System(s)' is set as no **********")

                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_next_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_modifyAll_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_yes_id))
                )
                btn.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Set Bulk Edit Sync Options Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
                time.sleep(3)
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Is Not Opened For Wave, " + waveName + " **********")
            time.sleep(2)
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
            )
            btn.click()
            if type(hostNames) == str:
                time.sleep(2)
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
                )
                btn.click()

    def changeTargetType(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end

        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            hostName = sheet.cell(row=r, column=2).value
            targetType = sheet.cell(row=r, column=3).value
            co = CommonObjects(driver)
            val = co.findWave(waveName)
            if val == 2:
                continue
            if val == 1:
                driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
                else:
                    self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                    continue
            if val == 0:
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
            totalHosts = len(driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
            if totalHosts == 1:
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                if hostName == tmp:
                    driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
                else:
                    self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
                    continue
            else:
                count = 0
                for hostNo in range(1, totalHosts + 1):
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                    if hostName == tmp:
                        driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[8]/span/div/i[2]').click()
                        break
                    count += 1
                if count > totalHosts:
                    self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
                    continue
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_edit_xpath)) != 0:
                self.logger.info("********** Wave System Options Pop-up Banner Is Opened For Wave, " + waveName + " **********")
                if targetType == "Autoprovision":
                    driver.find_element(By.ID, LOC.rd_autoprovision_id).click()
                elif targetType == "Existing System":
                    imageName = sheet.cell(row=r, column=4).value
                    syncType = sheet.cell(row=r, column=5).value
                    passthrough = sheet.cell(row=r, column=6).value
                    targetIP = sheet.cell(row=r, column=7).value
                    friendlyName = sheet.cell(row=r, column=8).value
                    userName = sheet.cell(row=r, column=9).value
                    driver.find_element(By.ID, LOC.rd_existingSystem_id).click()
                    time.sleep(5)
                    if syncType == "Direct Sync":
                        driver.find_element(By.ID, LOC.rd_directSync_id).click()
                        time.sleep(3)
                        tmp = driver.find_element(By.ID, LOC.ch_passthrough_id).is_selected()
                        if passthrough ^ tmp:
                            driver.find_element(By.ID, LOC.ch_passthrough_id).click()
                        driver.find_element(By.ID, LOC.txt_targetIP_id).clear()
                        driver.find_element(By.ID, LOC.txt_targetIP_id).send_keys(targetIP)
                        driver.find_element(By.ID, LOC.txt_friendlyName_id).clear()
                        driver.find_element(By.ID, LOC.txt_friendlyName_id).send_keys(friendlyName)
                    elif syncType == "Stage 1 & 2" or syncType == "Stage 2":
                        if syncType == "Stage 1 & 2":
                            driver.find_element(By.ID, LOC.rd_stage12_id).click()
                        elif syncType == "Stage 2":
                            driver.find_element(By.ID, LOC.rd_stage2_id).click()
                        time.sleep(3)
                        driver.find_element(By.ID, LOC.txt_imageName_id).clear()
                        driver.find_element(By.ID, LOC.txt_imageName_id).send_keys(imageName)
                        tmp = driver.find_element(By.ID, LOC.ch_passthrough_id).is_selected()
                        if passthrough ^ tmp:
                            driver.find_element(By.ID, LOC.ch_passthrough_id).click()
                        driver.find_element(By.ID, LOC.txt_targetIP_id).clear()
                        driver.find_element(By.ID, LOC.txt_targetIP_id).send_keys(targetIP)
                        driver.find_element(By.ID, LOC.txt_friendlyName_id).clear()
                        driver.find_element(By.ID, LOC.txt_friendlyName_id).send_keys(friendlyName)
                    elif syncType == "Stage 1":
                        driver.find_element(By.ID, LOC.rd_stage1_id).click()
                        time.sleep(3)
                        driver.find_element(By.ID, LOC.txt_imageName_id).clear()
                        driver.find_element(By.ID, LOC.txt_imageName_id).send_keys(imageName)
                    driver.find_element(By.ID, LOC.txt_userName_id).clear()
                    driver.find_element(By.ID, LOC.txt_userName_id).send_keys(userName)
                elif targetType == "Capture":
                    imageName = sheet.cell(row=r, column=4).value
                    driver.find_element(By.ID, LOC.rd_capture_id).click()
                    driver.find_element(By.XPATH, LOC.txt_captureImage_xpath).clear()
                    driver.find_element(By.XPATH, LOC.txt_captureImage_xpath).send_keys(imageName)
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_modify_id))
                )
                btn.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Changed Target Type For Host : " + hostName + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
                time.sleep(5)
            else:
                self.logger.info("********** Wave System Options Pop-up Banner Is Not Opened For Wave, " + waveName + " **********")

    def moveHosts(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        for r in range(st, ed+1):
            sourceWave = sheet.cell(row=r, column=1).value
            targetWave = sheet.cell(row=r, column=2).value
            hostNames = sheet.cell(row=r, column=3).value
            time.sleep(5)
            co = CommonObjects(driver)
            val = co.findWave(sourceWave)
            if val == 2:
                return
            if val == 1:
                driver.find_element(By.LINK_TEXT, sourceWave).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + sourceWave + " Was Opened Successfully **********")
                else:
                    self.logger.info("********** Failed To Open Wave : " + sourceWave + " **********")
                    return
            if val == 0:
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + sourceWave + " Was Already Open **********")
            res = tuple(map(str, hostNames.split(', ')))
            totalHosts = len(driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
            for hostName in res:
                if totalHosts == 1:
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                    if hostName == tmp:
                        driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[3]').click()
                    else:
                        self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + sourceWave + " **********")
                        continue
                else:
                    count = 0
                    for hostNo in range(1, totalHosts + 1):
                        tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                        if hostName == tmp:
                            driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[8]/span/div/i[3]').click()
                            break
                        count += 1
                    if count > totalHosts:
                        self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + sourceWave + " **********")
                        continue
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.pop_moveHost_xpath)) != 0:
                    self.logger.info("********** Move Machine To New Wave Pop-up Banner Was Opened For Host : " + hostName + " **********")
                    env = Select(driver.find_element(By.XPATH, LOC.drp_selectWave_xpath))
                    env.select_by_visible_text(targetWave)
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, LOC.btn_moveMachine_id))
                    )
                    btn.click()
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                    )
                    note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                    self.logger.info("********** Changed Target Type For Host : " + hostName + ",")
                    self.logger.info(note + "\n")
                    time.sleep(2)
                    driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
                else:
                    self.logger.info("********** Move Machine To New Wave Pop-up Banner Was Not Opened For Host : " + hostName + " **********")

    def changeVcenterData(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        for r in range(st, ed + 1):
            waveName = sheet.cell(row=r, column=1).value
            hostNames = sheet.cell(row=r, column=2).value
            co = CommonObjects(driver)
            val = co.findWave(waveName)
            if val == 2:
                continue
            if val == 1:
                driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
                else:
                    self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                    continue
            if val == 0:
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
            if type(hostNames) != str:
                self.changeBulkVcenterData(driver, path, r)
            else:
                res = tuple(map(str, hostNames.split(', ')))
                if len(res) == 1:
                    self.changeSingleVcenterData(driver, path, r)
                else:
                    self.changeBulkVcenterData(driver, path, r)

    def changeSingleVcenterData(self, driver, path, r):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active

        waveName = sheet.cell(row=r, column=1).value
        hostName = sheet.cell(row=r, column=2).value
        clusterName = sheet.cell(row=r, column=3).value
        ESXHost = sheet.cell(row=r, column=4).value
        datastore = sheet.cell(row=r, column=5).value
        VMFolder = sheet.cell(row=r, column=6).value
        resourcePool = sheet.cell(row=r, column=7).value
        routes = sheet.cell(row=r, column=8).value

        totalHosts = len(driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
        if totalHosts == 1:
            tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
            if hostName == tmp:
                driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[2]').click()
            else:
                self.logger.info("********** The Host " + str(hostName) + " Was Not Found In The Wave : " + waveName + " **********")
                return
        else:
            count = 0
            for hostNo in range(1, totalHosts + 1):
                tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                if hostName == tmp:
                    driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[8]/span/div/i[2]').click()
                    break
                count += 1
            if count > totalHosts:
                self.logger.info("********** The Host " + str(hostName) + " Was Not Found In The Wave : " + waveName + " **********")
                return
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_editHost_xpath)) != 0:
            self.logger.info("********** Wave Edit Sync Options Pop-up Banner Was Opened For Wave, " + waveName + " **********")
            driver.find_element(By.LINK_TEXT, "vCenter Options").click()
            time.sleep(5)
            vCenter_class = driver.find_element(By.XPATH, LOC.txt_editCU_xpath).get_attribute("class")
            if vCenter_class == "ui-state-default ui-corner-top ng-star-inserted ui-tabview-selected ui-state-active":
                self.logger.info("********** Wave Edit vCenter Options Pop-up Banner Was Opened For Wave, " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.txt_clusterName_xpath).clear()
                driver.find_element(By.XPATH, LOC.txt_clusterName_xpath).send_keys(clusterName)
                driver.find_element(By.XPATH, LOC.txt_esxHost_xpath).clear()
                driver.find_element(By.XPATH, LOC.txt_esxHost_xpath).send_keys(ESXHost)
                driver.find_element(By.XPATH, LOC.txt_dataStore_xpath).clear()
                driver.find_element(By.XPATH, LOC.txt_dataStore_xpath).send_keys(datastore)
                if VMFolder != "NA":
                    driver.find_element(By.XPATH, LOC.txt_vmFolder_xpath).clear()
                    driver.find_element(By.XPATH, LOC.txt_vmFolder_xpath).send_keys(VMFolder)
                if resourcePool != "NA":
                    driver.find_element(By.XPATH, LOC.txt_resourcePool_xpath).clear()
                    driver.find_element(By.XPATH, LOC.txt_resourcePool_xpath).send_keys(resourcePool)
                if routes != "NA":
                    driver.find_element(By.XPATH, LOC.txt_routes_xpath).clear()
                    driver.find_element(By.XPATH, LOC.txt_routes_xpath).send_keys(routes)
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_modify_id))
                )
                btn.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Edit VCEnter Options Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Wave Edit vCenter Options Pop-up Banner Was Not Opened For Wave, " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.btn_cancelSingleVC_xpath).click()
        else:
            self.logger.info("********** Wave Edit Sync Options Pop-up Banner Was Not Opened For Wave, " + waveName + " **********")

    def changeBulkVcenterData(self, driver, path, r):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active

        waveName = sheet.cell(row=r, column=1).value
        hostNames = sheet.cell(row=r, column=2).value
        clusterName = sheet.cell(row=r, column=3).value
        ESXHost = sheet.cell(row=r, column=4).value
        datastore = sheet.cell(row=r, column=5).value
        VMFolder = sheet.cell(row=r, column=6).value
        resourcePool = sheet.cell(row=r, column=7).value
        routes = sheet.cell(row=r, column=8).value

        if type(hostNames) != str:
            driver.find_element(By.XPATH, LOC.btn_selectAll_xpath).click()
        else:
            res = tuple(map(str, hostNames.split(', ')))
            totalHosts = len(driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
            for hostName in res:
                if totalHosts == 1:
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                    if hostName == tmp:
                        driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    else:
                        self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
                else:
                    for hostNo in range(1, totalHosts + 1):
                        tmp = driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
                        if hostName == tmp:
                            driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[1]/p-tablecheckbox/div/div[2]').click()
                            break
                        if hostNo == totalHosts:
                            self.logger.info("********** The Host " + str(hostName) + " Was Not Found In The Wave : " + waveName + " **********")
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_bulkEdit_xpath))
        )
        btn.click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_bulkEdit_xpath)) != 0:
            self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Was Opened For Wave, " + waveName + " **********")
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_vCenterOption_xpath))
            )
            btn.click()
            vCenter_class = driver.find_element(By.XPATH, LOC.txt_bulkEditCU_xpath).get_attribute("class")
            if vCenter_class == "ui-state-default ui-corner-top ng-star-inserted ui-tabview-selected ui-state-active":
                self.logger.info("********** Wave Edit vCenter Options Pop-up Banner Was Opened For Wave, " + waveName + " **********")
                cn = Select(driver.find_element(By.ID, LOC.drp_clusterName_id))
                cn.select_by_visible_text(clusterName)
                esx = Select(driver.find_element(By.ID, LOC.drp_esxHost_id))
                esx.select_by_visible_text(ESXHost)
                ds = Select(driver.find_element(By.ID, LOC.drp_datastore_id))
                ds.select_by_visible_text(datastore)
                if VMFolder != "NA":
                    driver.find_element(By.ID, LOC.txt_vmFolder_id).clear()
                    driver.find_element(By.ID, LOC.txt_vmFolder_id).send_keys(VMFolder)
                if resourcePool != "NA":
                    driver.find_element(By.ID, LOC.txt_resourcePool_id).clear()
                    driver.find_element(By.ID, LOC.txt_resourcePool_id).send_keys(resourcePool)
                if routes != "NA":
                    driver.find_element(By.ID, LOC.txt_routes_id).clear()
                    driver.find_element(By.ID, LOC.txt_routes_id).send_keys(routes)
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_next_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_modifyAll_xpath))
                )
                btn.click()
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_yes_id))
                )
                btn.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Edit VCEnter Options Status of Wave : " + waveName + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Wave Edit vCenter Options Pop-up Banner Was Not Opened For Wave, " + waveName + " **********")
                driver.find_element(By.XPATH, LOC.btn_cancelVCEdit_xpath).click()
        else:
            self.logger.info("********** Wave Bulk Edit Sync Options Pop-up Banner Was Not Opened For Wave, " + waveName + " **********")
        time.sleep(3)
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
        )
        btn.click()
        if type(hostNames) == str:
            time.sleep(2)
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_selectAll_xpath))
            )
            btn.click()
