import time
import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from utilities.customLogger import LogGen


class Configuration:
    # Common Data for Cloud Users
    btn_add_id = "conf_cu_add_btn"
    btn_edit_id = "conf_cu_edit_btn"
    btn_delete_id = "conf_cu_del_btn"
    txt_name_id = "UserId"
    drp_cloudProvider_xpath = "//*[@id='conf_cu_add_cloud_modal']/div/div/div/form/div[2]/div/div/div[2]/div/select"
    btn_confirm_id = "conf_cu_add_cloud_modal_submit_btn"
    btn_cancel_id = "conf_cu_add_cloud_modal_cancel_btn"

    # AWS
    txt_AWSAccessKey_id = "configuration_clouduser_add_cloud_modal_aws_accessKey"
    txt_AWSSecretAccessKey_id = "configuration_clouduser_add_cloud_modal_aws_secretAccessKey"

    # Azure
    txt_AzureSubscriptionId_id = "configuration_clouduser_add_cloud_modal_azure_subscriptionId"
    txt_AzureTenantId_id = "configuration_clouduser_add_cloud_modal_azure_tenantId"
    txt_AzureClientId_id = "configuration_clouduser_add_cloud_modal_azure_clientId"
    txt_AzureClientSecret_id = "configuration_clouduser_add_cloud_modal_azure_clientSecret"
    drp_AzureCloudType_xpath = "//*[@id='configuration_clouduser_add_cloud_modal_azure_datacenter']"
    drp_AzureDataCentre_xpath = "//*[@id='configuration_clouduser_add_cloud_modal_azure_datacenter']/div/input"

    # Google
    rd_GGLUploadLocal_xpath = "//*[@id='conf_cu_add_cloud_modal']/div/div/div/form/div[2]/div/div/div[3]/div[1]/div[1]/div[1]/div/div/input"
    rd_GGLFilePath_xpath = "//*[@id='conf_cu_add_cloud_modal']/div/div/div/form/div[2]/div/div/div[3]/div[1]/div[1]/div[2]/div/div/input"
    src_GGLBrowse_xpath = "//*[@id='browse_dropzone']/div/div/label/i"
    txt_GGLPathOnRMM_id = "gcpFilePath"
    txt_GGLProjectId_id = "projectid"

    # IBM Cloud VPC
    drp_IBMRegion_xpath = "//*[@id='configuration_clouduser_add_cloud_modal_ibmgen2_region']/div/input"
    txt_IBMapiKey_id = "apikey"

    # CloudStack
    txt_CSApiUrl_id = "cloudstack_apiUrl"
    txt_CSApiKey_id = "cloudstack_apiKey"
    txt_CSSecretKey_id = "cloudstack_secretKey"
    txt_CSDomainId_id = "cloudstack_domainid"

    # OCI
    txt_OCIUserId_id = "oci_userId"
    rd_OCIUploadFile_id = "ociPrivateKeySourceUpload local File"
    rd_OCIFilePath_xpath = "ociPrivateKeySourceFile path on RMM"
    src_OCIBrowse_xpath = "//*[@id='browse_dropzone']/div/div/label/i"
    txt_OCIPathOnRMM_id = "oci_pkFilePath"
    txt_OCIFingerprint_id = "oci_fingerprint"
    txt_OCITenantId_id = "oci_tenantId"
    txt_OCIPassphrase_id = "oci_passphrase"
    drp_OCIRegion_xpath = "//*[@id='configuration_clouduser_add_cloud_modal_oci_datacenter']/div/input"
    txt_OCIApiUrl_id = "oci_apiUrl"
    rd_OCIParameterName_id = "ociParamTypeName"
    txt_OCICompartmentName_id = "oci_compartment_name"
    rd_OCIParameterId_id = "ociParamTypeID"
    txt_OCICompartmentId_id = "oci_compartment_id"
    rd_OCICertUploadFile_id = "ociCertificateSourceUpload local File"
    rd_OCICertFilePath_xpath = "ociCertificateSourceUpload local File"
    src_OCICertBrowse_xpath = "//*[@id='browse_dropzone']/div/div/label"
    txt_OCICertPathOnRMM_id = "oci_certificateSourcePath"

    # Softlayer
    txt_SLUserName_id = "userName"
    txt_SLApiKey_id = "apiKey"
    txt_SlDomainName_id = "domainName"
    txt_SLAccessRights_id = "accessrights"
    chBox_SLHourly_id = "check-hourly"

    # Zadara
    txt_ZadaraAccessKey_id = "configuration_clouduser_add_cloud_modal_zadara_accessKey"
    txt_ZadaraSecretAccessKey_id = "configuration_clouduser_add_cloud_modal_zadara_secretAccessKey"
    txt_ZadaraApiUrl_id = "zadara_apirul"
    txt_ZadaraRegion_id = "zadara_region"

    # VCenter data
    btn_createVC_id = "conf_vc_add_btn"
    btn_editVC_id = "conf_vc_edit_btn"
    btn_deleteVC_id = "conf_vc_del_btn"
    txt_VCName_id = "name"
    txt_VCipAddress_id = "address"
    txt_VCUserName_id = "username"
    txt_VCPassword_id = "password"
    txt_VCPort_id = "port"
    btn_addVC_id = "conf_vc_vc_add_modal_submit_btn"

    # Organization
    btn_addAdminOrg_xpath = "//*[@id='content']/div/article/div/div[2]/div[2]/tree-root/tree-viewport/div/div/tree-node-collection/div/tree-node/div/tree-node-wrapper/div[1]/div/span/i[1]"
    btn_addAdminUser_xpath = "//*[@id='content']/div/article/div/div[2]/div[2]/tree-root/tree-viewport/div/div/tree-node-collection/div/tree-node/div/tree-node-wrapper/div[1]/div/span/i[2]"
    btn_editAdminOrg_xpath = "//*[@id='content']/div/article/div/div[2]/div[2]/tree-root/tree-viewport/div/div/tree-node-collection/div/tree-node/div/tree-node-wrapper/div[1]/div/span/i[3]"

    # Pop-up Banners
    pop_addCloudUser_xpath = '//*[@id="conf_cu_add_cloud_modal"]/div/div/div/form/div[1]/h4'
    pop_addVC_xpath = '//*[@id="conf_vc_add_vc_modal_btn"]/div/div/div/form/div[1]/h4'
    not_addVC_xpath = '/html/body/app-root/simple-notifications/div/simple-notification[2]/div'

    logger = LogGen.loggen()

    def __init__(self, driver):
        self.driver = driver

    def addNewCloudUser(self, path):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row

        self.driver.find_element(By.LINK_TEXT, "Configuration").click()
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Clouduser").click()
        time.sleep(5)

        for r in range(3, rows+1):
            time.sleep(5)
            self.driver.find_element(By.ID, self.btn_add_id).click()
            time.sleep(5)

            name = sheet.cell(row=r, column=1).value
            cloudProvider = sheet.cell(row=r, column=2).value

            if len(self.driver.find_elements(By.XPATH, self.pop_addCloudUser_xpath)) != 0:
                self.logger.info("********** Pop-up Banner For Add Clouduser Was Opened **********")

                self.driver.find_element(By.ID, self.txt_name_id).send_keys(name)
                cp = Select(self.driver.find_element(By.XPATH, self.drp_cloudProvider_xpath))
                cp.select_by_visible_text(cloudProvider)

                if cloudProvider == "AWS":
                    self.logger.info("********** AWS Was Selected As Cloud Provider **********")
                    accessKey = sheet.cell(row=r, column=3).value
                    secretAccessKey = sheet.cell(row=r, column=4).value

                    self.driver.find_element(By.ID, self.txt_AWSAccessKey_id).send_keys(accessKey)
                    self.driver.find_element(By.ID, self.txt_AWSSecretAccessKey_id).send_keys(secretAccessKey)

                elif cloudProvider == "Azure":
                    self.logger.info("********** Azure Was Selected As Cloud Provider **********")
                    subscriptionId = sheet.cell(row=r, column=5).value
                    tenantId = sheet.cell(row=r, column=6).value
                    clientId = sheet.cell(row=r, column=7).value
                    clientSecret = sheet.cell(row=r, column=8).value
                    cloudType = sheet.cell(row=r, column=9).value
                    dataCentre = sheet.cell(row=r, column=10).value

                    self.driver.find_element(By.ID, self.txt_AzureSubscriptionId_id).send_keys(subscriptionId)
                    self.driver.find_element(By.ID, self.txt_AzureTenantId_id).send_keys(tenantId)
                    self.driver.find_element(By.ID, self.txt_AzureClientId_id).send_keys(clientId)
                    self.driver.find_element(By.ID, self.txt_AzureClientSecret_id).send_keys(clientSecret)
                    ct = Select(self.driver.find_element(By.XPATH, self.drp_AzureCloudType_xpath))
                    ct.select_by_visible_text(cloudType)
                    self.driver.find_element(By.XPATH, self.drp_AzureDataCentre_xpath).send_keys(dataCentre)

                elif cloudProvider == "Google":
                    self.logger.info("********** Google Was Selected As Cloud Provider **********")
                    fileUploadMethod = sheet.cell(row=r, column=11).value
                    filePath = sheet.cell(row=r, column=12).value
                    projectId = sheet.cell(row=r, column=13).value

                    if fileUploadMethod == "Upload local File":
                        self.driver.find_element(By.XPATH, self.rd_GGLUploadLocal_xpath).click()
                        self.driver.find_element(By.XPATH, self.src_GGLBrowse_xpath).send_keys(filePath)
                    elif fileUploadMethod == "File path on RMM":
                        self.driver.find_element(By.XPATH, self.rd_GGLFilePath_xpath).click()
                        self.driver.find_element(By.ID, self.txt_GGLPathOnRMM_id).send_keys(filePath)
                    self.driver.find_element(By.ID, self.txt_GGLProjectId_id).send_keys(projectId)

                elif cloudProvider == "IBM Cloud VPC":
                    self.logger.info("********** IBM Cloud VPC Was Selected As Cloud Provider **********")
                    region = sheet.cell(row=r, column=14).value
                    apiKey = sheet.cell(row=r, column=15).value

                    time.sleep(5)
                    self.driver.find_element(By.XPATH, self.drp_IBMRegion_xpath).send_keys(region)
                    self.driver.find_element(By.ID, self.txt_IBMapiKey_id).send_keys(apiKey)

                elif cloudProvider == "CloudStack":
                    self.logger.info("********** CloudStack Was Selected As Cloud Provider **********")
                    apiUrl = sheet.cell(row=r, column=16).value
                    apiKey = sheet.cell(row=r, column=17).value
                    secretKey = sheet.cell(row=r, column=18).value
                    domainId = sheet.cell(row=r, column=19).value

                    self.driver.find_element(By.ID, self.txt_CSApiUrl_id).clear()
                    self.driver.find_element(By.ID, self.txt_CSApiUrl_id).send_keys(apiUrl)
                    self.driver.find_element(By.ID, self.txt_CSApiKey_id).clear()
                    self.driver.find_element(By.ID, self.txt_CSApiKey_id).send_keys(apiKey)
                    self.driver.find_element(By.ID, self.txt_CSSecretKey_id).send_keys(secretKey)
                    self.driver.find_element(By.ID, self.txt_CSDomainId_id).send_keys(domainId)

                elif cloudProvider == "OCI":
                    self.logger.info("********** OCI Was Selected As Cloud Provider **********")
                    userId = sheet.cell(row=r, column=20).value
                    fileUploadMethod = sheet.cell(row=r, column=21).value
                    filePath = sheet.cell(row=r, column=22).value
                    fingerprint = sheet.cell(row=r, column=23).value
                    tenantId = sheet.cell(row=r, column=24).value
                    passphrase = sheet.cell(row=r, column=25).value
                    region = sheet.cell(row=r, column=26).value
                    apiUrl = sheet.cell(row=r, column=27).value
                    parameterType = sheet.cell(row=r, column=28).value
                    compartmentName = sheet.cell(row=r, column=29).value
                    certUploadMethod = sheet.cell(row=r, column=30).value
                    certFilePath = sheet.cell(row=r, column=31).value

                    self.driver.find_element(By.ID, self.txt_OCIUserId_id).send_keys(userId)
                    if fileUploadMethod == "Upload local File":
                        self.driver.find_element(By.ID, self.rd_OCIUploadFile_id).click()
                        self.driver.find_element(By.XPATH, self.src_OCIBrowse_xpath).send_keys(filePath)
                    elif fileUploadMethod == "File path on RMM":
                        self.driver.find_element(By.XPATH, self.rd_OCIFilePath_xpath).click()
                        self.driver.find_element(By.ID, self.txt_OCIPathOnRMM_id).send_keys(filePath)
                    self.driver.find_element(By.ID, self.txt_OCIFingerprint_id).send_keys(fingerprint)
                    self.driver.find_element(By.ID, self.txt_OCITenantId_id).send_keys(tenantId)
                    self.driver.find_element(By.ID, self.txt_OCIPassphrase_id).send_keys(passphrase)
                    self.driver.find_element(By.XPATH, self.drp_OCIRegion_xpath).send_keys(region)
                    self.driver.find_element(By.ID, self.txt_OCIApiUrl_id).send_keys(apiUrl)
                    if parameterType == "Name":
                        self.driver.find_element(By.ID, self.rd_OCIParameterName_id).click()
                        self.driver.find_element(By.ID, self.txt_OCICompartmentName_id).send_keys(compartmentName)
                    elif parameterType == "ID":
                        self.driver.find_element(By.ID, self.rd_OCIParameterId_id).click()
                        self.driver.find_element(By.ID, self.txt_OCICompartmentId_id).send_keys(compartmentName)
                    if certUploadMethod == "Upload local File":
                        self.driver.find_element(By.XPATH, self.rd_OCICertUploadFile_id).click()
                        self.driver.find_element(By.XPATH, self.src_OCICertBrowse_xpath).send_keys(certFilePath)
                    elif certUploadMethod == "File path on RMM":
                        self.driver.find_element(By.XPATH, self.rd_OCICertFilePath_xpath).click()
                        self.driver.find_element(By.XPATH, self.txt_OCICertPathOnRMM_id).send_keys(certFilePath)

                elif cloudProvider == "Softlayer":
                    self.logger.info("********** Softlayer Was Selected As Cloud Provider **********")
                    userName = sheet.cell(row=r, column=32).value
                    apiKey = sheet.cell(row=r, column=33).value
                    domainName = sheet.cell(row=r, column=34).value
                    accessRight = sheet.cell(row=r, column=35).value
                    hourly = sheet.cell(row=r, column=36).value

                    self.driver.find_element(By.ID, self.txt_SLUserName_id).send_keys(userName)
                    self.driver.find_element(By.ID, self.txt_SLApiKey_id).send_keys(apiKey)
                    self.driver.find_element(By.ID, self.txt_SlDomainName_id).send_keys(domainName)
                    self.driver.find_element(By.ID, self.txt_SLAccessRights_id).send_keys(accessRight)
                    if hourly:
                        self.driver.find_element(By.ID, self.chBox_SLHourly_id).click()

                elif cloudProvider == "Zadara":
                    self.logger.info("********** Zadara Was Selected As Cloud Provider **********")
                    accessKey = sheet.cell(row=r, column=37).value
                    secretAccessKey = sheet.cell(row=r, column=38).value
                    apiUrl = sheet.cell(row=r, column=39).value
                    region = sheet.cell(row=r, column=40).value

                    self.driver.find_element(By.ID, self.txt_ZadaraAccessKey_id).send_keys(accessKey)
                    self.driver.find_element(By.ID, self.txt_ZadaraSecretAccessKey_id).send_keys(secretAccessKey)
                    self.driver.find_element(By.ID, self.txt_ZadaraApiUrl_id).send_keys(apiUrl)
                    self.driver.find_element(By.ID, self.txt_ZadaraRegion_id).send_keys(region)

                time.sleep(5)
                self.driver.find_element(By.ID, self.btn_confirm_id).click()
                time.sleep(5)
                val = self.findClouduser(name)
                if val == 1:
                    self.logger.info("********** "+name+" Clouduser Was Added Successfully **********")
                else:
                    self.logger.info("********** Failed To Add "+name+" Clouduser **********")
            else:
                self.logger.info("********** Failed To Open Pop-up Banner For "+name+" Clouduser **********")
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, "Summary")) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def findClouduser(self, userName):
        flag = 0
        totalCU = len(self.driver.find_elements(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr'))
        for i in range(1, totalCU+1):
            tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span[1]').text
            if tmp == userName:
                flag += 1
        return flag

    def addVCenter(self, path):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row

        self.driver.find_element(By.LINK_TEXT, "Configuration").click()
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "vCenter").click()
        time.sleep(5)

        for r in range(3, rows + 1):
            name = sheet.cell(row=r, column=1).value
            ipAddress = sheet.cell(row=r, column=2).value
            userName = sheet.cell(row=r, column=3).value
            password = sheet.cell(row=r, column=4).value
            portNumber = sheet.cell(row=r, column=5).value

            element = WebDriverWait(self.driver, 30).until(
                    EC.element_to_be_clickable((By.ID, self.btn_createVC_id))
            )
            element.click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_addVC_xpath)) != 0:
                self.logger.info("********** Add vCenter Pop-up Banner Is Opened For "+str(name)+" vCenter **********")
                self.driver.find_element(By.ID, self.txt_VCName_id).send_keys(name)
                self.driver.find_element(By.ID, self.txt_VCipAddress_id).send_keys(ipAddress)
                self.driver.find_element(By.ID, self.txt_VCUserName_id).send_keys(userName)
                self.driver.find_element(By.ID, self.txt_VCPassword_id).send_keys(password)
                if portNumber != "NA":
                    self.driver.find_element(By.ID, self.txt_VCPort_id).send_keys(portNumber)
                time.sleep(5)
                self.driver.find_element(By.ID, self.btn_addVC_id).click()
                time.sleep(5)
                val = self.verify(name)
                if val == 1:
                    self.logger.info("********** vCenter " + str(name) + " Added Successfully **********")
                else:
                    self.logger.info("********** Failed To Add vCenter " + str(name) + " **********")
            else:
                self.logger.info("********** Add vCenter Pop-up Banner Is Not Opened For "+str(name)+" vCenter **********")
        if len(self.driver.find_elements(By.LINK_TEXT, "Waves")) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def verify(self, name):
        note = self.driver.find_element(By.XPATH, self.not_addVC_xpath).text
        self.logger.info(note + "\n")
        res = tuple(map(str, note.split(' ')))
        flag = 0
        time.sleep(10)
        if res[0] != "Failed":
            totalUser = len(self.driver.find_elements(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr'))
            for i in range(1, totalUser + 1):
                if totalUser == 1:
                    tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span[1]').text
                else:
                    tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span[1]').text
                if name == tmp:
                    flag += 1
                    break
        time.sleep(5)
        return flag
