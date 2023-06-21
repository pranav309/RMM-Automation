import time
import unittest
import openpyxl

import Locators.locConfiguration as LOC

from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from utilities.customLogger import LogGen


class Configuration(unittest.TestCase):

    logger = LogGen.loggen()

    def addNewCloudUser(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 3
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        conf_class = driver.find_element(By.XPATH, LOC.txt_config_xpath).get_attribute("class")
        if conf_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Configuration").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.txt_cu_xpath))
        )
        cu_class = driver.find_element(By.XPATH, LOC.txt_cu_xpath).get_attribute("class")
        if cu_class != "active":
            driver.find_element(By.LINK_TEXT, "Clouduser").click()
            time.sleep(5)

        for r in range(st, ed+1):
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, LOC.btn_add_id))
            )
            driver.find_element(By.ID, LOC.btn_add_id).click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, LOC.pop_addCloudUser_xpath))
            )
            name = sheet.cell(row=r, column=1).value
            cloudProvider = sheet.cell(row=r, column=2).value
            if len(driver.find_elements(By.XPATH, LOC.pop_addCloudUser_xpath)) != 0:
                self.logger.info("********** Pop-up Banner For Add Clouduser Was Opened **********")
                driver.find_element(By.ID, LOC.txt_name_id).send_keys(name)
                cp = Select(driver.find_element(By.XPATH, LOC.drp_cloudProvider_xpath))
                cp.select_by_visible_text(cloudProvider)

                if cloudProvider == "AWS":
                    self.logger.info("********** AWS Was Selected As Cloud Provider **********")
                    accessKey = sheet.cell(row=r, column=3).value
                    secretAccessKey = sheet.cell(row=r, column=4).value
                    driver.find_element(By.ID, LOC.txt_AWSAccessKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_AWSAccessKey_id).send_keys(accessKey)
                    driver.find_element(By.ID, LOC.txt_AWSSecretAccessKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_AWSSecretAccessKey_id).send_keys(secretAccessKey)
                elif cloudProvider == "Azure":
                    self.logger.info("********** Azure Was Selected As Cloud Provider **********")
                    subscriptionId = sheet.cell(row=r, column=5).value
                    tenantId = sheet.cell(row=r, column=6).value
                    clientId = sheet.cell(row=r, column=7).value
                    clientSecret = sheet.cell(row=r, column=8).value
                    cloudType = sheet.cell(row=r, column=9).value
                    dataCentre = sheet.cell(row=r, column=10).value

                    driver.find_element(By.ID, LOC.txt_AzureSubscriptionId_id).clear()
                    driver.find_element(By.ID, LOC.txt_AzureSubscriptionId_id).send_keys(subscriptionId)
                    driver.find_element(By.ID, LOC.txt_AzureTenantId_id).clear()
                    driver.find_element(By.ID, LOC.txt_AzureTenantId_id).send_keys(tenantId)
                    driver.find_element(By.ID, LOC.txt_AzureClientId_id).clear()
                    driver.find_element(By.ID, LOC.txt_AzureClientId_id).send_keys(clientId)
                    driver.find_element(By.ID, LOC.txt_AzureClientSecret_id).clear()
                    driver.find_element(By.ID, LOC.txt_AzureClientSecret_id).send_keys(clientSecret)
                    ct = Select(driver.find_element(By.XPATH, LOC.drp_AzureCloudType_xpath))
                    ct.select_by_visible_text(cloudType)
                    driver.find_element(By.XPATH, LOC.drp_AzureDataCentre_xpath).clear()
                    driver.find_element(By.XPATH, LOC.drp_AzureDataCentre_xpath).send_keys(dataCentre)

                elif cloudProvider == "Google":
                    self.logger.info("********** Google Was Selected As Cloud Provider **********")
                    fileUploadMethod = sheet.cell(row=r, column=11).value
                    filePath = sheet.cell(row=r, column=12).value
                    projectId = sheet.cell(row=r, column=13).value

                    if fileUploadMethod == "Upload local File":
                        driver.find_element(By.XPATH, LOC.rd_GGLUploadLocal_xpath).click()
                        driver.find_element(By.XPATH, LOC.src_GGLBrowse_xpath).clear()
                        driver.find_element(By.XPATH, LOC.src_GGLBrowse_xpath).send_keys(filePath)
                    elif fileUploadMethod == "File path on RMM":
                        driver.find_element(By.XPATH, LOC.rd_GGLFilePath_xpath).click()
                        driver.find_element(By.ID, LOC.txt_GGLPathOnRMM_id).clear()
                        driver.find_element(By.ID, LOC.txt_GGLPathOnRMM_id).send_keys(filePath)
                    driver.find_element(By.ID, LOC.txt_GGLProjectId_id).clear()
                    driver.find_element(By.ID, LOC.txt_GGLProjectId_id).send_keys(projectId)

                elif cloudProvider == "IBM Cloud VPC":
                    self.logger.info("********** IBM Cloud VPC Was Selected As Cloud Provider **********")
                    region = sheet.cell(row=r, column=14).value
                    apiKey = sheet.cell(row=r, column=15).value

                    driver.find_element(By.XPATH, LOC.drp_IBMRegion_xpath).clear()
                    driver.find_element(By.XPATH, LOC.drp_IBMRegion_xpath).send_keys(region)
                    driver.find_element(By.ID, LOC.txt_IBMapiKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_IBMapiKey_id).send_keys(apiKey)

                elif cloudProvider == "CloudStack":
                    self.logger.info("********** CloudStack Was Selected As Cloud Provider **********")
                    apiUrl = sheet.cell(row=r, column=16).value
                    apiKey = sheet.cell(row=r, column=17).value
                    secretKey = sheet.cell(row=r, column=18).value
                    domainId = sheet.cell(row=r, column=19).value

                    driver.find_element(By.ID, LOC.txt_CSApiUrl_id).clear()
                    driver.find_element(By.ID, LOC.txt_CSApiUrl_id).send_keys(apiUrl)
                    driver.find_element(By.ID, LOC.txt_CSApiKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_CSApiKey_id).send_keys(apiKey)
                    driver.find_element(By.ID, LOC.txt_CSSecretKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_CSSecretKey_id).send_keys(secretKey)
                    driver.find_element(By.ID, LOC.txt_CSDomainId_id).clear()
                    driver.find_element(By.ID, LOC.txt_CSDomainId_id).send_keys(domainId)

                elif cloudProvider == "OCI":
                    self.logger.info("********** OCI Was Selected As Cloud Provider **********")
                    userId = sheet.cell(row=r, column=20).value
                    fileUploadMethod = sheet.cell(row=r, column=21).value
                    filePath = sheet.cell(row=r, column=22).value
                    fingerprint = sheet.cell(row=r, column=23).value
                    tenantId = sheet.cell(row=r, column=24).value
                    passphrase = sheet.cell(row=r, column=25).value
                    region = sheet.cell(row=r, column=26).value
                    apiUrl = sheet.cell(row=r, column=27).value
                    parameterType = sheet.cell(row=r, column=28).value
                    compartmentName = sheet.cell(row=r, column=29).value
                    certUploadMethod = sheet.cell(row=r, column=30).value
                    certFilePath = sheet.cell(row=r, column=31).value

                    driver.find_element(By.ID, LOC.txt_OCIUserId_id).clear()
                    driver.find_element(By.ID, LOC.txt_OCIUserId_id).send_keys(userId)
                    if fileUploadMethod == "Upload local File":
                        driver.find_element(By.ID, LOC.rd_OCIUploadFile_id).click()
                        driver.find_element(By.XPATH, LOC.src_OCIBrowse_xpath).clear()
                        driver.find_element(By.XPATH, LOC.src_OCIBrowse_xpath).send_keys(filePath)
                    elif fileUploadMethod == "File path on RMM":
                        driver.find_element(By.XPATH, LOC.rd_OCIFilePath_xpath).click()
                        driver.find_element(By.ID, LOC.txt_OCIPathOnRMM_id).clear()
                        driver.find_element(By.ID, LOC.txt_OCIPathOnRMM_id).send_keys(filePath)
                    driver.find_element(By.ID, LOC.txt_OCIFingerprint_id).clear()
                    driver.find_element(By.ID, LOC.txt_OCIFingerprint_id).send_keys(fingerprint)
                    driver.find_element(By.ID, LOC.txt_OCITenantId_id).clear()
                    driver.find_element(By.ID, LOC.txt_OCITenantId_id).send_keys(tenantId)
                    driver.find_element(By.ID, LOC.txt_OCIPassphrase_id).clear()
                    driver.find_element(By.ID, LOC.txt_OCIPassphrase_id).send_keys(passphrase)
                    driver.find_element(By.XPATH, LOC.drp_OCIRegion_xpath).clear()
                    driver.find_element(By.XPATH, LOC.drp_OCIRegion_xpath).send_keys(region)
                    driver.find_element(By.ID, LOC.txt_OCIApiUrl_id).clear()
                    driver.find_element(By.ID, LOC.txt_OCIApiUrl_id).send_keys(apiUrl)
                    if parameterType == "Name":
                        driver.find_element(By.ID, LOC.rd_OCIParameterName_id).click()
                        driver.find_element(By.ID, LOC.txt_OCICompartmentName_id).clear()
                        driver.find_element(By.ID, LOC.txt_OCICompartmentName_id).send_keys(compartmentName)
                    elif parameterType == "ID":
                        driver.find_element(By.ID, LOC.rd_OCIParameterId_id).click()
                        driver.find_element(By.ID, LOC.txt_OCICompartmentId_id).clear()
                        driver.find_element(By.ID, LOC.txt_OCICompartmentId_id).send_keys(compartmentName)
                    if certUploadMethod == "Upload local File":
                        driver.find_element(By.XPATH, LOC.rd_OCICertUploadFile_id).click()
                        driver.find_element(By.XPATH, LOC.src_OCICertBrowse_xpath).clear()
                        driver.find_element(By.XPATH, LOC.src_OCICertBrowse_xpath).send_keys(certFilePath)
                    elif certUploadMethod == "File path on RMM":
                        driver.find_element(By.XPATH, LOC.rd_OCICertFilePath_xpath).click()
                        driver.find_element(By.XPATH, LOC.txt_OCICertPathOnRMM_id).clear()
                        driver.find_element(By.XPATH, LOC.txt_OCICertPathOnRMM_id).send_keys(certFilePath)

                elif cloudProvider == "Softlayer":
                    self.logger.info("********** Softlayer Was Selected As Cloud Provider **********")
                    userName = sheet.cell(row=r, column=32).value
                    apiKey = sheet.cell(row=r, column=33).value
                    domainName = sheet.cell(row=r, column=34).value
                    accessRight = sheet.cell(row=r, column=35).value
                    hourly = sheet.cell(row=r, column=36).value

                    driver.find_element(By.ID, LOC.txt_SLUserName_id).clear()
                    driver.find_element(By.ID, LOC.txt_SLUserName_id).send_keys(userName)
                    driver.find_element(By.ID, LOC.txt_SLApiKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_SLApiKey_id).send_keys(apiKey)
                    driver.find_element(By.ID, LOC.txt_SlDomainName_id).clear()
                    driver.find_element(By.ID, LOC.txt_SlDomainName_id).send_keys(domainName)
                    driver.find_element(By.ID, LOC.txt_SLAccessRights_id).clear()
                    driver.find_element(By.ID, LOC.txt_SLAccessRights_id).send_keys(accessRight)
                    if hourly:
                        driver.find_element(By.ID, LOC.chBox_SLHourly_id).click()

                elif cloudProvider == "Zadara":
                    self.logger.info("********** Zadara Was Selected As Cloud Provider **********")
                    accessKey = sheet.cell(row=r, column=37).value
                    secretAccessKey = sheet.cell(row=r, column=38).value
                    apiUrl = sheet.cell(row=r, column=39).value
                    region = sheet.cell(row=r, column=40).value

                    driver.find_element(By.ID, LOC.txt_ZadaraAccessKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_ZadaraAccessKey_id).send_keys(accessKey)
                    driver.find_element(By.ID, LOC.txt_ZadaraSecretAccessKey_id).clear()
                    driver.find_element(By.ID, LOC.txt_ZadaraSecretAccessKey_id).send_keys(secretAccessKey)
                    driver.find_element(By.ID, LOC.txt_ZadaraApiUrl_id).clear()
                    driver.find_element(By.ID, LOC.txt_ZadaraApiUrl_id).send_keys(apiUrl)
                    driver.find_element(By.ID, LOC.txt_ZadaraRegion_id).clear()
                    driver.find_element(By.ID, LOC.txt_ZadaraRegion_id).send_keys(region)

                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_confirm_id))
                )
                driver.find_element(By.ID, LOC.btn_confirm_id).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_deleteSuccessful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).text
                self.logger.info("********** Add New Clouduser Status For Clouduser: " + name + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Failed To Open Pop-up Banner For "+name+" Clouduser **********")

    def findClouduser(self, driver, cuNames):
        conf_class = driver.find_element(By.XPATH, LOC.txt_config_xpath).get_attribute("class")
        if conf_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Configuration").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.txt_cu_xpath))
        )
        cu_class = driver.find_element(By.XPATH, LOC.txt_cu_xpath).get_attribute("class")
        if cu_class != "active":
            driver.find_element(By.LINK_TEXT, "Clouduser").click()
            time.sleep(5)
        totalCU = len(driver.find_elements(By.XPATH, LOC.txt_totalUser_xpath))
        for cuName in cuNames:
            for i in range(1, totalCU+1):
                if totalCU == 1:
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span[1]').text
                else:
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span[1]').text
                if tmp == cuName:
                    self.logger.info("********** Found The Clouduser With Name: " + cuName + " At Location, " + str(i) + "**********")
                else:
                    self.logger.info("********** Failed To Found The Clouduser With Name: " + cuName + " **********")

    def addVCenter(self, driver, path, start, end):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        conf_class = driver.find_element(By.XPATH, LOC.txt_config_xpath).get_attribute("class")
        if conf_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Configuration").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.txt_cu_xpath))
        )
        vc_class = driver.find_element(By.XPATH, LOC.txt_vc_xpath).get_attribute("class")
        if vc_class != "active":
            driver.find_element(By.LINK_TEXT, "vCenter").click()
            time.sleep(5)

        for r in range(st, ed + 1):
            name = sheet.cell(row=r, column=1).value
            ipAddress = sheet.cell(row=r, column=2).value
            userName = sheet.cell(row=r, column=3).value
            password = sheet.cell(row=r, column=4).value
            portNumber = sheet.cell(row=r, column=5).value

            val = self.findVCenter(driver, name)
            if val == 1:
                self.logger.info("********** The vCenter With Name: " + name + " Was Already Present **********")
                continue

            element = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_createVC_id))
            )
            element.click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_addVC_xpath)) != 0:
                self.logger.info("********** Add vCenter Pop-up Banner Is Opened For "+str(name)+" vCenter **********")
                driver.find_element(By.ID, LOC.txt_VCName_id).clear()
                driver.find_element(By.ID, LOC.txt_VCName_id).send_keys(name)
                driver.find_element(By.ID, LOC.txt_VCipAddress_id).clear()
                driver.find_element(By.ID, LOC.txt_VCipAddress_id).send_keys(ipAddress)
                driver.find_element(By.ID, LOC.txt_VCUserName_id).clear()
                driver.find_element(By.ID, LOC.txt_VCUserName_id).send_keys(userName)
                driver.find_element(By.ID, LOC.txt_VCPassword_id).clear()
                driver.find_element(By.ID, LOC.txt_VCPassword_id).send_keys(password)
                if portNumber != "NA":
                    driver.find_element(By.ID, LOC.txt_VCPort_id).send_keys(portNumber)
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_addVC_id))
                )
                driver.find_element(By.ID, LOC.btn_addVC_id).click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_deleteSuccessful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).text
                self.logger.info("********** Add New Clouduser Status For Clouduser: " + name + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Add vCenter Pop-up Banner Is Not Opened For "+str(name)+" vCenter **********")

    def findVCenter(self, driver, vcNames):
        flag = 0
        conf_class = driver.find_element(By.XPATH, LOC.txt_config_xpath).get_attribute("class")
        if conf_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "Configuration").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, LOC.txt_cu_xpath))
        )
        vc_class = driver.find_element(By.XPATH, LOC.txt_vc_xpath).get_attribute("class")
        if vc_class != "active":
            driver.find_element(By.LINK_TEXT, "vCenter").click()
            time.sleep(5)
        res = tuple(map(str, vcNames.split(', ')))
        totalUser = len(driver.find_elements(By.XPATH, LOC.txt_totalUser_xpath))
        for vcName in res:
            for i in range(1, totalUser + 1):
                if totalUser == 1:
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span[1]').text
                else:
                    tmp = driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span[1]').text
                if tmp == vcName:
                    self.logger.info("********** Found The vCenter With Name: " + vcName + " At Location, " + str(i) + "**********")
                    flag += 1
                else:
                    self.logger.info("********** Failed To Found The vCenter With Name: " + vcName + " **********")
        return flag
