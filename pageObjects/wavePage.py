import time
import openpyxl
import keyboard

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from utilities.customLogger import LogGen


class WavePage:
    drp_count_xpath = "//*[@id='waves_rw_dynamic_pagination']/select"
    btn_start_id = "wave_policy_wave_policy_wave_detail_start_replications"
    img_createWave_xpath = "//*[@id='waves_add_machine']/div[1]/div/img"
    btn_createWave_xpath = "//*[@id='waves_add_wave']/span/i"
    img_createNewWave_xpath = "//*[@id='waves_create_wave_add_machine']/div/img"
    img_uploadCSV_xpath = "//*[@id='dropzone_modal_page']/div/div[1]/img"

    lnk_createWave_id = "waves_add_wave"
    rd_waveWithHost_id = "create_wave_type_1"

    img_uploadImage_xpath = "/html/body/app-root/app-main-layout/div/app-waves/create-wave-modal/div/div/div/div[2]/div[2]/div[1]/div/div[1]/img"
    img_uplaod_xpath = "/html/body/app-root/app-main-layout/div/app-waves/create-wave-modal/div/div/div/div[2]/div[2]/div[1]/div/div[1]"

    txt_waveName_id = "wave_name"
    chBox_passthrough_id = "wave_dashboard_add_machine_passthrough"
    btn_create_id = "wave_dashboard_add_machine_submit_btn"
    frame_createWave_xpath = "//*[@id='main']/app-waves/add-machine/div/div/div/form"
    btn_selectFile_id = "waves_upload_template"

    txt_sourceDNSIP_id = "source_dns_ip"
    txt_sourceFriendlyName_id = "source_friendlyname"
    rd_OS0_id = "wave_dashboard_add_machine_os_radio_0"
    rd_OS1_id = "wave_dashboard_add_machine_os_radio_1"
    txt_sourceSshPort_css = "div.form-group:nth-child(6) > div:nth-child(2) > input:nth-child(1)"
    txt_sourceUserName_id = "source_username"

    rd_targetType0_id = "wave_dashboard_add_machine_target_type_radio_0"
    rd_targetType1_id = "wave_dashboard_add_machine_target_type_radio_1"
    rd_targetType2_id = "wave_dashboard_add_machine_target_type_radio_2"

    rd_syncType0_id = "wave_dashboard_add_machine_sync_type_radio_0"
    rd_syncType1_id = "wave_dashboard_add_machine_sync_type_radio_1"
    rd_syncType2_id = "wave_dashboard_add_machine_sync_type_radio_2"
    rd_syncType3_id = "wave_dashboard_add_machine_sync_type_radio_3"
    txt_targetFriendlyName_id = "target_friendlyname"
    txt_imageName_id = "clone_name"
    txt_targetDNSIP_id = "target_dns_ip"
    chBox_sourceUserPassword_id = "applySrcValue"
    txt_targetUserName_id = "target_username"
    txt_targetSshPort_css = ".left-border > div:nth-child(5) > div:nth-child(2) > div:nth-child(2) > input:nth-child(1)"

    btn_deleteWaves_xpath = "//*[@id='waves_bulk_delete']/span/i"
    btn_delete_id = "waves_wave_delete_delete_btn"
    btn_bulkDelete_id = "waves_wave_bulk_delete_delete_btn"

    txt_search_xpath = "//*[@id='waves_search_input']"

    btn_addHost_id = "wave_policy_wave_policy_wave_detail_add_machine"
    btn_createHost_id = "wave_dashboard_add_machine_submit_btn"

    pop_delete_xpath = '//*[@id="waves_wave_delete_title_text"]'
    pop_delMultWaves_id = "waves_wave_delete_delMutiple_waves"
    pop_waveOpen_xpath = '//*[@id="rmm_lite_header"]/div/div[1]/div[2]'
    pop_addHost_xpath = '//*[@id="wave_policy_wave_policy_wave_detail_add_machine"]/div/div/div/form/div[1]/h4'
    pop_successful_xpath = '/html/body/app-root/simple-notifications/div/simple-notification/div'
    pop_deleteSuccessful_xpath = '/html/body/app-root/simple-notifications/div/simple-notification[2]/div'

    logger = LogGen.loggen()

    def __init__(self, driver):
        self.driver = driver

    def findWave(self, waveName):
        flag = 0
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
            if (len(self.driver.find_elements(By.LINK_TEXT, "Summary"))) == 0:
                self.driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
            if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                self.driver.find_element(By.LINK_TEXT, "DR").click()
                flag += 1
                time.sleep(5)
                self.driver.find_element(By.LINK_TEXT, "Waves").click()
                time.sleep(5)
                if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                    flag += 1
                    self.logger.info("********** Wave : " + waveName + " Is Not Present **********")
        return flag

    def createWaveWithoutHost(self, waveNames, passthrough):
        tmp = "None"
        res = tuple(map(str, waveNames.split(', ')))
        for waveName in res:
            time.sleep(5)
            if len(self.driver.find_elements(By.LINK_TEXT, "Summary")) == 0:
                self.driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)

            if tmp != waveName:
                if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                    if len(self.driver.find_elements(By.XPATH, self.img_createWave_xpath)) != 0:
                        self.driver.find_element(By.XPATH, self.img_createWave_xpath).click()
                    else:
                        self.driver.find_element(By.XPATH, self.btn_createWave_xpath).click()
                        time.sleep(3)
                        self.driver.find_element(By.XPATH, self.img_createNewWave_xpath).click()
                    time.sleep(5)
                    if len(self.driver.find_elements(By.XPATH, '//*[@id="main"]/app-waves/add-machine/div/div/div/form/div[1]/h4')) != 0:
                        self.logger.info("********** Create New Wave Pop-up Banner Was Opened For Wave, "+str(waveName)+" **********")
                        self.driver.find_element(By.ID, self.txt_waveName_id).send_keys(waveName)
                        if not passthrough:
                            self.driver.find_element(By.ID, self.chBox_passthrough_id).click()
                        time.sleep(3)
                        self.driver.find_element(By.ID, self.btn_create_id).click()
                        time.sleep(2)
                        note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                        self.logger.info("********** Create New Wave Status of Wave : " + waveName + ",")
                        self.logger.info(note + "\n")
                        time.sleep(5)
                    else:
                        self.logger.info("********** Create New Wave Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
            tmp = waveName
        if len(self.driver.find_elements(By.LINK_TEXT, "Summary")) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def createWaveWithFile(self, path):
        time.sleep(5)
        if len(self.driver.find_elements(By.LINK_TEXT, "Summary")) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()
        btn = WebDriverWait(self.driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, self.btn_createWave_xpath))
        )
        btn.click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH,'//*[@id="main"]/app-waves/create-wave-modal/div/div/div/div[1]/h4')) != 0:
            self.logger.info("********** Create New Wave Pop-up Banner Was Opened **********")
            self.driver.find_element(By.XPATH, self.img_uploadCSV_xpath).click()
            time.sleep(5)
            keyboard.write(path)
            time.sleep(5)
            keyboard.send('enter')
            time.sleep(5)
            note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
            self.logger.info("********** Create Status of Wave Name, ")
            self.logger.info(note + "\n")
        else:
            self.logger.info("********** Create New Wave Pop-up Banner Was Not Opened **********")
        if len(self.driver.find_elements(By.LINK_TEXT, "Summary")) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def createWaveWithHost(self, path):
        time.sleep(5)
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row

        for r in range(3, rows+1):
            waveName = sheet.cell(row=r, column=1).value
            passthrough = sheet.cell(row=r, column=2).value

            time.sleep(5)
            if len(self.driver.find_elements(By.LINK_TEXT, "Summary")) == 0:
                self.driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()

            if len(self.driver.find_elements(By.LINK_TEXT, waveName)) == 0:
                if len(self.driver.find_elements(By.XPATH, self.img_createWave_xpath)) != 0:
                    self.driver.find_element(By.XPATH, self.img_createWave_xpath).click()
                else:
                    time.sleep(5)
                    self.driver.find_element(By.XPATH, self.btn_createWave_xpath).click()
                    time.sleep(3)
                    self.driver.find_element(By.XPATH, self.img_createNewWave_xpath).click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH,'//*[@id="main"]/app-waves/add-machine/div/div/div/form/div[1]/h4')) != 0:
                    self.logger.info("********** Create Wave Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
                    self.logger.info("********** Creating no. " + str(r - 2) + " wave with Target Type " + sheet.cell(row=r,column=3).value + " **********")
                    self.driver.find_element(By.ID, self.rd_waveWithHost_id).click()

                    self.driver.find_element(By.ID, self.txt_waveName_id).send_keys(waveName)
                    if not passthrough:
                        self.driver.find_element(By.ID, self.chBox_passthrough_id).click()
                    time.sleep(3)
                    self.enterData(sheet, r)
                    self.driver.find_element(By.ID, self.btn_create_id).click()
                    time.sleep(5)
                    note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                    self.logger.info("********** Create New Wave Status of Wave : " + waveName + ",")
                    self.logger.info(note + "\n")
                else:
                    self.logger.info("********** Create Wave Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")

        if len(self.driver.find_elements(By.LINK_TEXT, "Summary")) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def addHostToWaves(self, path):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        print(rows)
        for r in range(2, rows + 1):
            print(r)
            waveName = sheet.cell(row=r, column=1).value
            hostName = sheet.cell(row=r, column=5).value
            time.sleep(5)
            val = self.findWave(waveName)
            if val == 2:
                return
            time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_waveOpen_xpath)) != 0:
                self.logger.info("********** " + str(waveName) + " Was Opened **********")
                addHost = WebDriverWait(self.driver, 30).until(
                    EC.element_to_be_clickable((By.ID, self.btn_addHost_id))
                )
                addHost.click()
                time.sleep(3)
                if len(self.driver.find_elements(By.XPATH, self.pop_addHost_xpath)) != 0:
                    self.logger.info("********** Add New Host Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
                    self.enterData(sheet, r)
                    self.driver.find_element(By.ID, self.btn_createHost_id).click()
                    time.sleep(5)
                    note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                    self.logger.info("********** Add New Host Status of Wave : " + waveName + ",")
                    self.logger.info(note + "\n")
                    time.sleep(5)
                    self.searchHost(hostName, waveName)
                else:
                    self.logger.info("********** Add New Host Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
            else:
                self.logger.info("********** " + str(waveName) + " Was Not Opened **********")
            time.sleep(5)
            if val == 1:
                self.driver.find_element(By.LINK_TEXT, "Replication").click()
                time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def addHostToWave(self, path):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        tmp = "None"
        for r in range(2, rows+1):
            time.sleep(5)
            waveName = sheet.cell(row=r, column=1).value
            hostName = sheet.cell(row=r, column=5).value
            if tmp != waveName:
                val = self.findWave(waveName)
                if val == 2:
                    return
            self.driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_waveOpen_xpath)) != 0:
                self.logger.info("********** " + str(waveName) + " Was Opened **********")
                addHost = WebDriverWait(self.driver, 30).until(
                    EC.element_to_be_clickable((By.ID, self.btn_addHost_id))
                )
                time.sleep(5)
                addHost.click()
                time.sleep(3)
                if len(self.driver.find_elements(By.XPATH, self.pop_addHost_xpath)) != 0:
                    self.logger.info("********** Add New Host Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
                    self.enterData(sheet, r)
                    time.sleep(5)
                    self.driver.find_element(By.ID, self.btn_createHost_id).click()
                    time.sleep(3)
                    note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                    self.logger.info("********** Add New Host Status of Wave : " + waveName + ",")
                    self.logger.info(note + "\n")
                    time.sleep(5)
                    self.searchHost(hostName, waveName)
                    if type(sheet.cell(row=r+1, column=3).value) == "NoneType":
                        return
                else:
                    self.logger.info("********** Add New Host Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
            else:
                self.logger.info("********** " + str(waveName) + " Was Not Opened **********")
            tmp = waveName
        if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def enterData(self, sheet, r):
        time.sleep(5)
        # Source Data
        sourceDNS_IP = sheet.cell(row=r, column=4).value
        sourceFriendlyName = sheet.cell(row=r, column=5).value
        sourceOS = sheet.cell(row=r, column=6).value
        sourceSSHport = sheet.cell(row=r, column=7).value
        sourceUserName = sheet.cell(row=r, column=8).value

        self.driver.find_element(By.ID, self.txt_sourceDNSIP_id).send_keys(sourceDNS_IP)
        self.driver.find_element(By.ID, self.txt_sourceFriendlyName_id).clear()
        self.driver.find_element(By.ID, self.txt_sourceFriendlyName_id).send_keys(sourceFriendlyName)
        if sourceOS == "Linux":
            self.driver.find_element(By.ID, self.rd_OS0_id).click()
        elif sourceOS == "Windows":
            self.driver.find_element(By.ID, self.rd_OS1_id).click()
        self.driver.find_element(By.CSS_SELECTOR, self.txt_sourceSshPort_css).send_keys(sourceSSHport)
        self.driver.find_element(By.ID, self.txt_sourceUserName_id).clear()
        self.driver.find_element(By.ID, self.txt_sourceUserName_id).send_keys(sourceUserName)
        time.sleep(3)

        # Target Data
        targetType = sheet.cell(row=r, column=3).value
        syncType = sheet.cell(row=r, column=9).value
        targetFriendlyName = sheet.cell(row=r, column=10).value
        targetImageName = sheet.cell(row=r, column=11).value
        targetDNS_IP = sheet.cell(row=r, column=12).value
        sourceUserPassword = sheet.cell(row=r, column=13).value
        targetUsername = sheet.cell(row=r, column=14).value
        targetSSHport = sheet.cell(row=r, column=15).value

        # Autoprovision
        if targetType == "Autoprovision":
            self.driver.find_element(By.ID, self.rd_targetType0_id).click()
            if syncType == "Direct Sync":
                self.driver.find_element(By.ID, self.rd_syncType0_id).click()
            else:
                if syncType == "Stage 1 & 2":
                    self.driver.find_element(By.ID, self.rd_syncType1_id).click()
                elif syncType == "Stage 1":
                    self.driver.find_element(By.ID, self.rd_syncType2_id).click()
                elif syncType == "Stage 2":
                    self.driver.find_element(By.ID, self.rd_syncType3_id).click()
                self.driver.find_element(By.ID, self.txt_imageName_id).clear()
                self.driver.find_element(By.ID, self.txt_imageName_id).send_keys(targetImageName)
            self.driver.find_element(By.ID, self.txt_targetFriendlyName_id).clear()
            self.driver.find_element(By.ID, self.txt_targetFriendlyName_id).send_keys(targetFriendlyName)

        # Exiting System
        elif targetType == "Existing System":
            self.driver.find_element(By.ID, self.rd_targetType1_id).click()
            if syncType == "Direct Sync":
                self.driver.find_element(By.ID, self.rd_syncType0_id).click()
                self.selectSyncType(targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
                self.driver.find_element(By.CSS_SELECTOR, self.txt_targetSshPort_css).send_keys(targetSSHport)
            elif syncType == "Stage 1 & 2":
                self.driver.find_element(By.ID, self.rd_syncType1_id).click()
                self.driver.find_element(By.ID, self.txt_imageName_id).clear()
                self.driver.find_element(By.ID, self.txt_imageName_id).send_keys(targetImageName)
                self.selectSyncType(targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
                self.driver.find_element(By.CSS_SELECTOR, self.txt_targetSshPort_css).send_keys(targetSSHport)
            elif syncType == "Stage 1":
                self.driver.find_element(By.ID, self.rd_syncType2_id).click()
                self.driver.find_element(By.ID, self.txt_imageName_id).clear()
                self.driver.find_element(By.ID, self.txt_imageName_id).send_keys(targetImageName)
                self.selectSyncType(targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
            elif syncType == "Stage 2":
                self.driver.find_element(By.ID, self.rd_syncType3_id).click()
                self.driver.find_element(By.ID, self.txt_imageName_id).clear()
                self.driver.find_element(By.ID, self.txt_imageName_id).send_keys(targetImageName)
                self.selectSyncType(targetDNS_IP, targetFriendlyName, sourceUserPassword, targetUsername)
                self.driver.find_element(By.CSS_SELECTOR, self.txt_targetSshPort_css).send_keys(targetSSHport)

        # Capture
        elif targetType == "Capture":
            self.driver.find_element(By.ID, self.rd_targetType2_id).click()
            self.driver.find_element(By.ID, self.txt_imageName_id).clear()
            self.driver.find_element(By.ID, self.txt_imageName_id).send_keys(targetImageName)

    def selectSyncType(self, DNS_IP, friendlyName, sourceUserPassword, targetUsername):
        time.sleep(5)
        self.driver.find_element(By.ID, self.txt_targetDNSIP_id).send_keys(DNS_IP)
        self.driver.find_element(By.ID, self.txt_targetFriendlyName_id).clear()
        self.driver.find_element(By.ID, self.txt_targetFriendlyName_id).send_keys(friendlyName)
        if sourceUserPassword:
            self.driver.find_element(By.ID, self.chBox_sourceUserPassword_id).click()
        else:
            self.driver.find_element(By.ID, self.txt_targetUserName_id).send_keys(targetUsername)

    def deleteWave(self, waveName):
        time.sleep(5)
        val = self.findWave(waveName)
        if val == 2:
            return
        time.sleep(5)
        self.driver.find_element(By.ID, "waves_"+waveName+"_wave_actions").click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.pop_delete_xpath)) != 0:
            self.logger.info("********** Delete Wave Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
            self.driver.find_element(By.ID, self.btn_delete_id).click()
            time.sleep(5)
            note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
            self.logger.info("********** Delete Wave Status of Wave : " + waveName + ",")
            self.logger.info(note + "\n")
        else:
            self.logger.info("********** Delete Wave Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
        time.sleep(5)
        if val == 1:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def deleteWaves(self, waveNames):
        time.sleep(5)
        res = tuple(map(str, waveNames.split(', ')))
        if (len(self.driver.find_elements(By.LINK_TEXT, "Summary"))) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()
        time.sleep(5)
        totalWaves = len(self.driver.find_elements(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr'))
        for waveName in res:
            count = 0
            for i in range(1, totalWaves+1):
                if totalWaves == 1:
                    tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span/div/a').text
                else:
                    tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span/div/a').text
                if tmp == waveName:
                    break
                count += 1
            if count > totalWaves:
                self.logger.info("********** "+waveName+" Was Not Found In Replication Waves **********")
            else:
                self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(count)+']/td[1]/p-tablecheckbox/div/div[2]').click()
            time.sleep(2)
        self.driver.find_element(By.XPATH, self.btn_deleteWaves_xpath).click()
        time.sleep(5)
        if len(self.driver.find_elements(By.ID, self.pop_delMultWaves_id)) != 0:
            self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Opened **********")
            self.driver.find_element(By.ID, self.btn_bulkDelete_id).click()
            time.sleep(5)
            note = self.driver.find_element(By.XPATH, self.pop_deleteSuccessful_xpath).text
            self.logger.info("********** Delete Wave Status,")
            self.logger.info(note + "\n")
        else:
            self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Not Opened **********")
        time.sleep(5)

    def searchWave(self, waveNames):
        time.sleep(5)
        res = tuple(map(str, waveNames.split(', ')))
        if (len(self.driver.find_elements(By.LINK_TEXT, "Summary"))) == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()
        time.sleep(5)
        totalWaves = len(self.driver.find_elements(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr'))
        for waveName in res:
            count = 0
            for i in range(1, totalWaves+1):
                if totalWaves == 1:
                    tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span/div/a').text
                else:
                    tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span/div/a').text
                if tmp == waveName:
                    self.logger.info("********** " + waveName + " Was Found In Replication Waves At Position "+str(count)+" **********")
                    break
                count += 1
            if count > totalWaves:
                self.logger.info("********** "+waveName+" Was Not Found In Replication Waves **********")

    def searchHost(self, waveName, hostName):
        time.sleep(5)
        val = self.findWave(waveName)
        if val == 2:
            return
        self.driver.find_element(By.LINK_TEXT, waveName).click()
        WebDriverWait(self.driver, 6000).until(
            EC.element_to_be_clickable((By.ID, self.btn_start_id))
        )
        time.sleep(5)
        totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
        for hostNo in range(1, totalHosts + 1):
            if totalHosts == 1:
                tmp = self.driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
            else:
                tmp = self.driver.find_element(By.XPATH,'//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(hostNo) + ']/td[3]/span/span').text
            if hostName == tmp:
                self.logger.info("********** The Host " + hostName + " Was Found In The Wave : " + waveName + " **********")
                break
            elif hostNo == totalHosts:
                self.logger.info("********** The Host " + hostName + " Was Not Found In The Wave : " + waveName + " **********")
        if val == 1:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Waves").click()
