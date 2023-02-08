import time

import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from utilities.customLogger import LogGen


class DRPolicy:
    btn_addNew_xpath = "//*[@id='policies_dr_policy_create_policy_btn']/span/i"

    txt_name_id = "policyName"
    rd_schedule_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[1]/div/div/input"
    rd_frequency_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[2]/div/div/input"
    rd_once_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[3]/div/div/input"
    rd_continuous_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[4]/div/div/input"
    txt_start_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[3]/div/p-calendar/span/input"
    txt_email_xpath = "//*[@id='email']"
    btn_add_xpath = "//*[@id='policies_create_dr_policy_sbf_addEmail_btn']/i"
    chBox_failNote_xpath = "//*[@id='notifyfailonly']"
    chBox_completeNote_xpath = "//*[@id='notifyonwavecomplete']"
    btn_create_id = "policies_create_dr_policy_sbf_create_btn"
    btn_clear_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[3]/div/p-calendar/span/div/div[3]/div/div[2]/button/span"
    btn_resume_xpath = "//*[@id='policies_dr_policy_resume_dr_policy']/div/div/div/form/div/div[3]/div/button[2]"

    # By Schedule
    drp_dw_xpath = "//*[@id='select_daily_weekly']"
    drp_hr_xpath = "//*[@id='policies_create_dr_policy_sbs_select_hour']"
    drp_min_xpath = "//*[@id='policies_create_dr_policy_sbs_select_minute']"
    drp_wkDay_xpath = "//*[@id='select_week']"

    # By Frequency
    drp_durationMin_id = "policies_create_dr_policy_sbf_select_minute"
    drp_durationHr_id = "policies_create_dr_policy_sbf_select_hour"
    drp_minHr_xpath = "//*[@id='select_hours_minutes']"
    drp_fromHr_id = "policies_create_dr_policy_sbf_exclude_from_hr"
    drp_fromMin_id = "policies_create_dr_policy_sbf_exclude_from_min"
    drp_toHr_id = "policies_create_dr_policy_sbf_exclude_to_hr"
    drp_toMin_id = "policies_create_dr_policy_sbf_exclude_to_min"

    # Once
    btn_startDate_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[3]/div/p-calendar/span/button/span[1]"

    logger = LogGen.loggen()

    def __init__(self, driver):
        self.driver = driver

    def createDRPolicy(self, path):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        for r in range(2, rows+1):
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_addNew_xpath).click()
            time.sleep(5)
            name = sheet.cell(row=r, column=1).value
            startTime = sheet.cell(row=r, column=2).value
            email = sheet.cell(row=r, column=3).value
            note1 = sheet.cell(row=r, column=4).value
            note2 = sheet.cell(row=r, column=5).value
            periodicity = sheet.cell(row=r, column=6).value

            self.driver.find_element(By.ID, self.txt_name_id).send_keys(name)

            if periodicity == "By Schedule":
                text = sheet.cell(row=r, column=7).value
                hrs = sheet.cell(row=r, column=8).value
                mins = sheet.cell(row=r, column=9).value
                day = sheet.cell(row=r, column=10).value

                self.driver.find_element(By.XPATH, self.rd_schedule_xpath).click()
                dailyWeekly = Select(self.driver.find_element(By.XPATH, self.drp_dw_xpath))
                hour = Select(self.driver.find_element(By.XPATH, self.drp_hr_xpath))
                minute = Select(self.driver.find_element(By.XPATH, self.drp_min_xpath))

                dailyWeekly.select_by_visible_text(text)
                if text == "Weekly":
                    weekDay = Select(self.driver.find_element(By.XPATH, self.drp_wkDay_xpath))
                    weekDay.select_by_visible_text(day)
                hour.select_by_index(hrs)
                minute.select_by_index(mins)

            elif periodicity == "By Frequency":
                mh = sheet.cell(row=r, column=11).value
                due = sheet.cell(row=r, column=12).value
                fhr = sheet.cell(row=r, column=13).value
                fmn = sheet.cell(row=r, column=14).value
                thr = sheet.cell(row=r, column=15).value
                tmn = sheet.cell(row=r, column=16).value

                self.driver.find_element(By.XPATH, self.rd_frequency_xpath).click()
                minHr = Select(self.driver.find_element(By.XPATH, self.drp_minHr_xpath))
                fromHr = Select(self.driver.find_element(By.ID, self.drp_fromHr_id))
                fromMin = Select(self.driver.find_element(By.ID, self.drp_fromMin_id))
                toHr = Select(self.driver.find_element(By.ID, self.drp_toHr_id))
                toMin = Select(self.driver.find_element(By.ID, self.drp_toMin_id))

                minHr.select_by_visible_text(mh)
                time.sleep(5)
                if mh == "Minutes":
                    duration = Select(self.driver.find_element(By.ID, self.drp_durationMin_id))
                else:
                    duration = Select(self.driver.find_element(By.ID, self.drp_durationHr_id))
                if mh == "Hours":
                    duration.select_by_index(due-1)
                    time.sleep(5)
                elif mh == "Minutes" and due == 5:
                    duration.select_by_index(0)
                elif mh == "Minutes" and due > 5:
                    duration.select_by_visible_text(str(due))

                fromHr.select_by_index(fhr+1)
                fromMin.select_by_index(fmn+1)
                toHr.select_by_index(thr+1)
                toMin.select_by_index(tmn+1)

            elif periodicity == "Once":
                self.driver.find_element(By.XPATH, self.rd_once_xpath).click()
                self.driver.find_element(By.XPATH, self.txt_start_xpath).click()
                self.driver.find_element(By.XPATH, self.btn_clear_xpath).click()
                self.driver.find_element(By.XPATH, self.txt_start_xpath).send_keys(startTime)
                time.sleep(5)
                self.driver.find_element(By.XPATH, self.btn_startDate_xpath).click()

            elif periodicity == "Continuous":
                self.driver.find_element(By.XPATH, self.rd_continuous_xpath).click()

            self.driver.find_element(By.XPATH, self.txt_email_xpath).send_keys(email)
            self.driver.find_element(By.XPATH, self.btn_add_xpath).click()
            if note1:
                self.driver.find_element(By.XPATH, self.chBox_failNote_xpath).click()
            if note2:
                self.driver.find_element(By.XPATH, self.chBox_completeNote_xpath).click()
            time.sleep(5)
            self.driver.find_element(By.ID, self.btn_create_id).click()
            self.logger.info("********** Created no. " + str(r - 1) + " DR Policy with name " + sheet.cell(row=r, column=1).value + " **********")

    def startPolicy(self, vals):
        for i in vals:
            time.sleep(5)
            self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[9]/span/i[1]')
            self.driver.find_element(By.XPATH, self.btn_resume_xpath).click()
        time.sleep(5)
