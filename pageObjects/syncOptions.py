import time
import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from utilities.customLogger import LogGen


class SyncOptions:
    # Sync Options
    btn_syncOptions_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/edit-item/div/div/div/form/div[2]/div/p-tabview/div/ul/li[2]/a/span"
    btn_modify_id = "wave_detail_edit_item_modify_btn"
    rd_tng_id = "wave_detail_edit_item_options_tng"
    rd_verbose_id = "wave_detail_edit_item_options_verbose"
    rd_directFScopy_id = "wave_detail_edit_item_options_allowdirectfscopy"
    rd_FSDeletion_id = "wave_detail_edit_item_options_allowFsDeletion"
    rd_NoTransfer_id = "wave_detail_edit_item_options_no_xfer"
    rd_transferCompress_id = "wave_detail_edit_item_options_xfer_compress"
    rd_noTransferCompress_id = "wave_detail_edit_item_options_no_xfer_compress"
    rd_ignoreMissing_id = "wave_detail_edit_item_options_ignore_missing"
    rd_noInPlace_id = "wave_detail_edit_item_options_noInPlace"
    rd_noReboot_id = "wave_detail_edit_item_options_noReboot"
    rd_includeSAN_id = "wave_detail_edit_item_options_include_san"
    rd_excludeSAN_id = "wave_detail_edit_item_options_exclude_san"
    rd_overrideRMMStorageCheck_id = "wave_detail_edit_item_options_storage_override"
    rd_deleteAllTargetFS_id = "wave_detail_edit_item_options_delete_all_target_fs"
    rd_keepTargetLayout_id = "wave_detail_edit_item_options_keep_target_layout"
    rd_cloudInit_id = "wave_detail_edit_item_options_cloud_init"

    # Bulk Edit Sync
    btn_selectAll_xpath = "/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/thead/tr/th[1]/p-tableheadercheckbox/div/div[2]"
    btn_bulkEdit_xpath = "//*[@id='content']/article/div/div[2]/p-table/div/div[1]/div[1]/button[8]/span/i"
    drp_goal_id = "wave_detail_bulk_edit_item_existing_capture_type"
    ch_tngYes_id = "tng_yes"
    ch_tngNo_id = "tng_no"
    ch_verboseYes_id = "verbose_yes"
    ch_verboseNo_id = "verbose_no"
    ch_passwordLessYes_id = "ssh_yes"
    ch_passwordLessNo_id = "ssh_no"
    ch_AllowDirectFScopyYes_id = "allowDirectFscopy_yes"
    ch_AllowDirectFScopyNo_id = "allowDirectFscopy_no"
    ch_AllowFSDeletionYes_id = "allowFsDeletion_yes"
    ch_AllowFSDeletionNo_id = "allowFsDeletion_no"
    ch_NoTransferYes_id = "no_xfer_yes"
    ch_NoTransferNo_id = "no_xfer_no"
    ch_TransferCompressYes_id = "xferCompress_yes"
    ch_TransferCompressNo_id = "xferCompress_no"
    ch_NoTransferCompressYes_id = "no_xfer_compress_yes"
    ch_NoTransferCompressNo_id = "no_xfer_compress_no"
    ch_IgnoreMissingYes_id = "ignoreMissing_yes"
    ch_IgnoreMissingNo_id = "ignoreMissing_no"
    ch_NoInPlaceYes_id = "noInPlace_yes"
    ch_NoInPlaceNo_id = "noInPlace_no"
    ch_NoRebootYes_id = "noReboot_yes"
    ch_NoRebootNo_id = "noReboot_no"
    ch_IncludeSANYes_id = "include_san_yes"
    ch_IncludeSANNo_id = "include_san_no"
    ch_ExcludeSANYes_id = "exclude_san_yes"
    ch_ExcludeSANNo_id = "exclude_san_no"
    ch_OverrideRMMStorageCheckYes_id = "storage_override_yes"
    ch_OverrideRMMStorageCheckNo_id = "storage_override_no"
    ch_DeleteAllTargetFSYes_id = "delete_all_target_fs_yes"
    ch_DeleteAllTargetFSNo_id = "delete_all_target_fs_no"
    ch_KeepTargetLayoutYes_id = "keep_target_layout_yes"
    ch_KeepTargetLayoutNo_id = "keep_target_layout_no"
    ch_CloudInitYes_id = "cloud_init_yes"
    ch_CloudInitNo_id = "cloud_init_no"
    btn_next_xpath = "//*[@id='main']/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[5]/div/button[2]"
    btn_modifyAll_xpath = "//*[@id='main']/rw-wave-detail/bulk-edit/div[1]/div/div/form/div[5]/div/button[2]"
    btn_yes_id = "wave_policy_wave_policy_wave_detail_autoprov_not_conf_yes_btn"

    logger = LogGen.loggen()

    def __init__(self, driver):
        self.driver = driver

    def setSyncOptions(self, r, sheet):
        edit = WebDriverWait(self.driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(r-2)+']/td[8]/span/div/i[2]'))
        )
        edit.click()
        # self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(r-2)+']/td[8]/span/div/i[2]').clikc()

        syncOption = WebDriverWait(self.driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, self.btn_syncOptions_xpath))
        )
        syncOption.click()
        time.sleep(5)

        tng = sheet.cell(row=r, column=16).value
        verbose = sheet.cell(row=r, column=17).value
        directFScopy = sheet.cell(row=r, column=18).value
        FSDeletion = sheet.cell(row=r, column=19).value
        NoTransfer = sheet.cell(row=r, column=20).value
        transferCompress = sheet.cell(row=r, column=21).value
        noTransferCompress = sheet.cell(row=r, column=22).value
        ignoreMissing = sheet.cell(row=r, column=23).value
        noInPlace = sheet.cell(row=r, column=24).value
        noReboot = sheet.cell(row=r, column=25).value
        includeSAN = sheet.cell(row=r, column=26).value
        excludeSAN = sheet.cell(row=r, column=27).value
        overrideRMMStorageCheck = sheet.cell(row=r, column=28).value
        deleteAllTargetFS = sheet.cell(row=r, column=29).value
        keepTargetLayout = sheet.cell(row=r, column=30).value
        cloudInit = sheet.cell(row=r, column=31).value

        if tng:
            self.driver.find_element(By.ID, self.rd_tng_id).click()
            self.logger.info("********** Sync Option 'TNG' is set **********")
        if verbose:
            self.driver.find_element(By.ID, self.rd_verbose_id).click()
            self.logger.info("********** Sync Option 'Verbose' is set **********")
        if directFScopy:
            self.driver.find_element(By.ID, self.rd_directFScopy_id).click()
            self.logger.info("********** Sync Option 'Direct FScopy' is set **********")
        if FSDeletion:
            self.driver.find_element(By.ID, self.rd_FSDeletion_id).click()
            self.logger.info("********** Sync Option 'FS Deletion' is set **********")
        if NoTransfer:
            self.driver.find_element(By.ID, self.rd_NoTransfer_id).click()
            self.logger.info("********** Sync Option 'No Transfer' is set **********")
        if transferCompress:
            self.driver.find_element(By.ID, self.rd_transferCompress_id).click()
            self.logger.info("********** Sync Option 'Transfer Compress' is set **********")
        if noTransferCompress:
            self.driver.find_element(By.ID, self.rd_noTransferCompress_id).click()
            self.logger.info("********** Sync Option 'No Transfer Compress' is set **********")
        if ignoreMissing:
            self.driver.find_element(By.ID, self.rd_ignoreMissing_id).click()
            self.logger.info("********** Sync Option 'Ignore Missing' is set **********")
        if noInPlace:
            self.driver.find_element(By.ID, self.rd_noInPlace_id).click()
            self.logger.info("********** Sync Option 'No In Place' is set **********")
        if noReboot:
            self.driver.find_element(By.ID, self.rd_noReboot_id).click()
            self.logger.info("********** Sync Option 'No Reboot' is set **********")
        if includeSAN:
            self.driver.find_element(By.ID, self.rd_includeSAN_id).click()
            self.logger.info("********** Sync Option 'Include SAN' is set **********")
        if excludeSAN:
            self.driver.find_element(By.ID, self.rd_excludeSAN_id).click()
            self.logger.info("********** Sync Option 'Exclude SAN' is set **********")
        if overrideRMMStorageCheck:
            self.driver.find_element(By.ID, self.rd_overrideRMMStorageCheck_id).click()
            self.logger.info("********** Sync Option 'Override RMM Storage' Check is set **********")
        if deleteAllTargetFS:
            self.driver.find_element(By.ID, self.rd_deleteAllTargetFS_id).click()
            self.logger.info("********** Sync Option 'Delete All Target' FS is set **********")
        if keepTargetLayout:
            self.driver.find_element(By.ID, self.rd_keepTargetLayout_id).click()
            self.logger.info("********** Sync Option 'Keep Target Layout' is set **********")
        if cloudInit:
            self.driver.find_element(By.ID, self.rd_cloudInit_id).click()
            self.logger.info("********** Sync Option 'Cloud Init' is set **********")

        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_modify_id).click()

    def bulkEditOption(self, path):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row

        for r in range(2, rows+1):
            waveName = sheet.cell(row=r, column=1).value
            goal = sheet.cell(row=r, column=2).value
            tng = sheet.cell(row=r, column=3).value
            verbose = sheet.cell(row=r, column=4).value
            passwordLess = sheet.cell(row=r, column=5).value
            directFScopy = sheet.cell(row=r, column=6).value
            FSDeletion = sheet.cell(row=r, column=7).value
            NoTransfer = sheet.cell(row=r, column=8).value
            transferCompress = sheet.cell(row=r, column=9).value
            noTransferCompress = sheet.cell(row=r, column=10).value
            ignoreMissing = sheet.cell(row=r, column=11).value
            noInPlace = sheet.cell(row=r, column=12).value
            noReboot = sheet.cell(row=r, column=13).value
            includeSAN = sheet.cell(row=r, column=14).value
            excludeSAN = sheet.cell(row=r, column=15).value
            overrideRMMStorageCheck = sheet.cell(row=r, column=16).value
            deleteAllTargetFS = sheet.cell(row=r, column=17).value
            keepTargetLayout = sheet.cell(row=r, column=18).value
            cloudInit = sheet.cell(row=r, column=19).value

            self.driver.find_element(By.LINK_TEXT, waveName).click()
            # selectAll = WebDriverWait(self.driver, 30).until(
            #     EC.element_to_be_clickable((By.XPATH, self.btn_selectAll_xpath))
            # )
            # selectAll.click()
            # bulkEdit = WebDriverWait(self.driver, 30).until(
            #     EC.element_to_be_clickable((By.XPATH, self.btn_bulkEdit_xpath))
            # )
            # bulkEdit.click()
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_selectAll_xpath).click()

            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_bulkEdit_xpath).click()

            time.sleep(5)
            if tng == "Yes":
                self.driver.find_element(By.ID, self.ch_tngYes_id).click()
                self.logger.info("********** Sync Option 'TNG' is set as yes **********")
            if tng == "No":
                self.driver.find_element(By.ID, self.ch_tngNo_id).click()
                self.logger.info("********** Sync Option 'TNG' is set as no **********")

            if verbose == "Yes":
                self.driver.find_element(By.ID, self.ch_verboseYes_id).click()
                self.logger.info("********** Sync Option 'Verbose' is set as yes **********")
            if verbose == "No":
                self.driver.find_element(By.ID, self.ch_verboseNo_id).click()
                self.logger.info("********** Sync Option 'Verbose' is set as no **********")

            if passwordLess == "Yes":
                self.driver.find_element(By.ID, self.ch_passwordLessYes_id).click()
                self.logger.info("********** Sync Option 'Passwordless' is set as yes **********")
            if passwordLess == "No":
                self.driver.find_element(By.ID, self.ch_passwordLessNo_id).click()
                self.logger.info("********** Sync Option 'Passwordless' is set as no **********")

            if directFScopy == "Yes":
                self.driver.find_element(By.ID, self.ch_AllowDirectFScopyYes_id).click()
                self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as yes **********")
            if directFScopy == "No":
                self.driver.find_element(By.ID, self.ch_AllowDirectFScopyNo_id).click()
                self.logger.info("********** Sync Option 'Allow Direct Fscopy' is set as no **********")

            if FSDeletion == "Yes":
                self.driver.find_element(By.ID, self.ch_AllowFSDeletionYes_id).click()
                self.logger.info("********** Sync Option 'Allow FS Deletion' is set as yes **********")
            if FSDeletion == "No":
                self.driver.find_element(By.ID, self.ch_AllowFSDeletionNo_id).click()
                self.logger.info("********** Sync Option 'Allow FS Deletion' is set as no **********")

            if NoTransfer == "Yes":
                self.driver.find_element(By.ID, self.ch_NoTransferYes_id).click()
                self.logger.info("********** Sync Option 'No Transfer' is set as yes **********")
            if NoTransfer == "No":
                self.driver.find_element(By.ID, self.ch_NoTransferNo_id).click()
                self.logger.info("********** Sync Option 'No Transfer' is set as no **********")

            if transferCompress == "Yes":
                self.driver.find_element(By.ID, self.ch_TransferCompressYes_id).click()
                self.logger.info("********** Sync Option 'Transfer Compress' is set as yes **********")
            if transferCompress == "No":
                self.driver.find_element(By.ID, self.ch_TransferCompressNo_id).click()
                self.logger.info("********** Sync Option 'Transfer Compress' is set as no **********")

            if noTransferCompress == "Yes":
                self.driver.find_element(By.ID, self.ch_NoTransferCompressYes_id).click()
                self.logger.info("********** Sync Option 'No Transfer Compress' is set as yes **********")
            if noTransferCompress == "No":
                self.driver.find_element(By.ID, self.ch_NoTransferCompressNo_id).click()
                self.logger.info("********** Sync Option 'No Transfer Compress' is set as no **********")

            if ignoreMissing == "Yes":
                self.driver.find_element(By.ID, self.ch_IgnoreMissingYes_id).click()
                self.logger.info("********** Sync Option 'Ignore Missing' is set as yes **********")
            if ignoreMissing == "No":
                self.driver.find_element(By.ID, self.ch_IgnoreMissingNo_id).click()
                self.logger.info("********** Sync Option 'Ignore Missing' is set as no **********")

            if noInPlace == "Yes":
                self.driver.find_element(By.ID, self.ch_NoInPlaceYes_id).click()
                self.logger.info("********** Sync Option 'No In Place' is set as yes **********")
            if noInPlace == "No":
                self.driver.find_element(By.ID, self.ch_NoInPlaceNo_id).click()
                self.logger.info("********** Sync Option 'No In Place' is set as no **********")

            if noReboot == "Yes":
                self.driver.find_element(By.ID, self.ch_NoRebootYes_id).click()
                self.logger.info("********** Sync Option 'No Reboot' is set as yes **********")
            if noReboot == "No":
                self.driver.find_element(By.ID, self.ch_NoRebootNo_id).click()
                self.logger.info("********** Sync Option 'No Reboot' is set as no **********")

            if includeSAN == "Yes":
                self.driver.find_element(By.ID, self.ch_IncludeSANYes_id).click()
                self.logger.info("********** Sync Option 'Include SAN' is set as yes **********")
            if includeSAN == "No":
                self.driver.find_element(By.ID, self.ch_IncludeSANNo_id).click()
                self.logger.info("********** Sync Option 'Include SAN' is set as no **********")

            if excludeSAN == "Yes":
                self.driver.find_element(By.ID, self.ch_ExcludeSANYes_id).click()
                self.logger.info("********** Sync Option 'Exclude SAN' is set as yes **********")
            if excludeSAN == "No":
                self.driver.find_element(By.ID, self.ch_ExcludeSANNo_id).click()
                self.logger.info("********** Sync Option 'Exclude SAN' is set as no **********")

            if overrideRMMStorageCheck == "Yes":
                self.driver.find_element(By.ID, self.ch_OverrideRMMStorageCheckYes_id).click()
                self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as yes **********")
            if overrideRMMStorageCheck == "No":
                self.driver.find_element(By.ID, self.ch_OverrideRMMStorageCheckNo_id).click()
                self.logger.info("********** Sync Option 'Override RMM Storage Check' is set as no **********")

            if deleteAllTargetFS == "Yes":
                self.driver.find_element(By.ID, self.ch_DeleteAllTargetFSYes_id).click()
                self.logger.info("********** Sync Option 'Delete All Target FS' is set as yes **********")
            if deleteAllTargetFS == "No":
                self.driver.find_element(By.ID, self.ch_DeleteAllTargetFSNo_id).click()
                self.logger.info("********** Sync Option 'Delete All Target FS' is set as no **********")

            if keepTargetLayout == "Yes":
                self.driver.find_element(By.ID, self.ch_KeepTargetLayoutYes_id).click()
                self.logger.info("********** Sync Option 'Keep Target Layout' is set as yes **********")
            if keepTargetLayout == "No":
                self.driver.find_element(By.ID, self.ch_KeepTargetLayoutNo_id).click()
                self.logger.info("********** Sync Option 'Keep Target Layout' is set as no **********")

            if cloudInit == "Yes":
                self.driver.find_element(By.ID, self.ch_CloudInitYes_id).click()
                self.logger.info("********** Sync Option 'Cloud Init' is set as yes **********")
            if cloudInit == "No":
                self.driver.find_element(By.ID, self.ch_CloudInitNo_id).click()
                self.logger.info("********** Sync Option 'Cloud Init' is set as no **********")

            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_next_xpath).click()
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_modifyAll_xpath).click()
            time.sleep(5)
            self.driver.find_element(By.ID, self.btn_yes_id).click()
            time.sleep(5)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
