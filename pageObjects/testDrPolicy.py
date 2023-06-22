import time
import unittest
import openpyxl

import Locators.locDrPolicy as LOC

from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from utilities.customLogger import LogGen
from utilities.commonObjects import CommonObjects


class DRPolicy(unittest.TestCase):

    logger = LogGen.loggen()

    def createDRPolicy(self, driver, path, start, end):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(3)
        driver.find_element(By.LINK_TEXT, "Policies").click()
        for r in range(st, ed+1):
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.btn_addNew_xpath))
            )
            btn.click()
            time.sleep(3)
            if len(driver.find_elements(By.XPATH, LOC.pop_createDrPolicy_xpath)) != 0:
                self.logger.info("********** Create New DR Policy Pop-up Banner Is Opened **********")
                name = sheet.cell(row=r, column=1).value
                startTime = sheet.cell(row=r, column=2).value
                email = sheet.cell(row=r, column=3).value
                note1 = sheet.cell(row=r, column=4).value
                note2 = sheet.cell(row=r, column=5).value
                periodicity = sheet.cell(row=r, column=6).value

                driver.find_element(By.ID, LOC.txt_name_id).send_keys(name)

                if periodicity == "By Schedule":
                    text = sheet.cell(row=r, column=7).value
                    hrs = sheet.cell(row=r, column=8).value
                    mins = sheet.cell(row=r, column=9).value
                    day = sheet.cell(row=r, column=10).value

                    driver.find_element(By.XPATH, LOC.rd_schedule_xpath).click()
                    time.sleep(5)
                    if driver.find_element(By.XPATH, LOC.rd_schedule_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: By Schedule **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: By Schedule **********")
                        self.logger.info("********** Failed To Create A DR Policy "+name+" **********")
                        driver.find_element(By.ID, LOC.btn_cancel_xpath).click()
                        continue
                    dailyWeekly = Select(driver.find_element(By.XPATH, LOC.drp_dw_xpath))
                    hour = Select(driver.find_element(By.XPATH, LOC.drp_hr_xpath))
                    minute = Select(driver.find_element(By.XPATH, LOC.drp_min_xpath))

                    dailyWeekly.select_by_visible_text(text)
                    if text == "Weekly":
                        weekDay = Select(driver.find_element(By.XPATH, LOC.drp_wkDay_xpath))
                        weekDay.select_by_visible_text(day)
                    hour.select_by_index(hrs)
                    minute.select_by_index(mins)

                elif periodicity == "By Frequency":
                    mh = sheet.cell(row=r, column=11).value
                    due = sheet.cell(row=r, column=12).value
                    fhr = sheet.cell(row=r, column=13).value
                    fmn = sheet.cell(row=r, column=14).value
                    thr = sheet.cell(row=r, column=15).value
                    tmn = sheet.cell(row=r, column=16).value

                    driver.find_element(By.XPATH, LOC.rd_frequency_xpath).click()
                    time.sleep(3)
                    if driver.find_element(By.XPATH, LOC.rd_frequency_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: By Frequency **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: By Frequency **********")
                        self.logger.info("********** Failed To Create A DR Policy " + name + " **********")
                        driver.find_element(By.ID, LOC.btn_cancel_xpath).click()
                        continue
                    minHr = Select(driver.find_element(By.XPATH, LOC.drp_minHr_xpath))
                    minHr.select_by_visible_text(mh)
                    time.sleep(5)
                    if mh == "Minutes":
                        duration = Select(driver.find_element(By.ID, LOC.drp_durationMin_id))
                    else:
                        duration = Select(driver.find_element(By.ID, LOC.drp_durationHr_id))
                    if mh == "Hours":
                        duration.select_by_index(due-1)
                    elif mh == "Minutes" and due == 5:
                        duration.select_by_index(0)
                    elif mh == "Minutes" and due > 5:
                        duration.select_by_visible_text(str(due))
                    if fhr != "NA":
                        fromHr = Select(driver.find_element(By.ID, LOC.drp_fromHr_id))
                        fromHr.select_by_index(fhr+1)
                    if fmn != "NA":
                        fromMin = Select(driver.find_element(By.ID, LOC.drp_fromMin_id))
                        fromMin.select_by_index(fmn+1)
                    if thr != "NA":
                        toHr = Select(driver.find_element(By.ID, LOC.drp_toHr_id))
                        toHr.select_by_index(thr+1)
                    if tmn != "NA":
                        toMin = Select(driver.find_element(By.ID, LOC.drp_toMin_id))
                        toMin.select_by_index(tmn+1)

                elif periodicity == "Once":
                    driver.find_element(By.XPATH, LOC.rd_once_xpath).click()
                    time.sleep(3)
                    if driver.find_element(By.XPATH, LOC.rd_once_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: Once **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: Once **********")
                        self.logger.info("********** Failed To Create A DR Policy " + name + " **********")
                        driver.find_element(By.ID, LOC.btn_cancel_xpath).click()
                        continue
                    driver.find_element(By.XPATH, LOC.txt_start_xpath).click()
                    driver.find_element(By.XPATH, LOC.btn_clear_xpath).click()
                    driver.find_element(By.XPATH, LOC.txt_start_xpath).send_keys(startTime)
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, LOC.btn_startDate_xpath))
                    )
                    btn.click()

                elif periodicity == "Continuous":
                    driver.find_element(By.XPATH, LOC.rd_continuous_xpath).click()
                    time.sleep(3)
                    if driver.find_element(By.XPATH, LOC.rd_continuous_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: Continuous **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: Continuous **********")
                        self.logger.info("********** Failed To Create A DR Policy " + name + " **********")
                        btn = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, LOC.btn_cancel_xpath))
                        )
                        btn.click()
                        continue

                driver.find_element(By.XPATH, LOC.txt_email_xpath).send_keys(email)
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, LOC.btn_add_xpath))
                )
                btn.click()

                if note1:
                    driver.find_element(By.XPATH, LOC.chBox_failNote_xpath).click()
                    if driver.find_element(By.XPATH, LOC.chBox_failNote_xpath).is_selected():
                        self.logger.info("********** Email Notification For Fail Cases Only Was Selected **********")
                    else:
                        self.logger.info("********** Failed To Select Email Notification For Fail Cases Only **********")
                if note2:
                    driver.find_element(By.XPATH, LOC.chBox_completeNote_xpath).click()
                    if driver.find_element(By.XPATH, LOC.chBox_completeNote_xpath).is_selected():
                        self.logger.info("********** Email Notification When Complete Was Selected **********")
                    else:
                        self.logger.info("********** Failed To Select Email Notification When Complete **********")
                btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, LOC.btn_create_id))
                )
                btn.click()
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                )
                note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                self.logger.info("********** Create DR Policy Status For Policy : " + name + ",")
                self.logger.info(note + "\n")
                time.sleep(2)
                driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
            else:
                self.logger.info("********** Create New DR Policy Pop-up Banner Is Not Opened **********")

    def findPolicy(self, driver, policyName):
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(3)
        driver.find_element(By.LINK_TEXT, "Policies").click()
        time.sleep(5)
        totalDrPolicies = len(driver.find_elements(By.ID, LOC.txt_policyName_id))
        for i in range(1, totalDrPolicies + 1):
            if totalDrPolicies == 1:
                tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            else:
                tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/span').text
            if tmp == policyName:
                self.logger.info("********** DR Policy with name " + policyName + " Is Present At Sr. No. " + str(i) + " **********")
                return
        self.logger.info("********** DR Policy with name " + policyName + " Is Not Present **********")

    def addDRPolicyToWave(self, driver, path, start, end):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            policyName = sheet.cell(row=r, column=2).value
            startNow = sheet.cell(row=r, column=3).value
            time.sleep(5)
            co = CommonObjects(driver)
            val = co.findWave(waveName)
            if val == 2:
                continue
            if val == 1:
                driver.find_element(By.LINK_TEXT, waveName).click()
                time.sleep(5)
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
                else:
                    self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                    continue
            if val == 0:
                if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                    self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
            tmp = len(driver.find_elements(By.XPATH, LOC.txt_noPolicy_xpath))
            if tmp != 0:
                self.logger.info("********** Currently The Wave : " + waveName + ", Don't Have Any Policy **********")
            else:
                currentPolicy = driver.find_element(By.XPATH, LOC.txt_assPolicyName_xpath).text
                res = tuple(map(str, currentPolicy.split(' ')))
                self.logger.info("********** Currently The Wave : " + waveName + ", Has Policy : " + res[0] + " Assigned To It **********")
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, LOC.txt_drPolicy_xpath))
            )
            btn.click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.pop_policyAssignment_xpath)) != 0:
                self.logger.info("********** Policy Assignment Pop-up Banner Is Opened For Wave, " + str(waveName) + " **********")
                driver.find_element(By.XPATH, LOC.drp_selectDrPolicy_xpath).click()
                time.sleep(3)
                if len(driver.find_elements(By.CSS_SELECTOR, "li[aria-label="+policyName+"]")) == 0:
                    self.logger.info("********** There Was No Such Policy With Name, " + str(policyName) + " **********")
                    driver.find_element(By.ID, LOC.btn_cancel_id).click()
                else:
                    driver.find_element(By.CSS_SELECTOR, "li[aria-label="+policyName+"]").click()
                    time.sleep(5)
                    if startNow:
                        driver.find_element(By.ID, LOC.ch_startPolicyNow_id).click()
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, LOC.btn_assignPolicy_id))
                    )
                    btn.click()
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
                    )
                    note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
                    self.logger.info("********** Assign DR Policy " + policyName + " To Wave " + waveName + ", Status : ")
                    self.logger.info(note + "\n")
                    time.sleep(2)
                    driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, LOC.txt_assPolicyName_xpath))
                    )
                    currentPolicy = driver.find_element(By.XPATH, LOC.txt_assPolicyName_xpath).text
                    res = tuple(map(str, currentPolicy.split(' ')))
                    if policyName == res[0]:
                        self.logger.info("********** Policy : " + policyName + ", Added To The Wave : " + waveName + " **********")
                        self.checkDrPolicyState(driver, policyName)
                    else:
                        self.logger.info("********** Failed To Add Policy : " + policyName + ", To The Wave : " + waveName + " **********")
            else:
                self.logger.info("********** Policy Assignment Pop-up Banner Is Not Opened For Wave, " + str(waveName) + " **********")

    def checkDrPolicyState(self, driver, policyName):
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(3)
        driver.find_element(By.LINK_TEXT, "Policies").click()
        time.sleep(5)
        totalDrPolicies = len(driver.find_elements(By.ID, LOC.txt_policyName_id))
        if totalDrPolicies == 1:
            tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            if tmp != policyName:
                self.logger.info("********** There Is No Such Policy With Name: "+policyName+" **********")
            else:
                status = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span').text
                self.logger.info("********** Policy : " + policyName + ", Is In " + status + " State **********")
        else:
            for i in range(1, totalDrPolicies+1):
                tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/span').text
                if tmp == policyName:
                    status = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span/span').text
                    self.logger.info("********** Policy : " + policyName + ", Is In " + status + " State **********")
                    return
            self.logger.info("********** There Is No Such Policy With Name: "+policyName+" **********")

    def resumePolicy(self, driver, policyName):
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(3)
        driver.find_element(By.LINK_TEXT, "Policies").click()
        time.sleep(5)
        totalDrPolicies = len(driver.find_elements(By.ID, LOC.txt_policyName_id))
        if totalDrPolicies == 1:
            tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            if tmp == policyName:
                driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[1]').click()
            else:
                self.logger.info("********** There Is No Such Policy With Name: " + policyName + " **********")
                return
        else:
            for i in range(1, totalDrPolicies + 1):
                tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/span').text
                if tmp == policyName:
                    driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[9]/span/i[1]').click()
                    break
                if i == totalDrPolicies:
                    self.logger.info("********** There Is No Such Policy With Name: " + policyName + " **********")
                    return
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_resumePolicy_xpath)) != 0:
            self.logger.info("********** Pop-up Banner For Resume Policy Was Opened For Policy: " + policyName + " **********")
            driver.find_element(By.XPATH, LOC.btn_resume_xpath).click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
            )
            note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
            self.logger.info("********** Start or Restart Status For Policy : " + policyName + ",")
            self.logger.info(note + "\n")
            time.sleep(2)
            driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
        else:
            self.logger.info("********** Pop-up Banner For Resume Policy Was Not Opened For Policy: " + policyName + " **********")

    def verifyDRHostSyncStatus(self, driver, waveName):
        time.sleep(5)
        co = CommonObjects(driver)
        val = co.findDrWave(waveName)
        if val == 2:
            return
        if val == 1:
            driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                return
        if val == 0:
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
        WebDriverWait(driver, 18000).until(
            EC.element_to_be_clickable((By.ID, LOC.btn_start_id))
        )
        co = CommonObjects(driver)
        co.hostSyncState()

    def pauseDRPolicy(self, driver, policyName):
        dr_class = driver.find_element(By.XPATH, LOC.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(3)
        driver.find_element(By.LINK_TEXT, "Policies").click()
        time.sleep(5)
        totalDrPolicies = len(driver.find_elements(By.ID,  LOC.txt_policyName_id))
        count = 1
        time.sleep(5)
        if totalDrPolicies == 1:
            tmp = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            if tmp == policyName:
                waveState = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span').text
                self.logger.info("********** Policy : " + policyName + ", Is In " + waveState + " State **********")
                wavePause = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[1]')
                wavePause_class = wavePause.get_attribute("title")
                if wavePause_class == "Pause Policy":
                    self.logger.info("********** Pausing The DR Policy **********")
                    driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[1]').click()
                else:
                    self.logger.info("********** Can't Pause The Policy : " + policyName + ", Because It Is In " + waveState + " State **********")
        else:
            for i in range(1, totalDrPolicies + 1):
                tmp = driver.find_element(By.XPATH,'/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/span').text
                if tmp == policyName:
                    waveState = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span/span').text
                    self.logger.info("********** Policy : " + policyName + ", Is In " + waveState + " State **********")
                    wavePause = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[9]/span/i[1]')
                    wavePause_class = wavePause.get_attribute("title")
                    if wavePause_class == "Pause Policy":
                        self.logger.info("********** Pausing The DR Policy **********")
                        driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[9]/span/i[1]').click()
                    else:
                        self.logger.info("********** Can't Pause The Policy : " + policyName + ", Because It Is In " + waveState + " State **********")
                    break
                if i == totalDrPolicies:
                    self.logger.info("********** There Is No Such Policy With Name: " + policyName + " **********")
                    return
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, LOC.pop_successful_xpath))
        )
        note = driver.find_element(By.XPATH, LOC.pop_successful_xpath).text
        self.logger.info("********** Pause Status For Policy : " + policyName + ",")
        self.logger.info(note + "\n")
        time.sleep(2)
        driver.find_element(By.XPATH, LOC.pop_successful_xpath).click()
        if totalDrPolicies == 1:
            WebDriverWait(driver, 18000).until(
                EC.text_to_be_present_in_element((By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span'), "Paused")
            )
            time.sleep(5)
            waveState = driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span').text
        else:
            WebDriverWait(driver, 18000).until(
                EC.text_to_be_present_in_element((By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(count)+']/td[2]/span/span'), "Paused")
            )
            time.sleep(5)
            waveState = driver.find_element(By.XPATH,'/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(count) + ']/td[2]/span/span').text
        self.logger.info("********** Policy : " + policyName + ", Is In " + waveState + " State **********")

    def failoverHost(self, driver, waveName, testMode):
        co = CommonObjects(driver)
        val = co.findDrWave(waveName)
        if val == 2:
            return
        if val == 1:
            driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                return
        if val == 0:
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
        driver.find_element(By.ID, LOC.btn_failOver_id).click()
        time.sleep(5)
        if len(driver.find_elements(By.XPATH, LOC.pop_failOver_xpath)) != 0:
            self.logger.info("********** Pop-up Banner Of Failover Was Opened **********")
            if testMode:
                driver.find_element(By.ID, LOC.ch_testMode_id).click()
            driver.find_element(By.ID, LOC.btn_failoverYes_id).click()
            time.sleep(5)
            WebDriverWait(driver, 18000).until(
                EC.text_to_be_present_in_element((By.XPATH, LOC.txt_waveState_xpath), "Failed Over")
            )
            time.sleep(5)
            co = CommonObjects(driver)
            co.hostSyncState()
            time.sleep(5)
        else:
            self.logger.info("********** Pop-up Banner Of Failover Was Not Opened **********")

    def fallbackHost(self, driver, waveName):
        co = CommonObjects(driver)
        val = co.findDrWave(waveName)
        if val == 2:
            return
        if val == 1:
            driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                return
        if val == 0:
            if len(driver.find_elements(By.XPATH, LOC.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
        btn = WebDriverWait(driver, 18000).until(
            EC.element_to_be_clickable((By.ID, LOC.btn_fallBack_id))
        )
        btn.click()
        btn = WebDriverWait(driver, 18000).until(
            EC.element_to_be_clickable((By.ID, LOC.btn_fallBackYes_id))
        )
        btn.click()
        time.sleep(5)
        WebDriverWait(driver, 18000).until(
            EC.element_to_be_clickable((By.XPATH, LOC.btn_failOver_xpath))
        )
