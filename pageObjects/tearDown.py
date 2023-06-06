import time
import paramiko
import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from utilities.customLogger import LogGen
from pageObjects.commonObjects import CommonObjects


class TearDown:
    # Text
    txt_replication_xpath = '//*[@id="nav-panel"]/nav/ul/li[1]'
    txt_wave_xpath = '//*[@id="nav-panel"]/nav/ul/li[1]/ul/li[2]'
    txt_dr_xpath = '//*[@id="nav-panel"]/nav/ul/li[2]'
    txt_policies_xpath = '//*[@id="nav-panel"]/nav/ul/li[2]/ul/li[2]'
    txt_rWave_xpath = '//*[@id="nav-panel"]/nav/ul/li[1]/ul/li[2]'
    txt_drWave_xpath = '//*[@id="nav-panel"]/nav/ul/li[2]/ul/li[3]'
    txt_hostNImages_xpath = '//*[@id="nav-panel"]/nav/ul/li[4]'
    txt_hosts_xpath = '//*[@id="nav-panel"]/nav/ul/li[4]/ul/li[1]'
    txt_images_xpath = '//*[@id="nav-panel"]/nav/ul/li[4]/ul/li[2]'
    txt_select_xpath = '//*[@id="hosts_images_host_delete_host_modal"]/div/div/div/div[2]/div[3]/div[2]/select'
    txt_waveName_xpath = '//*[@id="content"]/article/div/div[1]/div/h3/strong'
    txt_totalHosts_xpath = '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr'
    txt_totalPolicies_xpath = '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr'
    txt_hostSearch_xpath = '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[1]/div[2]/input'
    txt_waveSearch_xpath = '//*[@id="waves_search_input"]'
    txt_policySearch_xpath = '//*[@id="policies_dr_policy_search_text"]'

    # Buttons
    btn_deleteWaveHost_id = "wave_detail_delete_item_delete_btn"
    btn_deleteWave_id = "waves_wave_delete_delete_btn"
    btn_deleteWaves_xpath = "//*[@id='waves_bulk_delete']/span/i"
    btn_bulkDelete_id = "waves_wave_bulk_delete_delete_btn"
    btn_deleteDRP_id = "policies_delete_dr_policy_delete_btn"
    btn_deleteHost_id = "hosts_images_hosts_delete_host_btn"
    btn_delete_xpath = '//*[@id="conf_vc_del_btn"]/span'

    # Check Boxs
    ch_deleteVM_xpath = '//*[@id="hosts_delete_host_delete_vm"]'

    # Radio Button
    rd_cloudUSer_xpath = '//*[@id="hosts_images_host_delete_host_modal"]/div/div/div/div[2]/div[3]/div[1]/div[1]/div/div/input'
    rd_vCenter_xpath = '//*[@id="hosts_images_host_delete_host_modal"]/div/div/div/div[2]/div[3]/div[1]/div[2]/div/div/input'
    rd_hypervisor_xpath = '//*[@id="hosts_images_host_delete_host_modal"]/div/div/div/div[2]/div[3]/div[1]/div[3]/div/div/input'

    # Pop-ups and Validations
    var_waveDetails_xpath = '//*[@id="rmm_lite_header"]/div/div[1]/div[2]'
    pop_successful_xpath = '/html/body/app-root/simple-notifications/div/simple-notification/div'
    pop_delete_xpath = '//*[@id="waves_wave_delete_title_text"]'
    pop_delMultWaves_id = "waves_wave_delete_delMutiple_waves"
    pop_deleteSuccessful_xpath = '/html/body/app-root/simple-notifications/div/simple-notification[2]/div'
    pop_deletePolicy_xpath = '//*[@id="policies_delete_dr_policy_delete_selected_dr_policy"]'
    pop_deleteHost_xpath = '//*[@id="hosts_images_host_delete_host_modal"]/div/div/div/div[1]/h4'

    logger = LogGen.loggen()

    def __init__(self, driver):
        self.driver = driver

    def tearDown(self, path, start, end, count):
        workBook = openpyxl.load_workbook(path)
        sheet = workBook.active
        rows = sheet.max_row
        if start == "NA":
            st = 3
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        time.sleep(5)
        cnt = 1
        for r in range(st, ed+1):
            operation = sheet.cell(row=r, column=1).value.lower()

            if operation.find("delete") != -1 and operation.find("ssh") != -1 and operation.find("entry") != -1:
                source = sheet.cell(row=r, column=5).value
                target = sheet.cell(row=r, column=6).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete SSH Host Sync Entry **********")
                self.deleteSR(source, target)
                self.logger.info("********** Successfully Executed TestCase: Delete SSH Host Sync Entry **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("host") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=2).value
                hostNames = sheet.cell(row=r, column=3).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Host In Wave **********")
                self.deleteHostInWave(waveName, hostNames)
                self.logger.info("********** Successfully Executed TestCase: Delete Host In Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("replication") != -1 and (operation.find("wave") != -1 or operation.find("waves") != -1):
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Wave **********")
                self.deleteRepWaves(waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("DR") != -1 and (operation.find("wave") != -1 or operation.find("waves") != -1):
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Wave **********")
                self.deleteDrWaves(waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("wave") != -1:
                waveName = sheet.cell(row=r, column=2).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Wave **********")
                self.deleteWave(waveName)
                self.logger.info("********** Successfully Executed TestCase: Delete Wave **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("dr") != -1 and (operation.find("policy") != -1 or operation.find("policies") != -1):
                policyName = sheet.cell(row=r, column=4).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete SSH Host Sync Entry **********")
                self.deleteDRPolicy(policyName)
                self.logger.info("********** Successfully Executed TestCase: Delete SSH Host Sync Entry **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("source") != -1 and (operation.find("host") != -1 or operation.find("hosts") != -1):
                hostNames = sheet.cell(row=r, column=3).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Source Host **********")
                self.deleteSourceHost(hostNames)
                self.logger.info("********** Successfully Executed TestCase: Delete Source Host **********\n \n")
                cnt += 1

            elif operation.find("delete") != -1 and operation.find("target") != -1 and (operation.find("host") != -1 or operation.find("hosts") != -1):
                hostNames = sheet.cell(row=r, column=3).value
                VM = sheet.cell(row=r, column=7).value
                environment = sheet.cell(row=r, column=8).value
                name = sheet.cell(row=r, column=9).value
                self.logger.info("********** Starting TestCase "+str(count)+"."+str(cnt)+": Delete Target Host **********")
                self.deleteTargetHost(hostNames, VM, environment, name)
                self.logger.info("********** Successfully Executed TestCase: Delete Target Host **********\n \n")
                cnt += 1

            else:
                self.logger.info("********** TestCase "+str(count)+"."+str(cnt)+": There Are Some Mistakes In The Operation Keywords '" + str(operation) + "' ... Please Check Once **********\n \n")
                cnt += 1

    @staticmethod
    def deleteSR(source, target):
        vm_ip = "172.29.30.127"
        vm_username = "root"
        vm_password = "rackware"

        ssh = paramiko.SSHClient()
        ssh.load_system_host_keys()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(vm_ip,
                    username=vm_username,
                    password=vm_password,
                    look_for_keys=False)
        ssh.exec_command("rw ic srd "+str(source)+" --target "+str(target))

    def deleteHostInWave(self, waveName, hostNames):
        co = CommonObjects(self.driver)
        val = co.findWave(waveName)
        if val == 2:
            return
        if val == 1:
            self.driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Opened Successfully **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
                return
        if val == 0:
            if len(self.driver.find_elements(By.XPATH, self.txt_waveName_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Was Already Open **********")
        totalHosts = len(self.driver.find_elements(By.ID, "wave_policy_wave_policy_wave_detail_elapsed_time_info"))
        res = tuple(map(str, hostNames.split(', ')))
        for hostName in res:
            if totalHosts == 1:
                tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[3]/span/span').text
                if tmp == hostName:
                    self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[8]/span/div/i[4]').click()
                    self.hostDeleteState(hostName)
                else:
                    self.logger.info("********** There Was No Host With Name : " + hostName + ", In Wave : " + waveName + " **********")
            else:
                count = 1
                for i in range(1, totalHosts + 1):
                    tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[3]/span/span').text
                    if tmp == hostName:
                        self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/rw-wave-detail/div[1]/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(count) + ']/td[8]/span/div/i[4]').click()
                        self.hostDeleteState(hostName)
                        break
                    count += 1
                if count == totalHosts + 1:
                    self.logger.info("********** There Was No Host With Name : " + hostName + ", In Wave : " + waveName + " **********")
        # if val == 1:
        #     self.driver.find_element(By.LINK_TEXT, "Replication").click()
        #     time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def hostDeleteState(self, hostName):
        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_deleteWaveHost_id).click()
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, self.pop_successful_xpath))
        )
        note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
        self.logger.info("********** Delete Status For Host : " + hostName + ",")
        self.logger.info(note + "\n")
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.pop_successful_xpath).click()

    def deleteWave(self, waveName):
        co = CommonObjects(self.driver)
        val = co.findWave(waveName)
        if val == 2:
            return
        if val == 0:
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(3)
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
            val = co.findWave(waveName)
            if val == 2:
                return
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.ID, "waves_"+waveName+"_wave_actions"))
        )
        self.driver.find_element(By.ID, "waves_"+waveName+"_wave_actions").click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.pop_delete_xpath)) != 0:
            self.logger.info("********** Delete Wave Pop-up Banner Was Opened For Wave, " + str(waveName) + " **********")
            self.driver.find_element(By.ID, self.btn_deleteWave_id).click()
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.pop_successful_xpath))
            )
            note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
            self.logger.info("********** Delete Wave Status of Wave : " + waveName + ",")
            self.logger.info(note + "\n")
            time.sleep(2)
            self.driver.find_element(By.XPATH, self.pop_successful_xpath).click()
        else:
            self.logger.info("********** Delete Wave Pop-up Banner Was Not Opened For Wave, " + str(waveName) + " **********")
        time.sleep(5)

    def deleteRepWaves(self, waveNames):
        time.sleep(5)
        res = tuple(map(str, waveNames.split(', ')))
        replication_class = self.driver.find_element(By.XPATH, self.txt_replication_xpath).get_attribute("class")
        if replication_class == "ng-star-inserted":
            self.driver.find_element(By.LINK_TEXT, "Replication").click()
            time.sleep(5)
        wave_class = self.driver.find_element(By.XPATH, self.txt_rWave_xpath).get_attribute("class")
        if wave_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
        flag = 0
        for waveName in res:
            self.driver.find_element(By.XPATH, self.txt_waveSearch_xpath).click()
            self.driver.find_element(By.XPATH, self.txt_waveSearch_xpath).clear()
            self.driver.find_element(By.XPATH, self.txt_waveSearch_xpath).send_keys(waveName)
            time.sleep(2)
            totalWaves = len(self.driver.find_elements(By.XPATH, self.txt_totalPolicies_xpath))
            if totalWaves == 0:
                self.logger.info("********** " + waveName + " Was Not Found In Replication Waves **********")
            elif totalWaves == 1:
                tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span/div/a').text
                if tmp == waveName:
                    self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** " + waveName + " Was Not Found In Replication Waves **********")
            else:
                for i in range(1, totalWaves + 1):
                    tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span/div/a').text
                    if tmp == waveName:
                        self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalWaves:
                        self.logger.info("********** " + waveName + " Was Not Found In Replication Waves **********")
        if flag > 0:
            ele1 = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.btn_deleteWaves_xpath))
            )
            ele1.click()
            time.sleep(5)
            if len(self.driver.find_elements(By.ID, self.pop_delMultWaves_id)) != 0:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Opened **********")
                self.driver.find_element(By.ID, self.btn_bulkDelete_id).click()
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, self.pop_deleteSuccessful_xpath))
                )
                note = self.driver.find_element(By.XPATH, self.pop_deleteSuccessful_xpath).text
                self.logger.info("********** Delete Wave Status,")
                self.logger.info(note + "\n")
                time.sleep(2)
                self.driver.find_element(By.XPATH, self.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Not Opened **********")

    def deleteDrWaves(self, waveNames):
        time.sleep(5)
        res = tuple(map(str, waveNames.split(', ')))
        dr_class = self.driver.find_element(By.XPATH, self.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            self.driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(5)
        wave_class = self.driver.find_element(By.XPATH, self.txt_drWave_xpath).get_attribute("class")
        if wave_class == "ng-star-inserted":
            self.driver.find_element(By.LINK_TEXT, "Waves").click()
            time.sleep(5)
        flag = 0
        for waveName in res:
            self.driver.find_element(By.XPATH, self.txt_waveSearch_xpath).click()
            self.driver.find_element(By.XPATH, self.txt_waveSearch_xpath).clear()
            self.driver.find_element(By.XPATH, self.txt_waveSearch_xpath).send_keys(waveName)
            time.sleep(2)
            totalWaves = len(self.driver.find_elements(By.XPATH, self.txt_totalPolicies_xpath))
            if totalWaves == 0:
                self.logger.info("********** "+waveName+" Was Not Found In DR Waves **********")
            elif totalWaves == 1:
                tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span/div/a').text
                if tmp == waveName:
                    self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** " + waveName + " Was Not Found In DR Waves **********")
            else:
                for i in range(1, totalWaves+1):
                    tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span/div/a').text
                    if tmp == waveName:
                        self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/app-waves/div/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalWaves:
                        self.logger.info("********** "+waveName+" Was Not Found In DR Waves **********")
        if flag > 0:
            ele1 = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.btn_deleteWaves_xpath))
            )
            ele1.click()
            time.sleep(5)
            if len(self.driver.find_elements(By.ID, self.pop_delMultWaves_id)) != 0:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Opened **********")
                self.driver.find_element(By.ID, self.btn_bulkDelete_id).click()
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, self.pop_deleteSuccessful_xpath))
                )
                note = self.driver.find_element(By.XPATH, self.pop_deleteSuccessful_xpath).text
                self.logger.info("********** Delete Wave Status,")
                self.logger.info(note + "\n")
                time.sleep(2)
                self.driver.find_element(By.XPATH, self.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Delete Multiple Waves Pop-up Banner Was Not Opened **********")

    def deleteDRPolicy(self, policyNames):
        time.sleep(5)
        dr_class = self.driver.find_element(By.XPATH, self.txt_dr_xpath).get_attribute("class")
        if dr_class == "ng-star-inserted":
            self.driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(2)
        policies_class = self.driver.find_element(By.XPATH, self.txt_policies_xpath).get_attribute("class")
        if policies_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Policies").click()
            time.sleep(5)
        res = tuple(map(str, policyNames.split(', ')))
        flag = 0
        for policyName in res:
            self.driver.find_element(By.XPATH, self.txt_policySearch_xpath).click()
            self.driver.find_element(By.XPATH, self.txt_policySearch_xpath).clear()
            self.driver.find_element(By.XPATH, self.txt_policySearch_xpath).send_keys(policyName)
            time.sleep(2)
            totalPolicies = len(self.driver.find_elements(By.XPATH, self.txt_totalPolicies_xpath))
            if totalPolicies == 0:
                self.logger.info("********** Policy : " + policyName + ", Was Not Present **********")
            if totalPolicies == 1:
                tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
                if tmp == policyName:
                    self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[3]').click()
                    flag += 1
                else:
                    self.logger.info("********** Policy : " + policyName + ", Was Not Present **********")
            else:
                for i in range(1, totalPolicies+1):
                    tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/span').text
                    if tmp == policyName:
                        self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[9]/span/i[3]').click()
                        flag += 1
                        break
                    if i == totalPolicies:
                        self.logger.info("********** Policy : " + policyName + ", Was Not Present **********")
        if flag > 0:
            time.sleep(2)
            if len(self.driver.find_elements(By.XPATH, self.pop_deletePolicy_xpath)) != 0:
                self.logger.info("********** Delete Policy Pop Up Banner Was Opened **********")
                self.driver.find_element(By.ID, self.btn_deleteDRP_id).click()
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, self.pop_successful_xpath))
                )
                note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                self.logger.info("********** Delete Policy Status, ")
                self.logger.info(note + "\n")
                time.sleep(2)
                self.driver.find_element(By.XPATH, self.pop_successful_xpath).click()
            else:
                self.logger.info("********** Delete Policy Pop Up Banner Was Not Opened **********")

    def deleteSourceHost(self, hostNames):
        time.sleep(5)
        hi_class = self.driver.find_element(By.XPATH, self.txt_hostNImages_xpath).get_attribute("class")
        if hi_class == "ng-star-inserted":
            self.driver.find_element(By.LINK_TEXT, "Hosts & Images").click()
            time.sleep(2)
        hosts_class = self.driver.find_element(By.XPATH, self.txt_hosts_xpath).get_attribute("class")
        if hosts_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Hosts").click()
            time.sleep(5)
        res = tuple(map(str, hostNames.split(', ')))
        flag = 0
        for hostName in res:
            time.sleep(2)
            self.driver.find_element(By.XPATH, self.txt_hostSearch_xpath).click()
            self.driver.find_element(By.XPATH, self.txt_hostSearch_xpath).clear()
            self.driver.find_element(By.XPATH, self.txt_hostSearch_xpath).send_keys(hostName)
            time.sleep(3)
            totalHosts = len(self.driver.find_elements(By.XPATH, self.txt_totalHosts_xpath))
            if totalHosts == 0:
                self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            elif totalHosts == 1:
                if len(self.driver.find_elements(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span')) != 0:
                    tp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span').text
                    if tp == "* No hosts to show.":
                        continue
                tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span[1]').text
                if tmp == hostName:
                    WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]'))
                    )
                    self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            else:
                for i in range(1, totalHosts + 1):
                    tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span[1]').text
                    if tmp == hostName:
                        WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]'))
                        )
                        self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalHosts:
                        self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
        if flag > 0:
            ele = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.btn_delete_xpath))
            )
            ele.click()
            time.sleep(3)
            if len(self.driver.find_elements(By.XPATH, self.pop_deleteHost_xpath)) != 0:
                self.logger.info("********** Delete Host Pop Up Banner Was Opened **********")
                self.driver.find_element(By.ID, self.btn_deleteHost_id).click()
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, self.pop_deleteSuccessful_xpath))
                )
                self.logger.info("********** Delete Status For Hosts,")
                for i in range(1, len(res)+1):
                    if len(self.driver.find_elements(By.XPATH, self.pop_deleteSuccessful_xpath)) != 0:
                        WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, self.pop_deleteSuccessful_xpath))
                        )
                        note = self.driver.find_element(By.XPATH, self.pop_deleteSuccessful_xpath).text
                        self.logger.info(note + "\n")
                        self.driver.find_element(By.XPATH, self.pop_deleteSuccessful_xpath).click()
            else:
                self.logger.info("********** Delete Host Pop Up Banner Was Not Opened **********")

    def deleteTargetHost(self, hostNames, VM, environment, name):
        time.sleep(5)
        hi_class = self.driver.find_element(By.XPATH, self.txt_hostNImages_xpath).get_attribute("class")
        if hi_class == "ng-star-inserted":
            self.driver.find_element(By.LINK_TEXT, "Hosts & Images").click()
            time.sleep(2)
        hosts_class = self.driver.find_element(By.XPATH, self.txt_hosts_xpath).get_attribute("class")
        if hosts_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Hosts").click()
            time.sleep(5)
        res = tuple(map(str, hostNames.split(', ')))
        flag = 0
        for hostName in res:
            self.driver.find_element(By.XPATH, self.txt_hostSearch_xpath).click()
            self.driver.find_element(By.XPATH, self.txt_hostSearch_xpath).clear()
            self.driver.find_element(By.XPATH, self.txt_hostSearch_xpath).send_keys(hostName)
            time.sleep(2)
            totalHosts = len(self.driver.find_elements(By.XPATH, self.txt_totalHosts_xpath))
            if totalHosts == 0:
                self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            if totalHosts == 1:
                if len(self.driver.find_elements(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span')) != 0:
                    tp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td/span').text
                    if tp == "* No hosts to show.":
                        continue
                tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[2]/span[1]').text
                if tmp == hostName:
                    self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr/td[1]/p-tablecheckbox/div/div[2]').click()
                    flag += 1
                else:
                    self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
            else:
                for i in range(1, totalHosts + 1):
                    tmp = self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[2]/span[1]').text
                    if tmp == hostName:
                        self.driver.find_element(By.XPATH, '//*[@id="content"]/div/article/div/div[2]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/p-tablecheckbox/div/div[2]').click()
                        flag += 1
                        break
                    if i == totalHosts:
                        self.logger.info("********** Host : " + hostName + ", Was Not Present **********")
        if flag > 0:
            ele = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.btn_delete_xpath))
            )
            ele.click()
            time.sleep(3)
            if len(self.driver.find_elements(By.XPATH, self.pop_deleteHost_xpath)) != 0:
                self.logger.info("********** Delete Host Pop Up Banner Was Opened **********")
                if VM:
                    self.driver.find_element(By.XPATH, self.ch_deleteVM_xpath).click()
                    time.sleep(3)
                    if environment == "clouduser":
                        self.driver.find_element(By.XPATH, self.rd_cloudUSer_xpath).click()
                    elif environment == "vCenter":
                        self.driver.find_element(By.XPATH, self.rd_vCenter_xpath).click()
                    elif environment == "Hypervisor":
                        self.driver.find_element(By.XPATH, self.rd_hypervisor_xpath).click()
                    time.sleep(3)
                    select = Select(self.driver.find_element(By.XPATH, self.txt_select_xpath))
                    select.select_by_visible_text(str(name))
                ele2 = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.ID, self.btn_deleteHost_id))
                )
                ele2.click()
                time.sleep(10)
            else:
                self.logger.info("********** Delete Host Pop Up Banner Was Not Opened **********")
