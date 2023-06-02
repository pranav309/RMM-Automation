import time
import openpyxl

from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from utilities.customLogger import LogGen
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.commonObjects import CommonObjects


class DRPolicy:
    btn_addNew_xpath = "//*[@id='policies_dr_policy_create_policy_btn']/span/i"

    txt_name_id = "policyName"
    rd_schedule_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[1]/div/div/input"
    rd_frequency_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[2]/div/div/input"
    rd_once_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[3]/div/div/input"
    rd_continuous_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[2]/div[1]/div[4]/div/div/input"
    txt_start_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[3]/div/p-calendar/span/input"
    txt_email_xpath = "//*[@id='email']"
    btn_add_xpath = "//*[@id='policies_create_dr_policy_sbf_addEmail_btn']/i"
    chBox_failNote_xpath = "//*[@id='notifyfailonly']"
    chBox_completeNote_xpath = "//*[@id='notifyonwavecomplete']"
    btn_create_id = "policies_create_dr_policy_sbf_create_btn"
    btn_clear_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[3]/div/p-calendar/span/div/div[3]/div/div[2]/button/span"
    btn_resume_xpath = "//*[@id='policies_dr_policy_resume_dr_policy']/div/div/div/form/div/div[3]/div/button[2]"
    txt_policyName_id = "policies_dr_policy_policyname"
    btn_cancel_xpath = "policies_create_dr_policy_sbf_cancel_btn"
    txt_waveState_xpath = '//*[@id="content"]/article/div/div[2]/div[1]/div[1]/div[2]'
    txt_waveName_xpath = '//*[@id="content"]/article/div/div[1]/div/h3/strong'

    # By Schedule
    drp_dw_xpath = "//*[@id='select_daily_weekly']"
    drp_hr_xpath = "//*[@id='policies_create_dr_policy_sbs_select_hour']"
    drp_min_xpath = "//*[@id='policies_create_dr_policy_sbs_select_minute']"
    drp_wkDay_xpath = "//*[@id='select_week']"

    # By Frequency
    drp_durationMin_id = "policies_create_dr_policy_sbf_select_minute"
    drp_durationHr_id = "policies_create_dr_policy_sbf_select_hour"
    drp_minHr_xpath = "//*[@id='select_hours_minutes']"
    drp_fromHr_id = "policies_create_dr_policy_sbf_exclude_from_hr"
    drp_fromMin_id = "policies_create_dr_policy_sbf_exclude_from_min"
    drp_toHr_id = "policies_create_dr_policy_sbf_exclude_to_hr"
    drp_toMin_id = "policies_create_dr_policy_sbf_exclude_to_min"

    # Once
    btn_startDate_xpath = "//*[@id='policies_dr_policy_create_dr_policy']/div/div/div/form/div[2]/div[3]/div/p-calendar/span/button/span[1]"

    # Add DR Policy To Wave
    txt_drPolicy_xpath = "//*[@id='content']/article/div/div[2]/div[2]/div[3]/div[2]/div"
    drp_selectDrPolicy_xpath = "//*[@id='wave_detail_wave_policy_dr_policy']/div/div[3]/span"
    ch_startPolicyNow_id = "wave_detail_wave_policy_start_now"
    btn_assignPolicy_id = "wave_detail_wave_policy_assign_policy_btn"
    btn_start_id = "wave_policy_wave_policy_wave_detail_start_replications"
    btn_pause_id = "wave_policy_wave_policy_wave_detail_pause_replications"
    txt_assPolicyName_xpath = '//*[@id="wave_policy_wave_policy_wave_detail_drPolicy"]/span'
    txt_noPolicy_xpath = '//*[@id="content"]/article/div/div[2]/div[2]/div[3]/div[2]/div'
    btn_cancel_id = "wave_detail_wave_policy_cancel_btn"

    # Failover
    txt_failoverCheck_xpath = '//*[@id="content"]/article/div/div[2]/div[2]/div[3]/div[2]/div'
    btn_failOver_xpath = "//*[@id='wave_policy_wave_policy_wave_detail_drPolicyFailover']/span/i"
    btn_failOver_id = "wave_policy_wave_policy_wave_detail_drPolicyFailover"
    ch_testMode_id = "wave_detail_failover_testmode"
    btn_failoverYes_id = "wave_detail_failover_yes_btn"
    btn_endDRFailoverOK_xpath = "//*[@id='policies_dr_policy_stop_failover_test']/div/div/div/div/div[3]/div/button[2]"
    txt_policyName_xpath = "//*[@id='wave_policy_wave_policy_wave_detail_drPolicy']/span"

    # Fallback
    btn_fallBack_id = "wave_policy_wave_policy_wave_detail_drPolicyFallback"
    btn_fallBackYes_id = "wave_detail_dr_fallback_yes_btn"

    # Validations
    val_waveOpen_xpath = '//*[@id="content"]/article/div/div[1]/div/h3/strong'

    # Pop-up Banners
    pop_failOver_xpath = '//*[@id="wave_policy_wave_policy_wave_detail_failover"]/div/div/div/div[1]/h4'
    pop_createDrPolicy_xpath = '//*[@id="policies_dr_policy_create_dr_policy"]/div/div/div/form/div[1]/h4'
    pop_policyAssignment_xpath = '//*[@id="wave_policy_wave_policy_wave_detail_wave_policy_info"]/div[1]/div/div/div[1]/h4'
    pop_resumePolicy_xpath = '//*[@id="policies_dr_policy_resume_dr_policy"]/div/div/div/form/div/div[1]/h4'
    pop_successful_xpath = '/html/body/app-root/simple-notifications/div/simple-notification/div'

    txt_policies_xpath = '//*[@id="nav-panel"]/nav/ul/li[2]/ul/li[2]'

    logger = LogGen.loggen()

    def __init__(self, driver):
        self.driver = driver

    def createDRPolicy(self, path, start, end):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        time.sleep(5)
        if(len(self.driver.find_elements(By.LINK_TEXT, "Policies"))) == 0:
            self.driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, "Policies").click()
        for r in range(st, ed+1):
            time.sleep(5)
            self.driver.find_element(By.XPATH, self.btn_addNew_xpath).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_createDrPolicy_xpath)) != 0:
                self.logger.info("********** Create New DR Policy Pop-up Banner Is Opened **********")
                name = sheet.cell(row=r, column=1).value
                startTime = sheet.cell(row=r, column=2).value
                email = sheet.cell(row=r, column=3).value
                note1 = sheet.cell(row=r, column=4).value
                note2 = sheet.cell(row=r, column=5).value
                periodicity = sheet.cell(row=r, column=6).value

                self.driver.find_element(By.ID, self.txt_name_id).send_keys(name)

                if periodicity == "By Schedule":
                    text = sheet.cell(row=r, column=7).value
                    hrs = sheet.cell(row=r, column=8).value
                    mins = sheet.cell(row=r, column=9).value
                    day = sheet.cell(row=r, column=10).value

                    self.driver.find_element(By.XPATH, self.rd_schedule_xpath).click()
                    time.sleep(5)
                    if self.driver.find_element(By.XPATH, self.rd_schedule_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: By Schedule **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: By Schedule **********")
                        self.logger.info("********** Failed To Create A DR Policy "+name+" **********")
                        self.driver.find_element(By.ID, self.btn_cancel_xpath).click()
                        continue
                    dailyWeekly = Select(self.driver.find_element(By.XPATH, self.drp_dw_xpath))
                    hour = Select(self.driver.find_element(By.XPATH, self.drp_hr_xpath))
                    minute = Select(self.driver.find_element(By.XPATH, self.drp_min_xpath))

                    dailyWeekly.select_by_visible_text(text)
                    if text == "Weekly":
                        weekDay = Select(self.driver.find_element(By.XPATH, self.drp_wkDay_xpath))
                        weekDay.select_by_visible_text(day)
                    hour.select_by_index(hrs)
                    minute.select_by_index(mins)

                elif periodicity == "By Frequency":
                    mh = sheet.cell(row=r, column=11).value
                    due = sheet.cell(row=r, column=12).value
                    fhr = sheet.cell(row=r, column=13).value
                    fmn = sheet.cell(row=r, column=14).value
                    thr = sheet.cell(row=r, column=15).value
                    tmn = sheet.cell(row=r, column=16).value

                    self.driver.find_element(By.XPATH, self.rd_frequency_xpath).click()
                    time.sleep(5)
                    if self.driver.find_element(By.XPATH, self.rd_frequency_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: By Frequency **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: By Frequency **********")
                        self.logger.info("********** Failed To Create A DR Policy " + name + " **********")
                        self.driver.find_element(By.ID, self.btn_cancel_xpath).click()
                        continue
                    minHr = Select(self.driver.find_element(By.XPATH, self.drp_minHr_xpath))
                    minHr.select_by_visible_text(mh)
                    time.sleep(5)
                    if mh == "Minutes":
                        duration = Select(self.driver.find_element(By.ID, self.drp_durationMin_id))
                    else:
                        duration = Select(self.driver.find_element(By.ID, self.drp_durationHr_id))
                    if mh == "Hours":
                        duration.select_by_index(due-1)
                        time.sleep(5)
                    elif mh == "Minutes" and due == 5:
                        duration.select_by_index(0)
                    elif mh == "Minutes" and due > 5:
                        duration.select_by_visible_text(str(due))
                    if fhr != "NA":
                        fromHr = Select(self.driver.find_element(By.ID, self.drp_fromHr_id))
                        fromHr.select_by_index(fhr+1)
                    if fmn != "NA":
                        fromMin = Select(self.driver.find_element(By.ID, self.drp_fromMin_id))
                        fromMin.select_by_index(fmn+1)
                    if thr != "NA":
                        toHr = Select(self.driver.find_element(By.ID, self.drp_toHr_id))
                        toHr.select_by_index(thr+1)
                    if tmn != "NA":
                        toMin = Select(self.driver.find_element(By.ID, self.drp_toMin_id))
                        toMin.select_by_index(tmn+1)

                elif periodicity == "Once":
                    self.driver.find_element(By.XPATH, self.rd_once_xpath).click()
                    time.sleep(5)
                    if self.driver.find_element(By.XPATH, self.rd_once_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: Once **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: Once **********")
                        self.logger.info("********** Failed To Create A DR Policy " + name + " **********")
                        self.driver.find_element(By.ID, self.btn_cancel_xpath).click()
                        continue
                    self.driver.find_element(By.XPATH, self.txt_start_xpath).click()
                    self.driver.find_element(By.XPATH, self.btn_clear_xpath).click()
                    self.driver.find_element(By.XPATH, self.txt_start_xpath).send_keys(startTime)
                    time.sleep(5)
                    self.driver.find_element(By.XPATH, self.btn_startDate_xpath).click()

                elif periodicity == "Continuous":
                    self.driver.find_element(By.XPATH, self.rd_continuous_xpath).click()
                    time.sleep(5)
                    if self.driver.find_element(By.XPATH, self.rd_continuous_xpath).is_selected():
                        self.logger.info("********** Selected Periodicity Type: Continuous **********")
                    else:
                        self.logger.info("********** Failed To Select Periodicity Type: Continuous **********")
                        self.logger.info("********** Failed To Create A DR Policy " + name + " **********")
                        self.driver.find_element(By.ID, self.btn_cancel_xpath).click()
                        continue

                self.driver.find_element(By.XPATH, self.txt_email_xpath).send_keys(email)
                self.driver.find_element(By.XPATH, self.btn_add_xpath).click()

                if note1:
                    self.driver.find_element(By.XPATH, self.chBox_failNote_xpath).click()
                    if self.driver.find_element(By.XPATH, self.chBox_failNote_xpath).is_selected():
                        self.logger.info("********** Email Notification For Fail Cases Only Was Selected **********")
                    else:
                        self.logger.info("********** Failed To Select Email Notification For Fail Cases Only **********")
                if note2:
                    self.driver.find_element(By.XPATH, self.chBox_completeNote_xpath).click()
                    if self.driver.find_element(By.XPATH, self.chBox_completeNote_xpath).is_selected():
                        self.logger.info("********** Email Notification When Complete Was Selected **********")
                    else:
                        self.logger.info("********** Failed To Select Email Notification When Complete **********")
                time.sleep(5)
                self.driver.find_element(By.ID, self.btn_create_id).click()
                time.sleep(5)
                self.findPolicy(name)
            else:
                self.logger.info("********** Create New DR Policy Pop-up Banner Is Not Opened **********")
        # time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Replication").click()
        # time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def findPolicy(self, policyName):
        time.sleep(5)
        if (len(self.driver.find_elements(By.LINK_TEXT, "Policies"))) == 0:
            self.driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(5)
        element = self.driver.find_element(By.XPATH, self.txt_policies_xpath)
        element_class = element.get_attribute("class")
        if element_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Policies").click()
            time.sleep(5)
        totalDrPolicies = len(self.driver.find_elements(By.ID, self.txt_policyName_id))
        for i in range(1, totalDrPolicies + 1):
            if totalDrPolicies == 1:
                tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            else:
                tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/span').text
            if tmp == policyName:
                self.logger.info("********** DR Policy with name " + policyName + " Is Present At Sr. No. " + str(i) + " **********")
                return
        self.logger.info("********** DR Policy with name " + policyName + " Is Not Present **********")

    def addDRPolicyToWave(self, path, start, end):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = sheet.max_row
        if start == "NA":
            st = 2
        else:
            st = start
        if end == "NA":
            ed = rows
        else:
            ed = end
        for r in range(st, ed+1):
            waveName = sheet.cell(row=r, column=1).value
            policyName = sheet.cell(row=r, column=2).value
            startNow = sheet.cell(row=r, column=3).value
            time.sleep(5)
            co = CommonObjects(self.driver)
            val = co.findWave(waveName)
            if val == 2:
                return
            self.driver.find_element(By.LINK_TEXT, waveName).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.val_waveOpen_xpath)) != 0:
                self.logger.info("********** Wave : " + waveName + " Opened Successfully **********")
                if val == 0:
                    self.logger.info("********** Currently The Wave : " + waveName + ", Don't Have Any Policy **********")
                else:
                    currentPolicy = self.driver.find_element(By.XPATH, self.txt_assPolicyName_xpath).text
                    res = tuple(map(str, currentPolicy.split(' ')))
                    self.logger.info("********** Currently The Wave : " + waveName + ", Has Policy : " + res[0] + " **********")
                self.driver.find_element(By.XPATH, self.txt_drPolicy_xpath).click()
                time.sleep(5)
                if len(self.driver.find_elements(By.XPATH, self.pop_policyAssignment_xpath)) != 0:
                    self.logger.info("********** Policy Assignment Pop-up Banner Is Opened For Wave, " + str(waveName) + " **********")
                    self.driver.find_element(By.XPATH, self.drp_selectDrPolicy_xpath).click()
                    time.sleep(5)
                    if len(self.driver.find_elements(By.CSS_SELECTOR, "li[aria-label="+policyName+"]")) == 0:
                        self.logger.info("********** There Was No Such Policy With Name, " + str(policyName) + " **********")
                        self.driver.find_element(By.ID, self.btn_cancel_id).click()
                    else:
                        self.driver.find_element(By.CSS_SELECTOR, "li[aria-label="+policyName+"]").click()
                        time.sleep(5)
                        if startNow:
                            self.driver.find_element(By.ID, self.ch_startPolicyNow_id).click()
                        time.sleep(3)
                        self.driver.find_element(By.ID, self.btn_assignPolicy_id).click()
                        time.sleep(5)
                        note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
                        self.logger.info("********** Assign DR Policy " + policyName + " To Wave " + waveName + ", Status : ")
                        self.logger.info(note + "\n")
                        currentPolicy = self.driver.find_element(By.XPATH, self.txt_assPolicyName_xpath).text
                        res = tuple(map(str, currentPolicy.split(' ')))
                        if policyName == res[0]:
                            self.logger.info("********** Policy : " + policyName + ", Added To The Wave : " + waveName + " **********")
                            time.sleep(5)
                            self.checkDrPolicyState(policyName)
                        else:
                            self.logger.info("********** Failed To Add Policy : " + policyName + ", To The Wave : " + waveName + ", Because The Wave Has " + currentPolicy + " Policy Assigned To It Already i.e. In Running State **********")
                else:
                    self.logger.info("********** Policy Assignment Pop-up Banner Is Not Opened For Wave, " + str(waveName) + " **********")
            else:
                self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
        # time.sleep(5)
        # if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
        #     self.driver.find_element(By.LINK_TEXT, "Replication").click()
        #     time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def checkDrPolicyState(self, policyName):
        time.sleep(5)
        if (len(self.driver.find_elements(By.LINK_TEXT, "Policies"))) == 0:
            self.driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(5)
        element = self.driver.find_element(By.XPATH, self.txt_policies_xpath)
        element_class = element.get_attribute("class")
        if element_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Policies").click()
            time.sleep(5)
        totalDrPolicies = len(self.driver.find_elements(By.ID, self.txt_policyName_id))
        if totalDrPolicies == 1:
            tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            if tmp != policyName:
                self.logger.info("********** There Is No Such Policy With Name: "+policyName+" **********")
            else:
                status = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span').text
                self.logger.info("********** Policy : " + policyName + ", Is In " + status + " State **********")
        else:
            for i in range(1, totalDrPolicies+1):
                tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[1]/span').text
                if tmp == policyName:
                    status = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span/span').text
                    self.logger.info("********** Policy : " + policyName + ", Is In " + status + " State **********")
                    return
            self.logger.info("********** There Is No Such Policy With Name: "+policyName+" **********")
        # if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
        #     self.driver.find_element(By.LINK_TEXT, "Replication").click()
        #     time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def resumePolicy(self, policyName):
        time.sleep(5)
        if (len(self.driver.find_elements(By.LINK_TEXT, "Policies"))) == 0:
            self.driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(5)
        element = self.driver.find_element(By.XPATH, self.txt_policies_xpath)
        element_class = element.get_attribute("class")
        if element_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Policies").click()
            time.sleep(5)
        totalDrPolicies = len(self.driver.find_elements(By.ID, self.txt_policyName_id))
        if totalDrPolicies == 1:
            tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            if tmp == policyName:
                self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[1]').click()
            else:
                self.logger.info("********** There Is No Such Policy With Name: " + policyName + " **********")
                return
        else:
            for i in range(1, totalDrPolicies + 1):
                tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/span').text
                if tmp == policyName:
                    self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[9]/span/i[1]').click()
                    break
                if i == totalDrPolicies:
                    self.logger.info("********** There Is No Such Policy With Name: " + policyName + " **********")
                    return

        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.pop_resumePolicy_xpath)) != 0:
            self.logger.info("********** Pop-up Banner For Resume Policy Was Opened For Policy: " + policyName + " **********")
            self.driver.find_element(By.XPATH, self.btn_resume_xpath).click()
            time.sleep(5)
            note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
            self.logger.info("********** Start or Restart Status For Policy : " + policyName + ",")
            self.logger.info(note + "\n")
        else:
            self.logger.info("********** Pop-up Banner For Resume Policy Was Not Opened For Policy: " + policyName + " **********")
        # if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
        #     self.driver.find_element(By.LINK_TEXT, "Replication").click()
        #     time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Waves").click()
        # time.sleep(5)

    def verifyDRHostSyncStatus(self, waveName):
        time.sleep(5)
        co = CommonObjects(self.driver)
        val = co.findDrWave(waveName)
        if val == 1:
            return
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, waveName).click()
        if len(self.driver.find_elements(By.XPATH, self.val_waveOpen_xpath)) != 0:
            self.logger.info("********** Wave : " + waveName + " Opened Successfully **********")
            # time.sleep(300)
            # self.logger.info("********** Pausing DR Policy **********")
            # self.pauseDRPolicy(str(policyName))
            # self.logger.info("********** Successfully Paused DR Policy **********")
            # if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) == 0:
            #     self.driver.find_element(By.LINK_TEXT, "DR").click()
            #     time.sleep(5)
            # self.driver.find_element(By.LINK_TEXT, "Waves").click()
            # time.sleep(5)
            # self.driver.find_element(By.LINK_TEXT, waveName).click()
            WebDriverWait(self.driver, 18000).until(
                EC.text_to_be_present_in_element((By.XPATH, self.txt_waveState_xpath), "Idle" or "	Paused" or "Stopped")
            )
            co = CommonObjects(self.driver)
            co.hostSyncState()
            time.sleep(5)
        else:
            self.logger.info("********** Failed To Open Wave : " + waveName + " **********")

    def failoverHost(self, waveName, testMode):
        co = CommonObjects(self.driver)
        val = co.findDrWave(waveName)
        if val == 1:
            return
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, waveName).click()
        time.sleep(5)
        if len(self.driver.find_elements(By.XPATH, self.val_waveOpen_xpath)) != 0:
            self.logger.info("********** Wave : " + waveName + " Opened Successfully **********")
            # tmp = self.driver.find_element(By.XPATH, self.txt_failoverCheck_xpath).text
            # if tmp == "No Policy":
            #     self.logger.info("********** You Can't Perform Failover Over "+waveName+" Wave. As This Wave Don't Have Any Policy Assigned To It. **********")
            #     return
            self.driver.find_element(By.ID, self.btn_failOver_id).click()
            time.sleep(5)
            if len(self.driver.find_elements(By.XPATH, self.pop_failOver_xpath)) != 0:
                self.logger.info("********** Pop-up Banner Of Failover Was Opened **********")
                if testMode:
                    self.driver.find_element(By.ID, self.ch_testMode_id).click()
                self.driver.find_element(By.ID, self.btn_failoverYes_id).click()
                WebDriverWait(self.driver, 18000).until(
                    EC.text_to_be_present_in_element((By.XPATH, self.txt_waveState_xpath), "Failed Over")
                )
                time.sleep(5)
                co = CommonObjects(self.driver)
                co.hostSyncState()
                time.sleep(5)
            else:
                self.logger.info("********** Pop-up Banner Of Failover Was Not Opened **********")
        else:
            self.logger.info("********** Failed To Open Wave : " + waveName + " **********")
        # if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
        #     self.driver.find_element(By.LINK_TEXT, "Replication").click()
        #     time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Waves").click()

    def fallbackHost(self, waveName):
        time.sleep(5)
        self.driver.find_element(By.LINK_TEXT, waveName).click()
        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_fallBack_id).click()
        time.sleep(5)
        self.driver.find_element(By.ID, self.btn_fallBackYes_id).click()
        time.sleep(5)
        WebDriverWait(self.driver, 18000).until(
            EC.element_to_be_clickable((By.XPATH, self.btn_failOver_xpath))
        )
        self.driver.find_element(By.LINK_TEXT, "Waves")

    def pauseDRPolicy(self, policyName):
        time.sleep(5)
        if (len(self.driver.find_elements(By.LINK_TEXT, "Policies"))) == 0:
            self.driver.find_element(By.LINK_TEXT, "DR").click()
            time.sleep(5)
        element = self.driver.find_element(By.XPATH, self.txt_policies_xpath)
        element_class = element.get_attribute("class")
        if element_class != "active":
            self.driver.find_element(By.LINK_TEXT, "Policies").click()
            time.sleep(5)
        totalDrPolicies = len(self.driver.find_elements(By.ID,  self.txt_policyName_id))
        count = 1
        time.sleep(5)
        if totalDrPolicies == 1:
            tmp = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[1]/span').text
            if tmp == policyName:
                waveState = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span').text
                self.logger.info("********** Policy : " + policyName + ", Is In " + waveState + " State **********")
                wavePause = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[1]')
                wavePause_class = wavePause.get_attribute("title")
                if wavePause_class == "Pause Policy":
                    self.logger.info("********** Pausing The DR Policy **********")
                    self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[9]/span/i[1]').click()
                else:
                    self.logger.info("********** Can't Pause The Policy : " + policyName + ", Because It Is In " + waveState + " State **********")
        else:
            for i in range(1, totalDrPolicies + 1):
                tmp = self.driver.find_element(By.XPATH,'/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[1]/span').text
                if tmp == policyName:
                    waveState = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(i) + ']/td[2]/span/span').text
                    self.logger.info("********** Policy : " + policyName + ", Is In " + waveState + " State **********")
                    wavePause = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[9]/span/i[1]')
                    wavePause_class = wavePause.get_attribute("title")
                    if wavePause_class == "Pause Policy":
                        self.logger.info("********** Pausing The DR Policy **********")
                        self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(i)+']/td[9]/span/i[1]').click()
                    else:
                        self.logger.info("********** Can't Pause The Policy : " + policyName + ", Because It Is In " + waveState + " State **********")
                    break
                count += 1
                if i == totalDrPolicies:
                    self.logger.info("********** There Is No Such Policy With Name: " + policyName + " **********")
                    return
        time.sleep(5)
        note = self.driver.find_element(By.XPATH, self.pop_successful_xpath).text
        self.logger.info("********** Pause Status For Policy : " + policyName + ",")
        self.logger.info(note + "\n")
        if totalDrPolicies == 1:
            WebDriverWait(self.driver, 18000).until(
                EC.text_to_be_present_in_element((By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span'), "Paused")
            )
            time.sleep(5)
            waveState = self.driver.find_element(By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr/td[2]/span/span').text
        else:
            WebDriverWait(self.driver, 18000).until(
                EC.text_to_be_present_in_element((By.XPATH, '/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr['+str(count)+']/td[2]/span/span'), "Paused")
            )
            time.sleep(5)
            waveState = self.driver.find_element(By.XPATH,'/html/body/app-root/app-main-layout/div/dr-policy/div/div/article/div/div[3]/p-table/div/div[2]/table/tbody/tr[' + str(count) + ']/td[2]/span/span').text
        self.logger.info("********** Policy : " + policyName + ", Is In " + waveState + " State **********")
        time.sleep(5)
        # if len(self.driver.find_elements(By.LINK_TEXT, "Policies")) != 0:
        #     self.driver.find_element(By.LINK_TEXT, "Replication").click()
        #     time.sleep(5)
        # self.driver.find_element(By.LINK_TEXT, "Waves").click()
        # time.sleep(5)
